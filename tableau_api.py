import requests
import logging
import io
import os
import zipfile
import re
import calendar
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple, Optional
import pandas as pd
import tableauserverclient as TSC
from tableauhyperapi import HyperProcess, Telemetry, Connection

class TableauAPI:
    def __init__(self, server_url: str, site_id: str):
        self.server_url = server_url.rstrip('/')
        self.site_id = site_id
        self.token = None
        self.site_id_response = None
        self.user_id = None
        self.api_version = "3.20"
    
    def authenticate_pat(self, token_name: str, token_key: str) -> Tuple[str, str, str]:
        """Authenticate with Tableau Server using Personal Access Token (PAT)"""
        url = f"{self.server_url}/api/{self.api_version}/auth/signin"
        
        payload = {
            "credentials": {
                "personalAccessTokenName": token_name,
                "personalAccessTokenSecret": token_key,
                "site": {"contentUrl": self.site_id}
            }
        }
        
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        try:
            logging.info(f"Attempting PAT authentication for token: {token_name} on site: {self.site_id}")
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            
            data = response.json()
            self.token = data['credentials']['token']
            self.site_id_response = data['credentials']['site']['id']
            self.user_id = data['credentials']['user']['id']
            
            logging.info(f"Successfully authenticated with PAT: {token_name}")
            return self.token, self.site_id_response, self.user_id
            
        except requests.exceptions.RequestException as e:
            logging.error(f"PAT authentication failed: {str(e)}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_data = e.response.json()
                    error_msg = error_data.get('error', {}).get('detail', str(e))
                except:
                    error_msg = f"HTTP {e.response.status_code}: {e.response.text}"
                raise Exception(f"Tableau PAT authentication failed: {error_msg}")
            else:
                raise Exception(f"Network error during PAT authentication: {str(e)}")

    def authenticate(self, username: str, password: str) -> Tuple[str, str, str]:
        """Authenticate with Tableau Server and return token, site_id, user_id"""
        url = f"{self.server_url}/api/{self.api_version}/auth/signin"
        
        payload = {
            "credentials": {
                "name": username,
                "password": password,
                "site": {"contentUrl": self.site_id}
            }
        }
        
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        try:
            logging.info(f"Attempting authentication for user: {username} on site: {self.site_id}")
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            
            data = response.json()
            self.token = data['credentials']['token']
            self.site_id_response = data['credentials']['site']['id']
            self.user_id = data['credentials']['user']['id']
            
            logging.info(f"Successfully authenticated user: {username}")
            return self.token, self.site_id_response, self.user_id
            
        except requests.exceptions.RequestException as e:
            logging.error(f"Authentication failed: {str(e)}")
            if hasattr(e, 'response') and e.response is not None:
                try:
                    error_data = e.response.json()
                    error_msg = error_data.get('error', {}).get('detail', str(e))
                except:
                    error_msg = f"HTTP {e.response.status_code}: {e.response.text}"
                raise Exception(f"Tableau authentication failed: {error_msg}")
            else:
                raise Exception(f"Network error during authentication: {str(e)}")
    def _get_headers(self) -> Dict[str, str]:
        """Get headers with authentication token"""
        if not self.token:
            raise Exception("Not authenticated. Please call authenticate() first.")
        
        return {
            "X-Tableau-Auth": self.token,
            "Accept": "application/json"
        }
    
    def get_projects(self) -> List[Dict]:
        """Get all projects accessible to the authenticated user"""
        if not self.site_id_response:
            raise Exception("No site ID available. Please authenticate first.")
            
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/projects"
        
        try:
            logging.info(f"Requesting projects from: {url}")
            response = requests.get(url, headers=self._get_headers())
            
            logging.info(f"Projects response status: {response.status_code}")
            if response.status_code != 200:
                logging.error(f"Projects response text: {response.text}")
            
            response.raise_for_status()
            
            data = response.json()
            logging.info(f"Projects response data: {data}")
            
            projects = data.get("projects", {}).get("project", [])
            
            # Ensure projects is always a list
            if isinstance(projects, dict):
                projects = [projects]
            
            logging.info(f"Retrieved {len(projects)} projects")
            for project in projects[:3]:  # Log first 3 projects for debugging
                logging.info(f"Project: {project.get('name', 'Unknown')} (ID: {project.get('id', 'Unknown')})")
            
            return projects
            
        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to get projects: {str(e)}")
            if hasattr(e, 'response') and e.response is not None:
                logging.error(f"Response status: {e.response.status_code}")
                logging.error(f"Response text: {e.response.text}")
            raise Exception(f"Failed to retrieve projects: {str(e)}")
    
    # Cache the full site workbook list per auth token. The project.id REST
    # filter is rejected (400) on current API versions, so we always end up
    # listing every workbook and filtering client-side — paging the whole site
    # on every project click is the main "slow dropdown" cost. Cache it.
    _all_workbooks_cache: dict = {}
    _ALL_WB_TTL = 300  # seconds

    def _get_all_site_workbooks(self) -> List[Dict]:
        key = (self.server_url, self.site_id_response, (self.token or '')[:12])
        import time as _t
        cached = TableauAPI._all_workbooks_cache.get(key)
        if cached and (_t.time() - cached['ts']) < self._ALL_WB_TTL:
            return cached['workbooks']

        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks"
        all_wbs, page = [], 1
        while True:
            resp = requests.get(url, headers=self._get_headers(),
                                params={"pageSize": "1000", "pageNumber": str(page)},
                                timeout=(5, 20))
            resp.raise_for_status()
            container = (resp.json() or {}).get("workbooks", {}) or {}
            batch = container.get("workbook", [])
            if isinstance(batch, dict):
                batch = [batch]
            all_wbs.extend(batch)
            # Stop when the page wasn't full (no more pages)
            if len(batch) < 1000:
                break
            page += 1
            if page > 20:  # hard safety cap
                break
        TableauAPI._all_workbooks_cache[key] = {'ts': _t.time(), 'workbooks': all_wbs}
        logging.info(f"Fetched + cached {len(all_wbs)} site workbooks")
        return all_wbs

    def list_workbooks_in_project_by_id(self, project_id: str) -> List[Dict]:
        """Get all workbooks in a specific project (from the cached site list)."""
        try:
            all_workbooks = self._get_all_site_workbooks()
            workbooks = [wb for wb in all_workbooks
                         if wb.get('project', {}).get('id') == project_id]
            logging.info(f"{len(workbooks)} workbooks in project '{project_id}' "
                         f"(of {len(all_workbooks)} site-wide)")
            return workbooks
        except Exception as e:
            logging.error(f"Failed to get workbooks for project ID '{project_id}': {str(e)}")
            raise Exception(f"Failed to retrieve workbooks: {str(e)}")

    def get_workbook_by_id(self, workbook_id: str) -> Optional[Dict]:
        """Get details for a specific workbook by its LUID"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}"
        try:
            response = requests.get(url, headers=self._get_headers())
            if response.status_code == 200:
                data = response.json()
                return data.get('workbook')
            return None
        except Exception as e:
            logging.error(f"Failed to get workbook {workbook_id}: {str(e)}")
            return None

    def resolve_view_id_in_workbook(self, workbook_id: str, view_name: str) -> Optional[str]:
        """Resolve a view name to an ID using site-wide lookup (useful for hidden views)."""
        try:
            wb_info = self.get_workbook_by_id(workbook_id)
            if not wb_info: return None
            workbook_name = wb_info.get('name')
            if not workbook_name: return None

            import urllib.parse
            encoded_workbook_name = urllib.parse.quote(workbook_name)
            url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/views?filter=workbookName:eq:{encoded_workbook_name}"
            headers = {"X-Tableau-Auth": self.token, "Accept": "application/json"}
            res = requests.get(url, headers=headers)
            
            if res.status_code == 200:
                data = res.json()
                site_views = data.get('views', {}).get('view', [])
                workbook_views = [sv for sv in site_views if sv.get('workbook', {}).get('id') == workbook_id]
                
                if not workbook_views:
                    logging.warning(f"No views found for workbook ID {workbook_id} in matching views.")
                    return None

                def normalize(name):
                    return ''.join(name.split()).lower()

                normalized_target = normalize(view_name)
                
                # 1. Exact match (case insensitive)
                match = next((sv for sv in workbook_views if sv.get('name', '').lower() == view_name.lower()), None)
                
                # 2. Normalized match (strip all spaces)
                if not match:
                    match = next((sv for sv in workbook_views if normalize(sv.get('name', '')) == normalized_target), None)

                # 3. Match without "Dashboard > " prefix
                if not match and " > " in view_name:
                    short_name = view_name.split(" > ")[-1]
                    normalized_short = normalize(short_name)
                    match = next((sv for sv in workbook_views if normalize(sv.get('name', '')) == normalized_short), None)
                
                # 4. Match if site_view name is contained in discovery name
                if not match:
                    match = next((sv for sv in workbook_views if normalize(sv.get('name', '')) in normalized_target), None)
                    
                if match:
                    logging.info(f"✓ Resolved '{view_name}' to ID {match.get('id')} in workbook {workbook_id} via match '{match.get('name')}'")
                    return match.get('id')
        except Exception as e:
            logging.warning(f"Failed to resolve ID for '{view_name}': {e}")
        return None

    def list_workbooks_in_project(self, project_name: str) -> List[Dict]:
        """Get all workbooks in a specific project (by name)"""
        try:
            # First get the project ID
            projects = self.get_projects()
            project_id = None
            
            for project in projects:
                if project['name'].lower() == project_name.lower():
                    project_id = project['id']
                    break
            
            if not project_id:
                logging.warning(f"Project '{project_name}' not found")
                return []
            
            return self.list_workbooks_in_project_by_id(project_id)
            
        except Exception as e:
            logging.error(f"Failed to get workbooks for project '{project_name}': {str(e)}")
            raise Exception(f"Failed to retrieve workbooks: {str(e)}")
    
    def get_views_in_workbook(self, workbook_id: str) -> List[Dict]:
        """Get all views (dashboards) in a specific workbook"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/views"

        try:
            response = requests.get(url, headers=self._get_headers(), timeout=(5, 15))
            response.raise_for_status()

            data = response.json()
            views = data.get("views", {}).get("view", [])
            
            # Ensure views is always a list
            if isinstance(views, dict):
                views = [views]
            
            logging.info(f"Retrieved {len(views)} views for workbook {workbook_id}")
            return views
            
        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to get views for workbook {workbook_id}: {str(e)}")
            raise Exception(f"Failed to retrieve dashboards: {str(e)}")
    
    def get_view_filters(self, view_id: str) -> List[Dict]:
        """Get all filters applied to a view"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/views/{view_id}/filters"
        
        try:
            logging.info(f"Fetching filters for view {view_id}")
            response = requests.get(url, headers=self._get_headers())
            
            # handle common error codes
            if response.status_code == 404 or response.status_code == 405:
                logging.warning(f"Filters endpoint not supported or view not found: {view_id}")
                return []
                
            response.raise_for_status()
            
            data = response.json()
            logging.info(f"DEBUG: Raw filter response for view {view_id}: {data}")
            
            # Try multiple common paths for filters in the JSON response
            filters = []
            if data and 'filters' in data:
                filter_data = data['filters']
                if isinstance(filter_data, dict) and 'filter' in filter_data:
                    filters = filter_data['filter']
                elif isinstance(filter_data, list):
                    filters = filter_data
            
            if isinstance(filters, dict):
                filters = [filters]
                
            if not isinstance(filters, list):
                filters = []
                
            # Filter out and format
            formatted_filters = []
            for f in filters:
                name = f.get('fieldName') or f.get('name')
                if name:
                    formatted_filters.append({
                        'name': name,
                        'type': 'filter',
                        'values': f.get('values', [])
                    })

            logging.info(f"Retrieved {len(formatted_filters)} formatted filters for view {view_id}")
            return formatted_filters
            
        except Exception as e:
            logging.error(f"Failed to get filters for view {view_id}: {str(e)}")
            return []

    def get_workbook_datasources(self, workbook_id: str) -> List[Dict]:
        """Get data sources/connections for a workbook with updated timestamps"""
        # First get the connections list
        conn_url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/connections"
        
        try:
            logging.info(f"Fetching data source connections for workbook {workbook_id}")
            response = requests.get(conn_url, headers=self._get_headers())
            
            if response.status_code == 404 or response.status_code == 405:
                logging.debug(f"Connections endpoint not supported for workbook {workbook_id}")
                return []
                
            response.raise_for_status()
            data = response.json()
            logging.info(f"DEBUG: Raw connections response for workbook {workbook_id}: {data}")
            
            connections = []
            if data and 'connections' in data:
                conn_data = data['connections']
                if isinstance(conn_data, dict) and 'connection' in conn_data:
                    connections = conn_data['connection']
                elif isinstance(conn_data, list):
                    connections = conn_data
                    
            if isinstance(connections, dict):
                connections = [connections]
                
            if not isinstance(connections, list):
                connections = []
            
            # Get datasource details for connections that have datasource references
            datasources = []
            seen_datasource_ids = set()
            
            for conn in connections:
                # Check if connection has a datasource reference
                ds = conn.get('datasource', {})
                ds_id = ds.get('id')
                
                if ds_id and ds_id not in seen_datasource_ids:
                    seen_datasource_ids.add(ds_id)
                    # Fetch detailed datasource info
                    ds_info = self._get_datasource_details(ds_id)
                    if ds_info:
                        datasources.append(ds_info)
            
            # If no datasources found through connections, try fetching all site datasources 
            # and filtering (fallback approach)
            if not datasources:
                logging.info(f"No datasource references in connections, attempting alternative approach")
                # The connection type itself can give hints about the data
                for conn in connections:
                    conn_type = conn.get('type', 'unknown')
                    datasources.append({
                        'name': conn.get('serverAddress') or conn_type,
                        'type': conn_type,
                        'hasExtracts': conn_type == 'dataengine',
                        'updatedAt': None,
                        'connection_based': True
                    })
                    
            logging.info(f"Retrieved {len(datasources)} data sources for workbook {workbook_id}")
            return datasources
            
        except Exception as e:
            logging.error(f"Failed to get data sources for workbook {workbook_id}: {str(e)}")
            return []
    
    def _get_datasource_details(self, datasource_id: str) -> Optional[Dict]:
        """Get detailed information about a specific data source"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/datasources/{datasource_id}"
        
        try:
            response = requests.get(url, headers=self._get_headers())
            
            if response.status_code != 200:
                logging.debug(f"Could not fetch datasource {datasource_id}: {response.status_code}")
                return None
                
            data = response.json()
            ds = data.get('datasource', {})
            ds_name = ds.get('name')
            
            # Try to get the actual extract refresh time from Metadata API
            extract_refresh_time = self._get_extract_refresh_time(ds_name)
            
            return {
                'id': ds.get('id'),
                'name': ds_name,
                'type': ds.get('type'),
                'hasExtracts': ds.get('hasExtracts', False),
                'updatedAt': extract_refresh_time or ds.get('updatedAt'),  # Prefer extract refresh time
                'createdAt': ds.get('createdAt'),
                'contentUrl': ds.get('contentUrl')
            }
        except Exception as e:
            logging.error(f"Error fetching datasource details {datasource_id}: {str(e)}")
            return None

    def _get_extract_refresh_time(self, datasource_name: str) -> Optional[str]:
        """Get actual extract refresh time from Tableau Metadata API (GraphQL)"""
        if not datasource_name:
            return None
            
        # Metadata API endpoint
        metadata_url = f"{self.server_url}/api/metadata/graphql"
        
        # GraphQL query to get the extractLastUpdateTime for embedded datasources
        query = """
        query GetExtractRefreshTime($name: String!) {
            embeddedDatasources(filter: {name: $name}) {
                name
                hasExtracts
                extractLastRefreshTime
                extractLastUpdateTime
            }
            publishedDatasources(filter: {name: $name}) {
                name
                hasExtracts
                extractLastRefreshTime
                extractLastUpdateTime
            }
        }
        """
        
        try:
            response = requests.post(
                metadata_url,
                headers=self._get_headers(),
                json={
                    "query": query,
                    "variables": {"name": datasource_name}
                }
            )
            
            if response.status_code != 200:
                logging.debug(f"Metadata API request failed: {response.status_code}")
                return None
            
            data = response.json()
            logging.debug(f"Metadata API response for {datasource_name}: {data}")
            
            # Check embedded datasources first (most common for workbook datasources)
            embedded = data.get('data', {}).get('embeddedDatasources', [])
            for ds in embedded:
                if ds.get('name') == datasource_name:
                    # Prefer extractLastUpdateTime (includes both refresh and incremental updates)
                    return ds.get('extractLastUpdateTime') or ds.get('extractLastRefreshTime')
            
            # Check published datasources as fallback
            published = data.get('data', {}).get('publishedDatasources', [])
            for ds in published:
                if ds.get('name') == datasource_name:
                    return ds.get('extractLastUpdateTime') or ds.get('extractLastRefreshTime')
            
            return None
            
        except Exception as e:
            logging.debug(f"Error fetching extract refresh time from Metadata API: {str(e)}")
            return None

    def get_workbook_upstream_tables(self, workbook_id: str) -> Optional[str]:
        """Get the upstream databases, tables, and Custom SQL queries for a workbook"""
        metadata_url = f"{self.server_url}/api/metadata/graphql"

        query = """
        query GetWorkbookUpstreamTables($luid: String!) {
          workbooks(filter: { luid: $luid }) {
            name
            upstreamDatabases {
              name
              connectionType
              tables {
                name
                schema
              }
              referencedByQueries {
                id
                name
                query
              }
            }
          }
        }
        """

        try:
            response = requests.post(
                metadata_url,
                headers=self._get_headers(),
                json={
                    "query": query,
                    "variables": {"luid": workbook_id}
                }
            )

            if response.status_code != 200:
                logging.error(f"Metadata API request failed: {response.status_code} - {response.text}")
                return None

            data = response.json()
            logging.info(f"Metadata API Response: {data}")
            
            if 'errors' in data:
                error_msgs = ", ".join([err.get('message', 'Unknown Error') for err in data.get('errors', [])])
                logging.error(f"GraphQL errors: {error_msgs}")
                return f"Error retrieving schema: {error_msgs}"

            workbooks = data.get('data', {}).get('workbooks', [])
            
            if not workbooks:
                return "No upstream database information found for this workbook."

            wb = workbooks[0]
            if not wb:
                return "No upstream database information found for this workbook."

            output_lines = [f"Data Sources for Workbook: {wb.get('name', 'Unknown')}"]
            
            databases = wb.get('upstreamDatabases', [])
            if not databases:
                return "No upstream database information found for this workbook."

            custom_sqls_found = False

            for db in databases:
                if not db:
                    continue
                
                db_name = db.get('name', 'Unknown')
                conn_type = db.get('connectionType', 'Unknown')
                
                tables = db.get('tables', [])
                custom_queries = db.get('referencedByQueries', [])
                
                if not tables and not custom_queries:
                    continue
                    
                output_lines.append(f"\n[Database: {db_name} | Type: {conn_type}]")
                
                for table in tables:
                    if not table:
                        continue
                    table_name = table.get('name', 'Unknown')
                    schema = table.get('schema', '')
                    full_name = f"{schema}.{table_name}" if schema else table_name
                    output_lines.append(f"Standard Table: {full_name}")
                    
                for query_obj in custom_queries:
                    if not query_obj:
                        continue
                    custom_sqls_found = True
                    query_name = query_obj.get('name', 'Unknown Query Name')
                    query_text = query_obj.get('query', '')
                    output_lines.append(f"\nCustom SQL Query ({query_name}):")
                    if query_text:
                        # Indent the query for better readability
                        indented_query = "\n".join([f"  {line}" for line in query_text.split('\n')])
                        output_lines.append(indented_query)
                    else:
                        output_lines.append("  <Query text missing or empty>")

            # FALLBACK: If Tableau Metadata API (GraphQL) failed to associate Custom SQL
            # We download the workbook XML structure directly using the REST API and parse it
            if not custom_sqls_found:
                try:
                    import zipfile
                    import io
                    import xml.etree.ElementTree as ET
                    
                    url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/content"
                    content_response = requests.get(url, headers=self._get_headers())
                    
                    if content_response.status_code == 200:
                        content = content_response.content
                        xml_content = None
                        
                        try:
                            with zipfile.ZipFile(io.BytesIO(content)) as z:
                                twb_files = [f for f in z.namelist() if f.endswith('.twb')]
                                if twb_files:
                                    xml_content = z.read(twb_files[0])
                        except zipfile.BadZipFile:
                            # It's not a zip (.twbx), it must be a raw .twb XML file
                            xml_content = content
                            
                        if xml_content:
                            root = ET.fromstring(xml_content)
                            scraped_sqls = []
                            for relation in root.findall('.//relation'):
                                if relation.get('type') == 'text' and relation.text:
                                    scraped_sqls.append(relation.text.strip())
                            
                            if scraped_sqls:
                                output_lines.append("\n[Direct XML Custom SQL Extraction]")
                                for i, sql in enumerate(scraped_sqls):
                                    output_lines.append(f"\nCustom SQL Query #{i+1}:")
                                    indented_query = "\n".join([f"  {line}" for line in sql.split('\n')])
                                    output_lines.append(indented_query)
                except Exception as xml_err:
                    logging.error(f"Failed to parse XML for Custom SQL: {xml_err}")

            if len(output_lines) == 1:
                output_lines.append("No specific tables or Custom SQL identified in upstream databases.")
            
            return "\n".join(output_lines)

        except Exception as e:
            logging.error(f"Error fetching upstream tables from Metadata API: {str(e)}")
            import traceback
            logging.error(traceback.format_exc())
            return f"Error retrieving schema: {str(e)}"

    def get_workbook_parameters(self, workbook_id: str) -> List[Dict]:
        """Get parameters for a workbook (often used as filters)"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/parameters"
        
        try:
            logging.info(f"Fetching parameters for workbook {workbook_id}")
            response = requests.get(url, headers=self._get_headers())
            
            if response.status_code == 404 or response.status_code == 405:
                logging.debug(f"Parameters endpoint not supported for workbook {workbook_id}")
                return []
                
            response.raise_for_status()
            data = response.json()
            logging.info(f"DEBUG: Raw parameters response for workbook {workbook_id}: {data}")
            
            parameters = []
            if data and 'parameters' in data:
                param_data = data['parameters']
                if isinstance(param_data, dict) and 'parameter' in param_data:
                    parameters = param_data['parameter']
                elif isinstance(param_data, list):
                    parameters = param_data
                    
            if isinstance(parameters, dict):
                parameters = [parameters]
                
            if not isinstance(parameters, list):
                parameters = []
                
            formatted_params = []
            for p in parameters:
                name = p.get('name')
                if name:
                    formatted_params.append({
                        'name': name,
                        'type': 'parameter',
                        'dataType': p.get('dataType')
                    })

            logging.info(f"Retrieved {len(formatted_params)} formatted parameters for workbook {workbook_id}")
            return formatted_params
            
        except Exception as e:
            logging.error(f"Failed to get parameters for workbook {workbook_id}: {str(e)}")
            return []

    def get_workbook_worksheets_graphql(self, workbook_id: str) -> List[Dict]:
        """Get ALL worksheets in a workbook (including hidden ones) via GraphQL Metadata API."""
        metadata_url = f"{self.server_url}/api/metadata/graphql"
        
        query = """
        query GetWorksheets($luid: String!) {
          workbooks(filter: { luid: $luid }) {
            name
            luid
            sheets {
              name
              luid
              __typename
            }
            dashboards {
              name
              luid
              path
              sheets {
                name
                luid
              }
            }
          }
        }
        """
        
        try:
            logging.info(f"Discovering worksheets for workbook {workbook_id} via Metadata API...")
            # Fail fast: on sites where the Metadata API is slow or unindexed it
            # returns empty anyway, so a hung request would just stall the
            # dashboard dropdown. (5s connect, 12s read) → fall back to REST.
            response = requests.post(
                metadata_url,
                headers=self._get_headers(),
                json={
                    "query": query,
                    "variables": {"luid": workbook_id}
                },
                timeout=(5, 12)
            )

            if response.status_code != 200:
                logging.warning(f"Metadata API request failed: {response.status_code}")
                return []
                
            data = response.json()
            if 'errors' in data:
                logging.error(f"Metadata API returned errors: {data['errors']}")
                
            # Safely navigate the JSON structure
            graphql_data = data.get('data')
            if not graphql_data:
                logging.warning("Metadata API returned no data object")
                return []
                
            workbooks = graphql_data.get('workbooks', [])
            if not workbooks or not workbooks[0]:
                logging.warning(f"No workbook found in GraphQL for LUID: {workbook_id}")
                return []
                
            wb_data = workbooks[0]
            if not wb_data:
                logging.warning("First workbook in GraphQL list is null")
                return []
                
            sheets = wb_data.get('sheets', [])
            dashboards = wb_data.get('dashboards', [])
            stories = wb_data.get('stories', [])
            
            # Additional safety for lists
            if not isinstance(sheets, list): sheets = []
            if not isinstance(dashboards, list): dashboards = []
            if not isinstance(stories, list): stories = []
            
            logging.info(f"GraphQL discovered {len(sheets)} sheets, {len(dashboards)} dashboards, and {len(stories)} stories in workbook '{wb_data.get('name')}'")
            
                # Format to match standard view dict as much as possible
            formatted_views = []
            seen_ws_ids = set()
            seen_ws_names = set() # normalization for name-based dedup
            
            def normalize_name(name):
                if not name: return ""
                return "".join(name.lower().split())
            
            # 1. Add Dashboards as primary entries first
            for db in dashboards:
                if not db: continue
                db_id = db.get('luid')
                formatted_views.append({
                    'id': db_id,
                    'name': (db.get('name') or 'Dashboard').strip(),
                    'sheetType': 'dashboard'
                })
            
            # 2. Add Worksheets from dashboards (prefixed)
            for db in dashboards:
                if not db: continue
                db_name = (db.get('name') or 'Dashboard').strip()
                db_ws = db.get('sheets', []) or db.get('worksheets', [])
                if not isinstance(db_ws, list): continue
                for ws in db_ws:
                    if not ws: continue
                    ws_id = ws.get('luid') or None
                    ws_name = (ws.get('name') or 'Sheet').strip()
                    if ws_name:
                        formatted_views.append({
                            'id': ws_id,
                            'name': f"{db_name} > {ws_name}",
                            'sheetType': 'worksheet',
                            'parent_dashboard_id': db_id,
                            'parent_dashboard_name': db_name
                        })
                        if ws_id: seen_ws_ids.add(ws_id)
                        seen_ws_names.add(normalize_name(ws_name))
            
            # 3. Add standalone worksheets ONLY if they weren't seen in a dashboard
            for s in sheets:
                if not s: continue
                s_id = s.get('luid') or None
                s_name = (s.get('name') or 'Sheet').strip()
                s_type = s.get('__typename')
                
                # If it's a dashboard type, and not already in list, add it
                if s_type == 'Dashboard':
                    if not any(v['id'] == s_id for v in formatted_views if v['id']):
                         formatted_views.append({
                            'id': s_id,
                            'name': s_name,
                            'sheetType': 'dashboard'
                        })
                    continue

                # If it's a worksheet, only add if not already seen in a dashboard
                if s_id and s_id in seen_ws_ids:
                    continue
                if normalize_name(s_name) in seen_ws_names:
                    continue
                
                if s_name and s_name not in [v['name'] for v in formatted_views]:
                    formatted_views.append({
                        'id': s_id,
                        'name': s_name,
                        'sheetType': 'worksheet'
                    })
                    if s_id: seen_ws_ids.add(s_id)
            
            # 4. Add worksheets from stories (with double nesting check)
            for story in stories:
                if not story: continue
                story_name = story.get('name', 'Story')
                story_points = story.get('storyPoints', [])
                if not isinstance(story_points, list): continue
                
                for sp in story_points:
                    if not sp: continue
                    nested_dbs = sp.get('nestedDashboards', [])
                    if isinstance(nested_dbs, list):
                        for ndb in nested_dbs:
                            if not ndb: continue
                            ndb_name = ndb.get('name', 'Dashboard')
                            ndb_ws = ndb.get('worksheets', [])
                            if isinstance(ndb_ws, list):
                                for nws in ndb_ws:
                                    nws_id = nws.get('luid')
                                    if nws_id and nws_id not in seen_ws_ids:
                                        formatted_views.append({
                                            'id': nws_id,
                                            'name': f"{story_name} > {ndb_name} > {nws.get('name', 'Sheet')}",
                                            'sheetType': 'worksheet',
                                            'parent_dashboard_id': ndb.get('luid'),
                                            'parent_dashboard_name': ndb_name,
                                            'parent_story_id': story.get('luid'),
                                            'parent_story_name': story_name
                                        })
                                        seen_ws_ids.add(nws_id)
                    
                    # Handle direct nested worksheets in story points
                    nested_ws = sp.get('nestedWorksheets', [])
                    if isinstance(nested_ws, list):
                        for nws in nested_ws:
                            if nws and nws.get('luid') and nws.get('luid') not in seen_ws_ids:
                                formatted_views.append({
                                    'id': nws.get('luid'),
                                    'name': f"{story_name} > {nws.get('name', 'Sheet')}",
                                    'sheetType': 'worksheet'
                                })
                                seen_ws_ids.add(nws.get('luid'))
            
            return formatted_views
            
        except Exception as e:
            logging.error(f"GraphQL worksheet discovery failed: {str(e)}")
            return []

    def get_views_via_twb(self, workbook_id: str) -> List[Dict]:
        """Nuclear Option: Download .twb content and parse ALL worksheets/dashboards from XML.
        This follows the strategy in the user-provided blog post.
        """
        import xml.etree.ElementTree as ET
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/content"
        
        try:
            logging.info(f"Fetching workbook content (TWB) for discovery: {workbook_id}")
            response = requests.get(url, headers=self._get_headers())
            response.raise_for_status()
            
            # The content might be a ZIP (TWBX) or plain XML (TWB)
            content = response.content
            if content.startswith(b'PK'):
                import zipfile
                logging.info("Workbook is TWBX (zipped). Extracting TWB...")
                with zipfile.ZipFile(io.BytesIO(content)) as myzip:
                    # Find the first .twb file
                    twb_files = [f for f in myzip.namelist() if f.endswith('.twb')]
                    if not twb_files:
                        logging.warning("No TWB file found inside TWBX")
                        return []
                    content = myzip.read(twb_files[0])
            
            root = ET.fromstring(content)
            views = []
            
            # Find worksheets
            for ws in root.findall('.//worksheet'):
                name = ws.get('name')
                if name:
                    views.append({'id': None, 'name': name, 'sheetType': 'worksheet'})
            
            # Find dashboards
            for db in root.findall('.//dashboard'):
                name = db.get('name')
                if name:
                    views.append({'id': None, 'name': name, 'sheetType': 'dashboard'})
            
            logging.info(f"TWB discovery found {len(views)} sheets/dashboards in XML")
            return views
            
        except Exception as e:
            logging.error(f"TWB content discovery failed: {str(e)}")
            return []

    async def get_sheets_in_dashboard(self, workbook_id: str, dashboard_id: str) -> List[Dict]:
        """Get worksheets that are specifically inside a given dashboard using GraphQL."""
        metadata_url = f"{self.server_url}/api/metadata/graphql"
        query = """
        query GetDashboardSheets($luid: String!) {
          dashboards(filter: { luid: $luid }) {
            name
            luid
            sheets {
              name
              luid
              __typename
            }
          }
        }
        """
        try:
            response = requests.post(
                metadata_url,
                headers=self._get_headers(),
                json={
                    "query": query,
                    "variables": {"luid": dashboard_id}
                }
            )
            if response.status_code != 200:
                return []
            
            data = response.json()
            dashboards = data.get('data', {}).get('dashboards', [])
            if not dashboards:
                return []
            
            sheets = dashboards[0].get('sheets', [])
            return [{
                'id': s.get('luid'),
                'name': s.get('name'),
                'sheetType': 'worksheet'
            } for s in sheets if s.get('__typename') in ['Worksheet', 'Sheet']]
        except Exception as e:
            logging.error(f"Error getting sheets in dashboard {dashboard_id}: {e}")
            return []

    def _get_sheets_for_dashboard_from_twb(self, workbook_id: str, dashboard_name: str) -> List[str]:
        """Download workbook content and parse TWB XML to get the exact sheet names
        that belong to a specific dashboard. Returns an ordered list of sheet names."""
        import xml.etree.ElementTree as ET
        import urllib.parse as _urlparse

        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/content"
        try:
            logging.info(f"Downloading TWB to discover sheets for dashboard '{dashboard_name}'...")
            response = requests.get(url, headers=self._get_headers())
            response.raise_for_status()
            content = response.content

            # Unzip .twbx if needed
            if content[:2] == b'PK':
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    twb_files = [f for f in z.namelist() if f.endswith('.twb')]
                    if not twb_files:
                        logging.warning("No .twb found inside .twbx")
                        return []
                    content = z.read(twb_files[0])

            root = ET.fromstring(content)
            target_norm = _urlparse.unquote(dashboard_name).replace(' ', '').replace('_', '').lower()

            for dashboard in root.iter('dashboard'):
                name_attr = dashboard.attrib.get('name', '')
                name_norm = _urlparse.unquote(name_attr).replace(' ', '').replace('_', '').lower()
                if name_attr == dashboard_name or name_norm == target_norm or target_norm in name_norm:
                    # Collect all zone names (preserving order, deduplicating)
                    seen = set()
                    sheets = []
                    for zone in dashboard.iter('zone'):
                        sheet = zone.attrib.get('name')
                        if sheet and sheet not in seen:
                            seen.add(sheet)
                            sheets.append(sheet)
                    logging.info(f"TWB XML: dashboard '{name_attr}' contains {len(sheets)} sheet(s): {sheets}")
                    return sheets

            logging.warning(f"Dashboard '{dashboard_name}' not found in TWB XML (available: "
                            f"{[d.attrib.get('name') for d in root.iter('dashboard')]})")
        except Exception as e:
            logging.warning(f"TWB sheet discovery failed for dashboard '{dashboard_name}': {e}")
        return []

    def export_all_sheets_as_csv(self, workbook_id: str, filters: Dict[str, str] = None, max_rows_per_sheet: int = 500, scope_view_id: str = None, dashboard_name: str = None) -> str:
        """Export CSV data from sheets in a workbook.

        If dashboard_name is provided, the workbook TWB is downloaded and parsed first to
        get the exact sheets that belong to that dashboard (most accurate).
        If scope_view_id is provided, it attempts to only fetch data for sheets belonging to that dashboard.
        Otherwise, it fetches everything discovered.
        """
        try:
            views_to_fetch = []

            # STEP 0 (PRIMARY): Download the TWB and parse the XML to find exactly which
            # sheets belong to this dashboard.  This is the most accurate method because
            # it reads the authoring-level layout directly — no GraphQL guessing needed.
            if dashboard_name:
                twb_sheet_names = self._get_sheets_for_dashboard_from_twb(workbook_id, dashboard_name)
                if twb_sheet_names:
                    logging.info(f"TWB primary scope: resolving {len(twb_sheet_names)} dashboard sheets to REST API IDs...")
                    for sheet_name in twb_sheet_names:
                        resolved_id = self.resolve_view_id_in_workbook(workbook_id, sheet_name)
                        views_to_fetch.append({
                            'id': resolved_id or '',
                            'name': sheet_name,
                            'sheetType': 'worksheet'
                        })
                    logging.info(f"TWB scope resolved {sum(1 for v in views_to_fetch if v['id'])} / {len(views_to_fetch)} IDs")

            # STEP 1: If scoped to a dashboard/view ID, try to get only its related sheets
            if not views_to_fetch and scope_view_id:
                logging.info(f"Scoping data extraction to selection: {scope_view_id}")
                
                # Use GraphQL to determine if selection is a worksheet or a dashboard
                graphql_views = self.get_workbook_worksheets_graphql(workbook_id)
                
                # Find the target in the GraphQL results
                target_view = next((v for v in graphql_views if v.get('id') == scope_view_id), None)
                
                if target_view and target_view.get('sheetType') == 'worksheet':
                    logging.info(f"Selection is a specific WORKSHEET: {target_view.get('name')}")
                    views_to_fetch = [target_view]
                elif target_view and target_view.get('sheetType') == 'dashboard':
                    # It's a dashboard! Expand to its worksheets.
                    dashboard_worksheets = [v for v in graphql_views if v.get('parent_dashboard_id') == scope_view_id]
                    if dashboard_worksheets:
                        logging.info(f"Selection is a DASHBOARD: {target_view.get('name')}. Expanding to {len(dashboard_worksheets)} constituent worksheets.")
                        views_to_fetch = dashboard_worksheets
                    else:
                        logging.warning(f"Dashboard {target_view.get('name')} has no worksheets in GraphQL. Falling back to REST API views...")
                        # GraphQL didn't return worksheet children — try REST API view list.
                        # Filter out anything that matches a known dashboard name so we only
                        # attempt proper worksheet views for CSV export.
                        try:
                            rest_views = self.get_views_in_workbook(workbook_id)
                            dashboard_names_lower = {v.get('name', '').lower() for v in graphql_views if v.get('sheetType') == 'dashboard'}
                            rest_worksheets = [v for v in rest_views if v.get('name', '').lower() not in dashboard_names_lower]
                            if rest_worksheets:
                                logging.info(f"Found {len(rest_worksheets)} worksheet view(s) via REST API.")
                                views_to_fetch = rest_worksheets
                            else:
                                logging.warning("No distinct worksheet views via REST either. Using dashboard view as last resort.")
                                views_to_fetch = [target_view]
                        except Exception as rest_err:
                            logging.warning(f"REST fallback for worksheet discovery failed: {rest_err}. Using dashboard view.")
                            views_to_fetch = [target_view]
                else:
                    # Fallback if ID not found in GraphQL structured list
                    logging.warning(f"ID {scope_view_id} not explicitly found in GraphQL worksheets/dashboards. Fallback to basic view.")
                    all_wb_views = self.get_views_in_workbook(workbook_id)
                    single_view = next((v for v in all_wb_views if v.get('id') == scope_view_id), None)
                    views_to_fetch = [single_view] if single_view else []
                    # Try GraphQL for the dashboard context
                    try:
                        import requests
                        metadata_url = f"{self.server_url}/api/metadata/graphql"
                        query = """
                        query GetDashboardSheets($luid: String!) {
                          dashboards(filter: { luid: $luid }) {
                            name
                            sheets { name luid }
                          }
                        }
                        """
                        resp = requests.post(metadata_url, headers=self._get_headers(), 
                                             json={"query": query, "variables": {"luid": scope_view_id}})
                        
                        # FALLBACK 1: If dashboard query fails or returns no sheets, try workbook-level
                        if resp.status_code != 200 or not resp.json().get('data', {}).get('dashboards'):
                            logging.info(f"Dashboard query failed or no dashboard found for {scope_view_id}. Trying workbook sheets...")
                            query_wb = """
                            query GetWorkbookSheets($luid: String!) {
                              workbooks(filter: { luid: $luid }) {
                                name
                                sheets { name luid }
                              }
                            }
                            """
                            resp = requests.post(metadata_url, headers=self._get_headers(), 
                                                json={"query": query_wb, "variables": {"luid": workbook_id}})

                        if resp.status_code == 200:
                            data = resp.json().get('data', {})
                            db_data = data.get('dashboards', [])
                            wb_data = data.get('workbooks', [])
                            
                            raw_sheets = []
                            if db_data:
                                logging.info(f"Selection is a dashboard: {db_data[0].get('name')}")
                                raw_sheets = db_data[0].get('sheets', [])
                            elif wb_data:
                                logging.info(f"Using workbook sheets for scope: {wb_data[0].get('name')}")
                                raw_sheets = wb_data[0].get('sheets', [])

                            for rs in raw_sheets:
                                rs_name = rs.get('name')
                                rs_id = rs.get('luid')
                                
                                # Resolve ID by name if missing or empty
                                if not rs_id or rs_id == "":
                                    resolved_id = self.resolve_view_id_in_workbook(workbook_id, rs_name)
                                    if resolved_id: 
                                        rs_id = resolved_id
                                    else:
                                        logging.warning(f"Sheet '{rs_name}' has no LUID and could not be resolved. It will likely be skipped by REST API.")
                                
                                views_to_fetch.append({
                                    'id': rs_id or '',
                                    'name': rs_name,
                                    'sheetType': 'worksheet'
                                })
                    except Exception as e:
                        logging.warning(f"Scoped GraphQL fetch failed: {e}")
                
                if not views_to_fetch:
                    logging.warning(f"Could not resolve any worksheets for scope_view_id '{scope_view_id}'.")

            # 2. If no scoped views found, use workbook-wide discovery
            if not views_to_fetch:
                views_to_fetch = self.get_workbook_worksheets_graphql(workbook_id)
            
            # Fallback to TWB/REST as before
            if not views_to_fetch:
                logging.info(f"GraphQL discovery yielded no worksheets. Falling back to TWB content parsing (XML)...")
                views_to_fetch = self.get_views_via_twb(workbook_id)
            
            if not views_to_fetch:
                logging.info(f"TWB parsing yielded no results. Falling back to REST API views list.")
                views_to_fetch = self.get_views_in_workbook(workbook_id)
            
            # 3. Deduplicate by sheet ID or name
            seen_ids = set()
            seen_names = set()
            unique_views = []
            for v in views_to_fetch:
                v_id = v.get('id')
                v_name = v.get('name')
                
                if v_id:
                    if v_id not in seen_ids:
                        seen_ids.add(v_id)
                        unique_views.append(v)
                elif v_name:
                    if v_name not in seen_names:
                        seen_names.add(v_name)
                        unique_views.append(v)
            
            logging.info(f"Proceeding to fetch CSV from {len(unique_views)} discovered views/worksheets")
            
            all_csv_parts = []
            total_data_rows = 0
            successful_sheets = []
            
            for view in unique_views:
                view_id = view.get('id', '')
                view_name = view.get('name', 'Unknown')
                sheet_type = view.get('sheetType', 'N/A')
                
                try:
                    logging.info(f"Fetching CSV data for '{view_name}' (type: {sheet_type}, ID: {view_id})")
                    # If we have an ID, we use it. If not (from TWB), we try to fetch by name.
                    if not view_id:
                        view_id = self.resolve_view_id_in_workbook(workbook_id, view_name)
                        if view_id:
                            logging.info(f"✓ Resolved ID for '{view_name}': {view_id}")

                    if view_id:
                        csv_text = self.export_view_as_csv(view_id, filters=filters, max_rows=max_rows_per_sheet)
                    else:
                        # Last Resort: Check if it's in the standard workbook views list (unlikely if we are here)
                        all_views = self.get_views_in_workbook(workbook_id)
                        target_view = next((v for v in all_views if v.get('name') == view_name), None)
                        if target_view:
                            csv_text = self.export_view_as_csv(target_view.get('id'), filters=filters, max_rows=max_rows_per_sheet)
                        else:
                            logging.warning(f"Could not find ID for hidden sheet name '{view_name}'. Skipping.")
                            continue
                    
                    if csv_text and csv_text.strip():
                        # Check if this looks like actual data (has multiple lines/rows)
                        line_count = len([l for l in csv_text.split('\n') if l.strip()])
                        if line_count > 1:  # Has at least header + 1 data row
                            # Add sheet label header
                            all_csv_parts.append(f"=== Sheet: {view_name} ===")
                            all_csv_parts.append(csv_text)
                            all_csv_parts.append("")  # blank line between sheets
                            
                            row_count = len([l for l in csv_text.split('\n') if l.strip() and not l.strip().startswith('...')])
                            total_data_rows += max(0, row_count - 1)  # minus header
                            logging.info(f"✓ SUCCESS: Got {row_count} lines from '{view_name}' (type: {sheet_type})")
                            successful_sheets.append(view_name)
                        else:
                            logging.warning(f"CSV returned but appears empty (only {line_count} lines) for '{view_name}'")
                    else:
                        logging.warning(f"No CSV data returned for '{view_name}' (type: {sheet_type})")
                except Exception as view_err:
                    logging.warning(f"Failed to fetch CSV for '{view_name}': {str(view_err)}")
                    continue
            
            # If we still haven't got any data, try a more aggressive approach
            if not all_csv_parts:
                logging.info("No data from views list. Trying alternative extraction method...")
                
                # Try fetching datasources and see if that helps identify the right endpoint
                datasources = self.get_workbook_datasources(workbook_id)
                logging.info(f"Workbook has {len(datasources)} datasources")
                
                # Try each view again but without filters (sometimes filters break things)
                if filters:
                    logging.info("Retrying all views WITHOUT filters...")
                    for view in unique_views:
                        # Skip views explicitly known to be dashboards; REST API views have no
                        # sheetType key and should always be retried (they may be worksheets).
                        if view.get('sheetType', '').lower() == 'dashboard':
                            continue
                        if view.get('id'):
                            try:
                                csv_text = self.export_view_as_csv(view['id'], filters=None, max_rows=max_rows_per_sheet)
                                if csv_text and len(csv_text.split('\n')) > 1:
                                    all_csv_parts.append(f"=== Sheet: {view['name']} (no filters) ===")
                                    all_csv_parts.append(csv_text)
                                    all_csv_parts.append("")
                                    successful_sheets.append(view['name'])
                                    logging.info(f"✓ Got data for '{view['name']}' without filters!")
                            except Exception as retry_err:
                                logging.warning(f"Retry failed for '{view['name']}: {retry_err}")
            
            if all_csv_parts:
                combined = '\n'.join(all_csv_parts)
                sheet_count = len([p for p in all_csv_parts if p.startswith('===')])
                logging.info(f"✓✓✓ SUCCESS: Combined CSV data from {sheet_count} sheets: {', '.join(successful_sheets)} ({total_data_rows} total data rows)")
                return combined
            
            logging.error("⚠️ CRITICAL: No CSV data collected from ANY sheet in workbook!")
            logging.error("   Possible causes:")
            logging.error("   1. Workbook uses custom SQL or stored procedures that don't support crosstab")
            logging.error("   2. All views are dashboard-level (not individual worksheets)")
            logging.error("   3. Permissions issue - user cannot export data")
            logging.error("   4. Tableau Server configuration blocks data exports")
            return ""
            
        except Exception as e:
            logging.error(f"Failed to export all sheets as CSV: {str(e)}", exc_info=True)
            return ""

    def export_view_as_csv(self, view_id: str, filters: Dict[str, str] = None, max_rows: int = 200) -> str:
        """Export the underlying data of a view as CSV text.
        
        Uses the Tableau REST API endpoint: GET /sites/{site-id}/views/{view-id}/data
        Falls back to crosstab Excel endpoint if CSV fails.
        Returns the CSV content as a string, truncated to max_rows for AI context limits.
        """
        params = {}
        if filters:
            for key, value in filters.items():
                if value:
                    if key.startswith('vf_'):
                        params[key] = value
                    else:
                        params[f"vf_{key}"] = value
        
        logging.debug(f"Attempting CSV export for view {view_id} with params: {params}")
        
        # Attempt 1: Query View Data (CSV endpoint)
        csv_text = self._try_csv_endpoint(view_id, params)
        
        if csv_text:
            logging.info(f"✓ CSV endpoint succeeded for view {view_id}")
        else:
            logging.debug(f"CSV endpoint failed, trying crosstab for view {view_id}")
            # Attempt 2: Download View Crosstab Excel and convert to CSV
            csv_text = self._try_crosstab_endpoint(view_id, params)
            if csv_text:
                logging.info(f"✓ Crosstab endpoint succeeded for view {view_id}")
            else:
                logging.warning(f"⚠️ Both CSV and crosstab endpoints failed for view {view_id}")
        
        if not csv_text:
            return ""
        
        # Truncate to max_rows to stay within AI token limits
        lines = csv_text.split('\n')
        original_line_count = len(lines)
        if len(lines) > max_rows + 1:  # +1 for header row
            csv_text = '\n'.join(lines[:max_rows + 1])
            csv_text += f"\n... (truncated, showing {max_rows} of {original_line_count - 1} total rows)"
        
        logging.info(f"Successfully fetched view data for view {view_id} ({original_line_count} rows, returning {len(csv_text)} chars)")
        return csv_text

    def _try_csv_endpoint(self, view_id: str, params: dict) -> str:
        """Try the Query View Data CSV endpoint."""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/views/{view_id}/data"
        
        try:
            logging.info(f"Trying CSV data endpoint for view {view_id}")
            # Only use X-Tableau-Auth header — do NOT set Accept header
            # The /data endpoint returns CSV in the response body by default
            headers = {"X-Tableau-Auth": self.token}
            
            response = requests.get(url, headers=headers, params=params)
            response.raise_for_status()
            
            csv_text = response.text.strip()
            if csv_text:
                logging.info(f"CSV endpoint succeeded for view {view_id}")
                return csv_text
            return ""
            
        except requests.exceptions.RequestException as e:
            logging.warning(f"CSV endpoint failed for view {view_id}: {str(e)}")
            return ""

    def _try_crosstab_endpoint(self, view_id: str, params: dict) -> str:
        """Try the Download View Crosstab Excel endpoint and convert to CSV."""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/views/{view_id}/crosstab/excel"
        
        try:
            logging.info(f"Trying crosstab Excel endpoint for view {view_id}")
            logging.debug(f"Crosstab URL: {url}")
            headers = {"X-Tableau-Auth": self.token}
            
            response = requests.get(url, headers=headers, params=params)
            logging.debug(f"Crosstab response status: {response.status_code}, content-type: {response.headers.get('Content-Type', 'N/A')}, size: {len(response.content)} bytes")
            
            if response.status_code != 200:
                logging.warning(f"Crosstab endpoint returned non-200 status: {response.status_code}")
                return ""
            
            # Check if response is actually Excel or an error
            content_type = response.headers.get('Content-Type', '')
            if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type and not content_type.startswith('application/vnd.ms-excel'):
                logging.warning(f"Crosstab response has unexpected content-type: {content_type}")
            
            # Convert Excel bytes to CSV string
            import io
            import csv as csv_module
            try:
                import openpyxl
                logging.debug(f"Parsing Excel file ({len(response.content)} bytes) with openpyxl...")
                wb = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
                
                all_csv_texts = []
                for ws in wb.worksheets:
                    logging.debug(f"Parsing Excel worksheet: {ws.title}, dimensions: {ws.dimensions}")
                    output = io.StringIO()
                    writer = csv_module.writer(output)
                    row_count = 0
                    for row in ws.iter_rows(values_only=True):
                        # Filter out rows that are entirely empty
                        if any(cell is not None for cell in row):
                            writer.writerow(row)
                            row_count += 1
                    
                    csv_part = output.getvalue().strip()
                    if csv_part:
                        all_csv_texts.append(f"--- Workbook Sheet: {ws.title} ---\n{csv_part}")
                        logging.debug(f"Converted '{ws.title}' to CSV: {row_count} rows")

                wb.close()
                csv_text = "\n\n".join(all_csv_texts)
                
                if csv_text:
                    logging.info(f"✓ Crosstab Excel endpoint succeeded for view {view_id} (Extracted {len(all_csv_texts)} sub-sheets)")
                    return csv_text
                else:
                    logging.warning(f"Crosstab produced empty CSV for view {view_id}")
            except ImportError:
                logging.error("openpyxl not installed — cannot parse crosstab Excel. Install with: pip install openpyxl")
            except Exception as excel_err:
                logging.error(f"Failed to parse Excel from crosstab endpoint: {excel_err}", exc_info=True)
            
            return ""
            
        except requests.exceptions.RequestException as e:
            logging.error(f"Crosstab endpoint failed for view {view_id}: {str(e)}")
            return ""

    def export_view_as_pdf(self, view_id: str, filters: Dict[str, str] = None) -> bytes:
        """Export a view as PDF and return the content, optionally applying filters"""
        url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/views/{view_id}/pdf"
        
        params = {}
        if filters:
            for key, value in filters.items():
                if value:
                    # HEURISTIC: If it already has vf_ prefix, keep it. 
                    # If not, try both? Or just use vf_ for now but allow override.
                    # Tableau REST API: View filters need vf_, Parameters do NOT.
                    if key.startswith('vf_'):
                        params[key] = value
                    else:
                        # We'll try applying as both for maximum compatibility 
                        # OR if we know it's a parameter, we drop vf_.
                        # For now, let's try applying as both vf_Key and Key
                        params[f"vf_{key}"] = value
                        params[key] = value
        
        try:
            logging.info(f"DEBUG: Exporting view {view_id} with combined params: {params}")
            logging.info(f"DEBUG: Full Export URL: {url}")
            response = requests.get(url, headers=self._get_headers(), params=params)
            response.raise_for_status()
            
            logging.info(f"Successfully exported view {view_id} as PDF (Size: {len(response.content)} bytes)")
            return response.content
            
        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to export view {view_id} as PDF: {str(e)}")
            raise Exception(f"Failed to export dashboard as PDF: {str(e)}")
    
    def get_dashboard_layout(self, workbook_id: str, dashboard_name: str) -> List[Dict]:
        """Download workbook XML and extract worksheet positions for a dashboard."""
        try:
            logging.info(f"Extracting layout for dashboard '{dashboard_name}' in workbook {workbook_id}...")
            url = f"{self.server_url}/api/{self.api_version}/sites/{self.site_id_response}/workbooks/{workbook_id}/content"
            response = requests.get(url, headers=self._get_headers())
            response.raise_for_status()
            
            content = response.content
            xml_content = None
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as z:
                    twb_files = [f for f in z.namelist() if f.endswith('.twb')]
                    if twb_files:
                        xml_content = z.read(twb_files[0])
            except zipfile.BadZipFile:
                xml_content = content
                
            if not xml_content:
                return []

            root = ET.fromstring(xml_content)
            
            # Find the target dashboard
            # Note: Dashboard names in XML might have spaces replaced or encoded.
            # We use a normalized comparison.
            def normalize(s): return re.sub(r'[^a-zA-Z0-9]', '', s).lower()
            target_norm = normalize(dashboard_name)
            
            dashboard_node = None
            for db in root.findall('.//dashboard'):
                if normalize(db.get('name', '')) == target_norm:
                    dashboard_node = db
                    break
            
            if dashboard_node is None:
                logging.warning(f"Dashboard '{dashboard_name}' not found in XML.")
                return []

            # Extract zones (worksheets)
            # Layout -> zones contains a hierarchy of zones. We look for those with name (worksheet name)
            # and non-zero dimensions.
            worksheet_zones = []
            for zone in dashboard_node.findall('.//zone'):
                ws_name = zone.get('name')
                # Only zones with type 'worksheet' or that have a name and dimensions are useful
                if ws_name and zone.get('x') is not None:
                    try:
                        worksheet_zones.append({
                            'name': ws_name,
                            'x': int(zone.get('x', 0)),
                            'y': int(zone.get('y', 0)),
                            'w': int(zone.get('w', 0)),
                            'h': int(zone.get('h', 0))
                        })
                    except ValueError:
                        continue
            
            logging.info(f"Extracted {len(worksheet_zones)} worksheet zones from dashboard layout.")
            return worksheet_zones
            
        except Exception as e:
            logging.error(f"Failed to get dashboard layout: {str(e)}")
            return []

    def sign_out(self):
        """Sign out and invalidate the authentication token"""
        if not self.token:
            return
        
        url = f"{self.server_url}/api/{self.api_version}/auth/signout"
        
        try:
            response = requests.post(url, headers=self._get_headers())
            response.raise_for_status()
            logging.info("Successfully signed out")
            
        except requests.exceptions.RequestException as e:
            logging.warning(f"Error during sign out: {str(e)}")
        
        finally:
            self.token = None
            self.site_id_response = None
            self.user_id = None

class TableauClient:
    """Official Tableau Server Client (TSC) wrapper for robust operations"""
    def __init__(self, server_url: str, site_id: str, token_name: str, token_key: str):
        self.server_url = server_url.rstrip('/')
        self.site_id = site_id
        self.token_name = token_name
        self.token_key = token_key
        self.tableau_auth = TSC.PersonalAccessTokenAuth(token_name, token_key, site_id)
        self.server = TSC.Server(self.server_url, use_server_version=True)
        
    def fetch_data_by_names(self, workbook_id: str, sheet_names: List[str]) -> str:
        """Fetch CSV data for multiple sheets by name using TSC"""
        combined_csv = ""
        try:
            with self.server.auth.sign_in(self.tableau_auth):
                logging.info(f"TSC: Signed in to {self.server_url}")
                
                # Get the workbook to find its views
                workbook = self.server.workbooks.get_by_id(workbook_id)
                self.server.workbooks.populate_views(workbook)
                
                for sheet_name in sheet_names:
                    # 1. Exact case-insensitive match
                    view = next((v for v in workbook.views if v.name.lower() == sheet_name.lower()), None)
                    
                    # 2. Fuzzy match: check if name is contained or contains
                    # CRITICAL: We avoid matching the generic "Dashboard" view to a specific sheet request
                    if not view:
                         view = next((v for v in workbook.views 
                                    if (sheet_name.lower() in v.name.lower() or v.name.lower() in sheet_name.lower())
                                    and v.name.lower() != "dashboard"), None)

                    if not view:
                        logging.warning(f"TSC: Could not find sheet matching '{sheet_name}' in workbook {workbook_id} (Views available: {[v.name for v in workbook.views]})")
                        continue
                    
                    logging.info(f"TSC: Fetching data for view '{view.name}' ({view.id})")
                    
                    # Fetch CSV data
                    csv_options = TSC.CSVRequestOptions()
                    self.server.views.populate_csv(view, csv_options)
                    
                    sheet_csv_data = b"".join(view.csv).decode('utf-8')
                    combined_csv += f"=== Sheet: {view.name} (via TSC) ===\n"
                    combined_csv += sheet_csv_data + "\n\n"
                    
            return combined_csv
        except Exception as e:
            logging.error(f"TSC: Failed to fetch data: {e}")
            return combined_csv
    def fetch_all_data(self, workbook_id: str) -> str:
        """Fallback method to fetch data for EVERY view in the workbook and combine them."""
        combined_csv = ""
        try:
            logging.info(f"TSC Fetch All: Discovery start for workbook {workbook_id}")
            with self.server.auth.sign_in(self.tableau_auth):
                workbook = self.server.workbooks.get_by_id(workbook_id)
                self.server.workbooks.populate_views(workbook)
                
                for view in workbook.views:
                    logging.info(f"TSC Fetch All: Capturing '{view.name}'...")
                    try:
                        csv_data = self._get_view_csv(view)
                        if csv_data:
                            combined_csv += f"=== Sheet: {view.name} ===\n"
                            combined_csv += csv_data + "\n\n"
                    except Exception as e:
                        logging.warning(f"TSC Fetch All: Failed to capture '{view.name}': {e}")
            return combined_csv
        except Exception as e:
            logging.error(f"TSC Fetch All Failed: {e}")
            return ""

    def fetch_data_from_workbook_views(self, workbook_id: str, dashboard_name: str = "Dashboard") -> str:
        """
        Implementation following the USER'S SUGGESTED PATTERN:
        Get workbook -> populate views -> iterate views -> populate_csv (skipping dashboard).
        """
        combined_csv = ""
        try:
            with self.server.auth.sign_in(self.tableau_auth):
                logging.info(f"TSC Pattern: Discovering views in workbook {workbook_id}")
                workbook = self.server.workbooks.get_by_id(workbook_id)
                self.server.workbooks.populate_views(workbook)
                
                for view in workbook.views:
                    # Skip the dashboard view per user snippet
                    if view.name.lower() == dashboard_name.lower():
                        logging.info(f"TSC Pattern: Skipping dashboard view '{view.name}'")
                        continue
                        
                    try:
                        logging.info(f"TSC Pattern: Downloading CSV for '{view.name}' ({view.id})")
                        self.server.views.populate_csv(view)
                        sheet_csv = b"".join(view.csv).decode('utf-8')
                        
                        combined_csv += f"=== Sheet: {view.name} (TSC Pattern) ===\n"
                        combined_csv += sheet_csv + "\n\n"
                    except Exception as e:
                        logging.warning(f"TSC Pattern: Failed to download '{view.name}': {e}")
                        
            return combined_csv
        except Exception as e:
            logging.error(f"TSC Pattern Failed: {e}")
            return ""

    def _get_view_csv(self, view_item) -> str:
        """Helper to get CSV from a view item using TSC populate_csv."""
        import io
        self.server.views.populate_csv(view_item)
        csv_bytes = b''.join(view_item.csv)
        return csv_bytes.decode('utf-8')

class TableauHyperExtractor:
    """
    High-fidelity data reconstruction engine for Tableau.
    Downloads .twbx, parses .twb XML for worksheet definitions,
    and processes the .hyper extract locally via pandas.
    """
    def __init__(self, server_url: str, site_id: str, token: str, output_dir: str = "./tableau_exports"):
        self.server_url = server_url.rstrip('/')
        self.site_id = site_id
        self.token = token
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def download_workbook(self, workbook_id: str) -> bytes:
        """Download workbook content (.twbx or .twb)"""
        url = f"{self.server_url}/api/3.20/sites/{self.site_id}/workbooks/{workbook_id}/content"
        headers = {"X-Tableau-Auth": self.token}
        logging.info(f"Downloading workbook {workbook_id} for reconstruction...")
        res = requests.get(url, headers=headers)
        res.raise_for_status()
        return res.content

    @staticmethod
    def _filter_field_from_column(column_attr: str):
        """Parse a .twb filter column attr into (field_name, kind), or None for
        Tableau internals that are not user-facing filters.
        '[federated.x].[none:channel:nk]'  → ('channel', 'categorical')
        '[federated.x].[yr:ORDER_DATE:ok]' → ('ORDER_DATE', 'year')
        '[federated.x].[:Measure Names]'   → None
        '[federated.x].[Action (Product)]' → None (dashboard action)"""
        if not column_attr:
            return None
        # Split on '].[' (not '.') — field names may themselves contain dots,
        # e.g. '[federated.x].[Action (Metric Calc. Symbol,Country)]'
        token = column_attr.split('].[')[-1].strip('[]')
        if token.startswith('Action ('):
            return None
        parts = token.split(':')
        kind = 'categorical'
        if len(parts) == 3:
            field = parts[1]
            if parts[0] == 'yr':
                kind = 'year'
        elif len(parts) == 2:
            field = parts[1] or parts[0]
        else:
            field = parts[0]
        field = field.strip()
        if not field or field.lower() in ('measure names', 'measure values'):
            return None
        # Tableau encoding tokens are never real field names — a parse that
        # lands on one means the column reference had a shape we don't model.
        if field.lower() in ('usr', 'none', 'attr', 'nk', 'ok', 'qk', 'yr', 'mn', 'qr', 'tdy'):
            return None
        return field, kind

    def get_dashboard_filter_options(self, workbook_content: bytes, dashboard_name: str, max_values: int = 200) -> List[Dict]:
        """Discover the categorical filters of a dashboard from the .twb XML and
        their possible values from the .hyper extract. Powers value-aware filter
        dropdowns so users pick exact (case-correct) values instead of typing.
        Returns [{'name': display_name, 'kind': 'categorical'|'year', 'values': [...]}]."""
        import re as _re
        import uuid as _uuid

        hyper_path = os.path.join(self.output_dir, f"filteropt_{_uuid.uuid4().hex}.hyper")
        has_hyper = False
        try:
            xml_bytes = workbook_content
            if zipfile.is_zipfile(io.BytesIO(workbook_content)):
                with zipfile.ZipFile(io.BytesIO(workbook_content)) as z:
                    twb_files = [f for f in z.namelist() if f.endswith('.twb')]
                    hyper_files = [f for f in z.namelist() if f.endswith('.hyper')]
                    if not twb_files:
                        return []
                    xml_bytes = z.read(twb_files[0])
                    if hyper_files:
                        with z.open(hyper_files[0]) as hf, open(hyper_path, 'wb') as out:
                            out.write(hf.read())
                        has_hyper = True

            root = ET.fromstring(xml_bytes)

            def norm(s):
                return _re.sub(r'[^a-z0-9]', '', (s or '').lower())

            # Caption map: internal names (e.g. Calculation_123...) → display names
            captions = {}
            for col in root.iter('column'):
                nm, cap = col.attrib.get('name', ''), col.attrib.get('caption')
                if nm and cap:
                    captions[nm.strip('[]').lower()] = cap

            # Worksheets that belong to the target dashboard
            target = norm(dashboard_name)
            sheet_names = set()
            for dash in root.iter('dashboard'):
                dn = dash.attrib.get('name', '')
                if not (dn == dashboard_name or norm(dn) == target
                        or (target and target in norm(dn))):
                    continue
                for zone in dash.iter('zone'):
                    zn = zone.attrib.get('name')
                    if zn:
                        sheet_names.add(norm(zn))
            if not sheet_names:
                sheet_names = {norm(ws.attrib.get('name', '')) for ws in root.iter('worksheet')}

            found = {}
            for ws in root.iter('worksheet'):
                if norm(ws.attrib.get('name', '')) not in sheet_names:
                    continue
                for f in ws.iter('filter'):
                    if f.attrib.get('class') != 'categorical':
                        continue
                    parsed = self._filter_field_from_column(f.attrib.get('column', ''))
                    if not parsed:
                        continue
                    field, kind = parsed
                    display = captions.get(field.lower(), field)
                    found.setdefault(field.lower(),
                                     {'name': display, 'kind': kind, 'values': []})

            # Distinct values from the extract — exact casing for the dropdowns.
            # Use targeted SELECT DISTINCT queries (not a full-table pandas load):
            # constant memory and fast even for large extracts / small instances.
            if found and has_hyper:
                try:
                    with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
                        with Connection(hyper.endpoint, hyper_path) as conn:
                            # Map normalized column name → (table, real column name)
                            col_map = {}
                            for schema in conn.catalog.get_schema_names():
                                for table in conn.catalog.get_table_names(schema=schema):
                                    tdef = conn.catalog.get_table_definition(table)
                                    for c in tdef.columns:
                                        cname = c.name.unescaped
                                        col_map.setdefault(norm(cname), (table, cname))
                            for fkey, opt in found.items():
                                key = norm(fkey)
                                hit = col_map.get(key) or next(
                                    (v for k, v in col_map.items()
                                     if key and (key in k or k in key)), None)
                                if not hit:
                                    continue
                                table, cname = hit
                                qcol = '"' + cname.replace('"', '""') + '"'
                                try:
                                    if opt['kind'] == 'year':
                                        rows = conn.execute_list_query(
                                            f'SELECT DISTINCT EXTRACT(YEAR FROM {qcol}) '
                                            f'FROM {table} WHERE {qcol} IS NOT NULL LIMIT 200')
                                        opt['values'] = sorted(
                                            {str(int(r[0])) for r in rows if r[0] is not None})
                                    else:
                                        rows = conn.execute_list_query(
                                            f'SELECT DISTINCT {qcol} FROM {table} '
                                            f'WHERE {qcol} IS NOT NULL LIMIT {int(max_values) + 1}')
                                        opt['values'] = sorted(
                                            {str(r[0]) for r in rows})[:max_values]
                                except Exception as qe:
                                    logging.debug(f"filter-options: distinct query failed for "
                                                  f"'{cname}' (non-fatal): {qe}")
                except Exception as e:
                    logging.warning(f"filter-options: Hyper value read failed (non-fatal): {e}")

            options = sorted(found.values(), key=lambda o: o['name'].lower())
            logging.info(f"filter-options: {len(options)} filter(s) for '{dashboard_name}': "
                         f"{[(o['name'], len(o['values'])) for o in options]}")
            return options
        except Exception as e:
            logging.error(f"get_dashboard_filter_options failed: {e}")
            return []
        finally:
            try:
                if os.path.exists(hyper_path):
                    os.remove(hyper_path)
            except Exception:
                pass

    @staticmethod
    def get_dashboard_design_size(workbook_content: bytes, dashboard_name: str) -> tuple:
        """
        Parse workbook content (.twbx or .twb) and return the (design_w, design_h)
        for the named dashboard (from its <size maxwidth=... maxheight=...> element).
        Returns (1000, 800) as a safe default if the element is not found.
        """
        import zipfile, io, xml.etree.ElementTree as ET, urllib.parse
        try:
            data = workbook_content
            if zipfile.is_zipfile(io.BytesIO(data)):
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    twb_files = [f for f in z.namelist() if f.endswith(".twb")]
                    if twb_files:
                        with z.open(twb_files[0]) as f:
                            data = f.read()

            root = ET.fromstring(data)
            target_norm = urllib.parse.unquote(dashboard_name).replace(" ", "").lower()

            for dash in root.iter("dashboard"):
                name = dash.attrib.get("name", "")
                name_norm = urllib.parse.unquote(name).replace(" ", "").replace("_", "").lower()
                is_target = (name == dashboard_name or name_norm == target_norm or target_norm in name_norm)
                if is_target:
                    size_el = dash.find("size")
                    if size_el is not None:
                        w = int(size_el.attrib.get("maxwidth", 1000))
                        h = int(size_el.attrib.get("maxheight", 800))
                        logging.info(f"get_dashboard_design_size: '{name}' → {w}×{h}")
                        return (w, h)
        except Exception as e:
            logging.warning(f"get_dashboard_design_size failed: {e}")
        return (1000, 800)

    def extract_and_parse(self, workbook_content: bytes, dashboard_name: str,
                          crop_data: Dict = None, image_size: Tuple[int, int] = None,
                          applied_filters: Dict = None) -> Tuple[pd.DataFrame, List[str], Dict]:
        """
        Extract .hyper and .twb, and parse worksheet definitions.
        - crop_data + image_size  → geometry mapping to find the specific cropped sheet(s).
        - applied_filters         → user-selected dashboard filter values (e.g. region, date
                                    range) that should narrow the Hyper data before
                                    reconstruction, matching what the dashboard actually shows.
        """
        hyper_path = os.path.join(self.output_dir, f"extract_{os.getpid()}.hyper")
        worksheet_defs = {}
        dashboard_sheets = []

        # Check if content is a zip file (.twbx) or plain XML (.twb)
        is_zip = zipfile.is_zipfile(io.BytesIO(workbook_content))
        
        if is_zip:
            with zipfile.ZipFile(io.BytesIO(workbook_content)) as z:
                twb_files = [f for f in z.namelist() if f.endswith(".twb")]
                hyper_files = [f for f in z.namelist() if f.endswith(".hyper")]
                
                if not hyper_files:
                    # Some workbooks might use TDS or other formats, but for our case we need Hyper
                    logging.warning("No .hyper file found in .twbx. If this is a live connection, reconstruction may fail.")
                else:
                    with z.open(hyper_files[0]) as hf:
                        with open(hyper_path, "wb") as out:
                            out.write(hf.read())
                
                if not twb_files:
                    raise Exception("No .twb file found in .twbx")
                
                # Extract the first .twb file to parse metadata
                with z.open(twb_files[0]) as f:
                    workbook_content = f.read() # Overwrite workbook_content with the .twb XML
        
        # Parse XML (from .twb content)
        root = ET.fromstring(workbook_content)
        
        # Map all worksheets and their definitions (using consolidated logic)
        worksheet_defs = self._parse_workbook_structure(root)
        logging.info(f"Parsed {len(worksheet_defs)} worksheet definitions.")

        import urllib.parse
        import html
        
        # List all dashboards for debugging
        all_dashboards = [d.attrib.get("name", "") for d in root.iter("dashboard")]
        logging.info(f"Available XML dashboards: {all_dashboards}")

        target_dash_norm = urllib.parse.unquote(dashboard_name).replace(" ", "").lower()
        zones_with_sheets = []     # zones with positive area — used for geometry mapping
        target_all_sheets = set()  # ALL sheet names referenced in the TARGET dashboard
                                   # (regardless of zone size) — used for dashboard_sheets fallback
        # Dashboard design dimensions — read from the target dashboard's <size> element
        design_w, design_h = 1000, 800  # sensible defaults if not found

        for dashboard in root.iter("dashboard"):
            name_attr = dashboard.attrib.get("name", "")
            name_attr_norm = urllib.parse.unquote(name_attr).replace(" ", "").replace("_", "").lower()

            is_target = (name_attr == dashboard_name or name_attr_norm == target_dash_norm or target_dash_norm in name_attr_norm)

            if is_target:
                logging.info(f"✓ Found exact or fuzzy matching dashboard: '{name_attr}'")
                # Read the fixed design dimensions so we can later compute the y_offset
                size_el = dashboard.find("size")
                if size_el is not None:
                    try:
                        design_w = int(size_el.attrib.get("maxwidth", design_w))
                        design_h = int(size_el.attrib.get("maxheight", design_h))
                        logging.info(f"  Dashboard design size from XML: {design_w}×{design_h}")
                    except Exception:
                        pass

            # Collect zones from every dashboard element.
            # Two separate collections:
            #   1. target_all_sheets  — ALL zone names in the TARGET dashboard (any size)
            #      Used later for dashboard_sheets fallback; captures hidden/floating panels.
            #   2. zones_with_sheets  — zones that have positive pixel area
            #      Used for geometry (crop) mapping where we need actual screen coordinates.
            for zone in dashboard.iter("zone"):
                sheet = zone.attrib.get("name")
                if not sheet:
                    continue
                # Collection 1: all sheets in target dashboard regardless of dimensions
                if is_target:
                    target_all_sheets.add(sheet)
                # Collection 2: zones with measurable area (for geometry mapping)
                try:
                    zx = int(zone.attrib.get("x", 0))
                    zy = int(zone.attrib.get("y", 0))
                    zw = int(zone.attrib.get("w", 0))
                    zh = int(zone.attrib.get("h", 0))
                    if zw > 0 and zh > 0:
                        zones_with_sheets.append({
                            'sheet': sheet,
                            'x': zx, 'y': zy, 'w': zw, 'h': zh,
                            'from_target': is_target,
                            'dash': name_attr
                        })
                except:
                    pass

        logging.info(f"Target dashboard sheet names (all zones): {sorted(target_all_sheets)}")

        # If we have a crop, find the best sheet among all collected zones
        # IMPORTANT: always prefer zones from the TARGET dashboard only.
        # Zones from other dashboards use a completely different layout and
        # their coordinates are meaningless relative to this crop.
        target_zones = [z for z in zones_with_sheets if z.get('from_target')]
        candidate_zones = target_zones if target_zones else zones_with_sheets
        if not target_zones:
            logging.warning(
                "[zone-match] No target-dashboard zones found; "
                "falling back to all-dashboard zones (may produce wrong sheet)."
            )

        if crop_data and image_size and candidate_zones:
            # replace zones_with_sheets with the filtered candidate list for scoring
            zones_with_sheets = candidate_zones
            # 1. Calculate scale factors
            # Tableau internal units for dashboards are always 0-100,000 for both X and Y.
            # When Tableau exports a fixed-size dashboard as a PDF on US Letter portrait
            # (1700×2200px at 200 DPI), it scales the content to fill the page *width*
            # and centres it vertically — leaving whitespace margins at top and bottom.
            orig_w, orig_h = image_size

            # ----------------------------------------------------------------
            # Determine y_offset (blank rows above dashboard content) by
            # pixel-scanning the actual PNG file.  This is reliable regardless
            # of how Tableau positions the content on the page — the mathematical
            # formula (page_h - content_h) / 2 can be wrong by hundreds of pixels.
            # ----------------------------------------------------------------
            y_offset  = 0.0
            content_h = float(orig_h)   # safe fallback: treat full image as content

            png_path_for_scan = crop_data.get('_png_path', '') if crop_data else ''
            if png_path_for_scan and os.path.exists(png_path_for_scan):
                try:
                    from PIL import Image as _PILImage
                    _img = _PILImage.open(png_path_for_scan).convert('L')
                    _pw, _ph = _img.size
                    BG = 245          # pixels brighter than this count as background
                    first_row, last_row = None, None
                    for _y in range(_ph):
                        if min(_img.crop((0, _y, _pw, _y + 1)).getdata()) < BG:
                            if first_row is None:
                                first_row = _y
                            last_row = _y
                    if first_row is not None and last_row is not None:
                        y_offset  = float(first_row)
                        content_h = float(last_row - first_row + 1)
                        logging.info(
                            f"Pixel-scan y_offset={y_offset:.0f}px, "
                            f"content_h={content_h:.0f}px (rows {first_row}–{last_row})"
                        )
                except Exception as _scan_err:
                    logging.warning(f"PNG pixel-scan for y_offset failed: {_scan_err}")
                    # Fall through to mathematical estimate below
                    png_path_for_scan = ''

            if not png_path_for_scan:
                # Mathematical fallback when we cannot scan the image
                if design_w > 0 and design_h > 0:
                    _scale    = orig_w / float(design_w)
                    content_h = design_h * _scale
                    y_offset  = max((orig_h - content_h) / 2.0, 0.0)
                    logging.info(
                        f"Math fallback y_offset={y_offset:.1f}px, "
                        f"content_h={content_h:.1f}px (design={design_w}×{design_h})"
                    )

            # X: no horizontal offset (content fills the full page width).
            x_ratio = orig_w / 100000.0
            # Y: map content_h pixels onto 100,000 Tableau units.
            y_ratio = content_h / 100000.0 if content_h > 0 else orig_h / 100000.0

            logging.info(
                f"Geometry Mapping: image={orig_w}×{orig_h}px, "
                f"y_offset={y_offset:.1f}px, content_h={content_h:.1f}px, "
                f"x_ratio={x_ratio:.6f}, y_ratio={y_ratio:.6f}"
            )

            # 2. Translate crop pixels → Tableau XML unit coordinates
            cx = crop_data.get('x', 0) / x_ratio
            cy = (crop_data.get('y', 0) - y_offset) / y_ratio
            cw = crop_data.get('width', 0) / x_ratio
            ch = crop_data.get('height', 0) / y_ratio

            logging.info(f"Geometry Mapping: Normalized crop pixels {crop_data} to units: {{'x': {cx:.1f}, 'y': {cy:.1f}, 'w': {cw:.1f}, 'h': {ch:.1f}}}")

            best_sheet = None
            max_score = -1

            # IoU (Intersection over Union) scoring:
            # score = overlap / (crop_area + zone_area - overlap)
            #
            # Why IoU instead of overlap/zone_area:
            #   overlap/zone_area gave small KPI tiles 100% whenever they sat
            #   entirely inside a large crop → returned 7+ sheets for one crop.
            # With IoU:
            #   • A tiny tile fully inside a large crop → score ≈ tile/crop (very low)
            #   • A zone whose size closely matches the crop → score near 1.0
            # This naturally selects the ONE chart zone that best matches the crop.
            crop_area = cw * ch if (cw > 0 and ch > 0) else 1.0

            # Track IoU for every intersecting zone
            all_iou_scores = {}   # sheet -> best IoU across all zone entries for that sheet

            for z in zones_with_sheets:
                overlap_x = max(cx, z['x'])
                overlap_y = max(cy, z['y'])
                overlap_w = min(cx + cw, z['x'] + z['w']) - overlap_x
                overlap_h = min(cy + ch, z['y'] + z['h']) - overlap_y

                if overlap_w > 0 and overlap_h > 0:
                    zone_area = float(z['w'] * z['h']) if (z['w'] * z['h']) > 0 else 1.0
                    overlap_area = overlap_w * overlap_h
                    union_area   = crop_area + zone_area - overlap_area
                    iou = overlap_area / union_area if union_area > 0 else 0.0

                    sheet_name_z = z['sheet']
                    logging.info(
                        f"  IoU with '{sheet_name_z}' (Dash: {z.get('dash')}): "
                        f"IoU={iou:.4f}, ZoneArea={zone_area:.0f}, CropArea={crop_area:.0f}, "
                        f"Overlap={overlap_area:.0f}"
                    )

                    # Keep best IoU per sheet name (a sheet can appear in multiple zones)
                    if sheet_name_z not in all_iou_scores or iou > all_iou_scores[sheet_name_z]:
                        all_iou_scores[sheet_name_z] = iou

            if all_iou_scores:
                # Sort all sheets by IoU descending
                sorted_iou = sorted(all_iou_scores.items(), key=lambda x: x[1], reverse=True)
                best_name, best_iou = sorted_iou[0]

                logging.info(f"[zone-match] IoU scores: " +
                             ", ".join(f"'{n}'={s:.4f}" for n, s in sorted_iou))

                # ── Multi-sheet threshold logic ───────────────────────────────
                # A crop can legitimately span multiple chart zones (e.g. a KPI
                # card stacked above a bar chart).  We want to return ALL zones
                # that genuinely "belong" to the crop, not just the single best.
                #
                # Rule: include every sheet whose IoU ≥ MIN_IOU_THRESHOLD.
                #
                # Why 0.10 works:
                #   • Two equal-sized charts in a crop each score ~0.40-0.50  → both included ✓
                #   • One chart (IoU 0.70) + tiny KPI tiles (IoU 0.01-0.03)   → only chart ✓
                #   • Crop just clips the edge of an adjacent zone (IoU 0.02)  → excluded ✓
                #
                # Hard cap at 5 sheets to prevent noise in very large crops.
                MIN_IOU_THRESHOLD = 0.10
                MAX_SHEETS = 5

                dashboard_sheets = [
                    name for name, iou in sorted_iou
                    if iou >= MIN_IOU_THRESHOLD
                ][:MAX_SHEETS]

                # Edge case: if nothing cleared the threshold (e.g. all scores
                # are tiny but the best is still the most relevant zone), fall
                # back to just the top match so we always return something.
                if not dashboard_sheets:
                    dashboard_sheets = [best_name]
                    logging.info(
                        f"[zone-match] No zone reached threshold {MIN_IOU_THRESHOLD}; "
                        f"returning best match '{best_name}' (IoU={best_iou:.4f})"
                    )

                logging.info(
                    f"✓ Geometry mapping (IoU≥{MIN_IOU_THRESHOLD}) matched crop to "
                    f"{len(dashboard_sheets)} sheet(s): {dashboard_sheets}"
                )
            else:
                logging.warning("Geometry mapping found no intersecting sheets in any dashboard.")
        
        logging.info(f"All collected zones with sheets: {zones_with_sheets}")

        # ── Fallback sheet resolution (no crop / no geometry match) ─────────────
        # Priority 1: ALL sheet names collected from the TARGET dashboard (any zone size)
        #   → `target_all_sheets` includes hidden/floating panels that have w=0 or h=0
        # Priority 2: sheets from zones with positive dimensions in the target dashboard
        #   → subset of #1 that have measurable screen area
        # Priority 3: sheets from any dashboard's zones (positive dimensions)
        #   → broadens to other dashboards if target had none
        # Priority 4: all worksheets in the workbook
        #   → last resort, produces large unfiltered dumps — avoid where possible
        if not dashboard_sheets:
            # Priority 1 — all zones of the target dashboard (regardless of size)
            p1_sheets = list(dict.fromkeys(
                s for s in sorted(target_all_sheets)   # sorted for determinism
                if s in worksheet_defs
            ))
            if p1_sheets:
                dashboard_sheets = p1_sheets
                logging.info(f"Fallback P1 → all target dashboard zones ({len(p1_sheets)}): {dashboard_sheets}")
            else:
                # Priority 2 — target dashboard zones with positive dimensions
                p2_sheets = list(dict.fromkeys(
                    z['sheet'] for z in zones_with_sheets
                    if z.get('from_target') and z['sheet'] in worksheet_defs
                ))
                if p2_sheets:
                    dashboard_sheets = p2_sheets
                    logging.info(f"Fallback P2 → target dashboard positive-dim zones: {dashboard_sheets}")
                else:
                    # Priority 3 — any dashboard's positive-dim zones
                    p3_sheets = list(dict.fromkeys(
                        z['sheet'] for z in zones_with_sheets
                        if z['sheet'] in worksheet_defs
                    ))
                    if p3_sheets:
                        dashboard_sheets = p3_sheets
                        logging.info(f"Fallback P3 → any dashboard zones: {dashboard_sheets}")
                    elif worksheet_defs:
                        dashboard_sheets = list(worksheet_defs.keys())
                        logging.warning(f"Fallback P4 (LAST RESORT) → all {len(dashboard_sheets)} workbook worksheets")

        # Load data from Hyper if available
        df_master = pd.DataFrame()
        if os.path.exists(hyper_path):
            try:
                df_master = self._read_hyper(hyper_path)
                logging.info(f"Loaded Hyper data: {len(df_master)} rows, Columns: {list(df_master.columns)}")

                # Apply user-selected dashboard filters so the extracted data matches
                # exactly what the dashboard is showing (e.g. date range, region).
                if applied_filters and not df_master.empty:
                    cols_lower = {c.lower().replace(' ', '_'): c for c in df_master.columns}
                    for filter_field, filter_value in applied_filters.items():
                        if filter_value in (None, '', [], {}):
                            continue
                        # Normalise the field name to find the matching column
                        key = filter_field.lower().replace(' ', '_').replace('-', '_')
                        matched_col = cols_lower.get(key)
                        if not matched_col:
                            # Try partial match
                            matched_col = next(
                                (c for k, c in cols_lower.items() if key in k or k in key), None
                            )
                        if not matched_col:
                            logging.info(f"  applied_filter '{filter_field}' — no matching column, skipping")
                            continue
                        values = filter_value if isinstance(filter_value, list) else [filter_value]
                        values_str = [str(v) for v in values]
                        before = len(df_master)
                        df_master = df_master[df_master[matched_col].astype(str).isin(values_str)]
                        logging.info(
                            f"  Applied dashboard filter [{matched_col} IN {values_str}]: "
                            f"{before} → {len(df_master)} rows"
                        )
            except Exception as e:
                logging.error(f"Failed to read Hyper file: {e}")
            
            # Cleanup hyper file
            try: os.remove(hyper_path)
            except: pass
        
        logging.info(f"Identified sheets for reconstruction: {dashboard_sheets}")
        return df_master, dashboard_sheets, worksheet_defs

    def reconstruct_csv(self, df_master: pd.DataFrame, dashboard_sheets: List[str], worksheet_defs: Dict) -> str:
        """
        For each worksheet in dashboard_sheets:
          1. Apply worksheet-level categorical + date/year filters (from TWB XML).
          2. Classify rows/cols shelf fields into dimensions and measures.
          3. Group the Hyper data by dimensions and aggregate measures.
          4. Emit a flat crosstab CSV section.

        Dimensions  → group-by columns (categories, dates, text fields).
        Measures    → aggregated numeric columns (SUM, AVG, COUNT, etc.).
        """
        if df_master.empty:
            logging.warning("reconstruct_csv: empty DataFrame — nothing to reconstruct.")
            return ""

        AGG_FUNCS   = {"sum", "avg", "cnt", "cntd", "max", "min"}
        PANDAS_AGG  = {"sum": "sum", "avg": "mean", "cnt": "count",
                       "cntd": "nunique", "max": "max", "min": "min"}
        combined_csv = ""

        for sheet_name in dashboard_sheets:
            if sheet_name not in worksheet_defs:
                logging.warning(f"[{sheet_name}] No worksheet definition found — skipping.")
                continue

            defn = worksheet_defs[sheet_name]
            logging.info(f"=== Reconstructing sheet: '{sheet_name}' ===")
            fc  = defn.get("field_catalog", {})
            df  = df_master.copy()

            # ── Step 1: Apply worksheet-level filters ─────────────────────────────────

            for flt in defn.get("categorical_filters", []):
                col = self._find_col(flt["col"], df.columns.tolist(), field_catalog=fc)
                if col:
                    str_vals = [str(v) for v in flt["values"]]
                    before   = len(df)
                    if flt.get("exclude"):
                        df = df[~df[col].astype(str).isin(str_vals)]
                    else:
                        df = df[df[col].astype(str).isin(str_vals)]
                    logging.info(
                        f"  Filter [{col} {'NOT IN' if flt.get('exclude') else 'IN'} "
                        f"{flt['values']}]: {before} → {len(df)} rows"
                    )
                else:
                    logging.warning(f"  Categorical filter field '{flt['col']}' not in Hyper columns.")

            for flt in defn.get("date_filters", []):
                col = self._find_col(flt["col"], df.columns.tolist(), field_catalog=fc)
                if col:
                    def _yr(v):
                        if hasattr(v, "year"): return int(v.year)
                        try: return int(pd.to_datetime(v, errors="coerce").year)
                        except: return 0
                    before = len(df)
                    df = df[df[col].apply(_yr).isin(flt["years"])]
                    logging.info(f"  Date filter [{col} IN years {flt['years']}]: {before} → {len(df)} rows")
                else:
                    logging.warning(f"  Date filter field '{flt['col']}' not in Hyper columns.")

            # Apply relative-date filters (Last N days / fixed date range)
            from datetime import datetime as _dt, timedelta as _td
            for flt in defn.get("relative_date_filters", []):
                col = self._find_col(flt["col"], df.columns.tolist(), field_catalog=fc)
                if not col:
                    logging.warning(
                        f"  Relative-date filter field '{flt['col']}' not found in Hyper columns.")
                    continue
                try:
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                    before  = len(df)
                    if flt["type"] == "fixed":
                        start = pd.Timestamp(flt["start_date"])
                        end   = pd.Timestamp(flt["end_date"])
                        df    = df[(df[col] >= start) & (df[col] <= end)]
                        logging.info(
                            f"  Fixed-date filter [{col} {flt['start_date']}→{flt['end_date']}]:"
                            f" {before}→{len(df)} rows")
                    elif flt["type"] == "relative":
                        periods_ago = flt.get("periods_ago")
                        if periods_ago is None:
                            logging.warning(
                                f"  Relative-date filter [{col}]: periods_ago unresolved "
                                f"(param_name='{flt.get('param_name','')}') — skipping.")
                            continue
                        gran  = flt.get("granularity", "day")
                        today = _dt.today()
                        if gran == "day":
                            cutoff = today - _td(days=periods_ago)
                        elif gran == "week":
                            cutoff = today - _td(weeks=periods_ago)
                        elif gran == "month":
                            cutoff = today - pd.DateOffset(months=int(periods_ago))
                        elif gran == "year":
                            cutoff = today - pd.DateOffset(years=int(periods_ago))
                        else:
                            cutoff = today - _td(days=periods_ago)
                        df = df[df[col] >= pd.Timestamp(cutoff)]
                        logging.info(
                            f"  Relative-date filter [{col} >= {cutoff.date()} "
                            f"({gran}×{periods_ago})]: {before}→{len(df)} rows")
                except Exception as exc:
                    logging.warning(f"  Relative-date filter error for '{flt['col']}': {exc}")

            if df.empty:
                logging.warning(f"  [{sheet_name}] No rows remain after filters — skipping.")
                continue

            # ── Step 2: Classify shelf fields into dimensions and measures ────────────
            # Walk rows_fields then cols_fields.  An aggregated field (agg in AGG_FUNCS)
            # is a measure; everything else is a dimension.

            dimensions: list = []   # actual DataFrame column names to group by
            measures:   list = []   # list of {col: str, agg: str, label: str}

            def _process_field(field: dict):
                agg     = (field.get("agg") or "none").lower()
                col_tok = field.get("col", "")
                caption = field.get("caption", col_tok)
                df_cols = df.columns.tolist()

                # ── 1. Try to resolve the token directly to a Hyper column ──────────
                actual = self._find_col(col_tok, df_cols, fc)
                if not actual:
                    actual = self._find_col(caption, df_cols, fc)

                # ── 2. For user calculations (agg="usr" or agg="tdy"/"qr"/"yr" on
                #       calculated fields), try formula resolution when the token itself
                #       is not in the Hyper data. ─────────────────────────────────────
                if not actual or agg == "usr":
                    fc_key = col_tok.upper().replace(" ", "_")
                    fc_entry = fc.get(fc_key, {})
                    formula  = fc_entry.get("formula", "")
                    role     = fc_entry.get("role", "")

                    if formula:
                        real_col, real_agg = self._parse_formula_field(formula, df_cols, fc)
                        if real_col:
                            lbl = fc_entry.get("caption", caption) or caption
                            if real_agg:
                                # Formula resolved to an aggregation → it's a measure
                                measures.append({"col": real_col,
                                                 "agg": real_agg,
                                                 "label": f"{real_agg.upper()}({lbl})"})
                            else:
                                # Formula resolved to a plain field → it's a dimension
                                dimensions.append(real_col)
                            return
                        # Formula exists but resolution failed — fall through to role check

                    if actual and role == "measure" and agg == "usr":
                        # Calculation IS in the Hyper extract — use as SUM measure
                        measures.append({"col": actual, "agg": "sum",
                                         "label": fc_entry.get("caption", caption) or caption})
                        return
                    elif not actual:
                        logging.debug(f"  Field '{col_tok}' (caption='{caption}', agg='{agg}') "
                                      f"not found in Hyper data and formula unresolvable — skipping.")
                        return
                    else:
                        # agg=="usr" but formula unresolvable and role is not "measure" — skip
                        logging.debug(f"  Field '{col_tok}' (agg='usr', role='{role}') unresolvable — skipping.")
                        return

                # Guard: if actual is still None after all resolution attempts, skip
                if actual is None:
                    logging.debug(f"  Field '{col_tok}' (agg='{agg}') unresolvable — skipping.")
                    return

                # ── 3. Standard aggregation → measure ────────────────────────────────
                if agg in AGG_FUNCS:
                    label = f"{agg.upper()}({caption})"
                    measures.append({"col": actual, "agg": PANDAS_AGG[agg], "label": label})

                # ── 4. Dimension with date truncation ─────────────────────────────────
                elif agg == "yr" or (agg in ("tdy", "none") and "date" in actual.lower()):
                    new_col = f"Year({actual})"
                    if new_col not in df.columns:
                        df[new_col] = df[actual].apply(
                            lambda v: int(v.year) if hasattr(v, "year")
                            else (lambda d: int(d.year) if pd.notnull(d) else None)(
                                pd.to_datetime(v, errors="coerce")))
                    dimensions.append(new_col)

                elif agg == "mn":
                    import calendar as _cal
                    new_col = f"Month({actual})"
                    if new_col not in df.columns:
                        def _month_name(v):
                            mo = int(v.month) if hasattr(v, "month") else (
                                lambda d: int(d.month) if pd.notnull(d) else None)(
                                pd.to_datetime(v, errors="coerce"))
                            return _cal.month_abbr[mo] if mo else None
                        df[new_col] = df[actual].apply(_month_name)
                    dimensions.append(new_col)

                elif agg == "qr":
                    new_col = f"Quarter({actual})"
                    if new_col not in df.columns:
                        def _qtr(v):
                            mo = int(v.month) if hasattr(v, "month") else (
                                lambda d: int(d.month) if pd.notnull(d) else None)(
                                pd.to_datetime(v, errors="coerce"))
                            return f"Q{(mo - 1) // 3 + 1}" if mo else None
                        df[new_col] = df[actual].apply(_qtr)
                    dimensions.append(new_col)

                # ── 5. Plain dimension ────────────────────────────────────────────────
                else:
                    dimensions.append(actual)

            for f in defn.get("rows_fields", []) + defn.get("cols_fields", []):
                _process_field(f)

            # Deduplicate while preserving order
            seen_d: set = set()
            dimensions = [c for c in dimensions if not (c in seen_d or seen_d.add(c))]
            seen_m: set = set()
            measures    = [m for m in measures
                           if not ((m["col"], m["agg"]) in seen_m
                                   or seen_m.add((m["col"], m["agg"])))]

            # If shelves had no measures, fall back to marks encodings
            if not measures:
                for f in defn.get("marks_measures", []):
                    _process_field(f)
                seen_m2: set = set()
                measures = [m for m in measures
                            if not ((m["col"], m["agg"]) in seen_m2
                                    or seen_m2.add((m["col"], m["agg"])))]

            logging.info(f"  Dimensions : {dimensions}")
            logging.info(f"  Measures   : {[(m['col'], m['agg'], m['label']) for m in measures]}")

            # ── Step 3: Build the crosstab (groupby + aggregate) ──────────────────────

            try:
                if dimensions and measures:
                    agg_dict = {m["col"]: m["agg"] for m in measures}
                    result   = df.groupby(dimensions, dropna=False).agg(agg_dict).reset_index()
                    # Rename measure columns to their human-readable labels
                    result.rename(columns={m["col"]: m["label"] for m in measures}, inplace=True)

                elif dimensions:
                    # Only dimensions present — count occurrences per group
                    result = (df.groupby(dimensions, dropna=False)
                                .size()
                                .reset_index(name="COUNT"))

                elif measures:
                    # Only measures — aggregate the entire dataset
                    agg_dict = {m["col"]: m["agg"] for m in measures}
                    row      = df.agg(agg_dict)
                    result   = row.to_frame(name="Value").T
                    result.rename(
                        columns={m["col"]: m["label"] for m in measures},
                        inplace=True)

                else:
                    # No fields resolved — emit a raw sample (first 100 rows)
                    logging.warning(f"  [{sheet_name}] No fields resolved — outputting raw sample.")
                    result = df.head(100)

                # ── Step 3b: Apply Top N filters ──────────────────────────────────────
                # Top N is applied AFTER aggregation so the ranking is on the
                # correct aggregated measure, not raw rows.
                for flt in defn.get("top_filters", []):
                    count      = flt.get("count", 10)
                    direction  = flt.get("direction", "top")
                    by_col_raw = flt.get("by_col")
                    ascending  = (direction == "bottom")

                    # Find the aggregated result column that matches by_col_raw
                    by_col_actual = None
                    if by_col_raw:
                        bk = by_col_raw.upper().replace(" ", "_")
                        for rc in result.columns:
                            if bk in rc.upper().replace(" ", "_"):
                                by_col_actual = rc
                                break
                    if not by_col_actual:
                        # Fallback: first numeric column in result
                        num_cols = result.select_dtypes(include="number").columns.tolist()
                        by_col_actual = num_cols[0] if num_cols else None

                    if by_col_actual:
                        before_n = len(result)
                        fn       = result.nlargest if not ascending else result.nsmallest
                        result   = fn(count, by_col_actual)
                        logging.info(
                            f"  Top-N [{direction} {count} by '{by_col_actual}']: "
                            f"{before_n}→{len(result)} rows")
                    else:
                        logging.warning(
                            f"  Top-N filter: could not resolve sort column "
                            f"for '{by_col_raw}' in result columns {list(result.columns)}")

                combined_csv += f"=== Sheet: {sheet_name} (Hyper-Reconstructed) ===\n"
                combined_csv += result.to_csv(index=False) + "\n\n"
                logging.info(f"  ✓ {len(result)} rows emitted for '{sheet_name}'")

            except Exception as exc:
                logging.error(f"  Reconstruction error for '{sheet_name}': {exc}", exc_info=True)
                # Graceful fallback: output the filtered rows for available columns
                fallback_cols = [c for c in (dimensions + [m["col"] for m in measures])
                                 if c in df.columns]
                if fallback_cols:
                    combined_csv += f"=== Sheet: {sheet_name} (Hyper-Reconstructed, raw fallback) ===\n"
                    combined_csv += df[fallback_cols].head(500).to_csv(index=False) + "\n\n"

        return combined_csv

    def _parse_workbook_structure(self, root):
        """
        Parse the .twb XML to build:
          - field_catalog : normalized internal name → {caption, role, datatype}
          - worksheet_defs: worksheet name →
              {rows_fields, cols_fields, categorical_filters, date_filters,
               marks_measures, field_catalog, field_map}

        rows_fields / cols_fields entries:
          {"agg": str, "col": str (normalized), "caption": str (human label), "token": str (raw)}
        categorical_filters entries:
          {"col": str, "values": [str], "exclude": bool}
        date_filters entries:
          {"col": str, "years": [int]}
        marks_measures entries:
          {"agg": str, "col": str}
        """
        # ── STEP 1: Build global field catalog ──────────────────────────────────────────
        # Entries carry: caption, role, datatype, AND formula (for calculated fields).
        # The formula lets us resolve usr: calculations to their underlying base columns.
        field_catalog = {}
        for ds in root.iter("datasource"):
            for col in ds.findall("column"):
                col_name    = col.attrib.get("name", "")
                col_caption = (col.attrib.get("caption") or
                               col_name.strip("[]").replace("&quot;", "").replace('"', ""))
                col_role    = col.attrib.get("role", "")
                col_dtype   = col.attrib.get("datatype", "")
                if not col_name:
                    continue
                # Extract calculation formula if present (e.g. "SUM([net_revenue_usd])")
                calc_el = col.find("calculation")
                formula = calc_el.attrib.get("formula", "") if calc_el is not None else ""
                clean = col_name.strip("[]").replace("&quot;", "").replace('"', "").strip()
                key   = clean.upper().replace(" ", "_")
                entry = {"caption": col_caption, "role": col_role,
                         "datatype": col_dtype, "formula": formula}
                field_catalog[key] = entry
                cap_key = col_caption.upper().replace(" ", "_")
                if cap_key and cap_key != key:
                    field_catalog[cap_key] = entry
                # IMPORTANT: also index by the normalized COLUMN PART only
                # (what _parse_field returns), so formula lookups work.
                # e.g.  "[usr:Calculation_123:qk]" → key = "USR:CALCULATION_123:QK"
                #       but _parse_field returns col = "CALCULATION_123"
                _, col_part = self._parse_field(col_name)
                if col_part and col_part not in field_catalog:
                    field_catalog[col_part] = entry

        logging.info(f"[field_catalog] Built {len(field_catalog)} entries (with formulas).")

        # ── STEP 1b: Parse global parameters ──────────────────────────────────────────
        # Parameters live in a datasource named "Parameters" (Tableau convention).
        parameters = {}
        for ds in root.iter("datasource"):
            if ds.attrib.get("name", "").lower() != "parameters":
                continue
            for col in ds.findall("column"):
                p_name    = col.attrib.get("name", "").strip("[]")
                p_caption = (col.attrib.get("caption") or p_name).strip()
                p_value   = col.attrib.get("value", "").strip()
                numeric_val = None
                # Direct integer value
                if p_value.lstrip("-").isdigit():
                    numeric_val = int(p_value)
                else:
                    # Extract leading number from strings like '21days', 'last_7'
                    nums = re.findall(r'\d+', p_value)
                    if nums:
                        numeric_val = int(nums[0])
                # Check alias list — current value key may map to a human label with a number
                for alias_el in col.findall("aliases/alias"):
                    if alias_el.attrib.get("key", "") == p_value:
                        alias_value = alias_el.attrib.get("value", "")
                        nums2 = re.findall(r'\d+', alias_value)
                        if nums2:
                            numeric_val = int(nums2[0])
                        break
                entry = {"caption": p_caption, "raw_value": p_value, "numeric": numeric_val}
                parameters[p_caption] = entry
                if p_name and p_name != p_caption:
                    parameters[p_name] = entry
        param_summary = ["%s=%s" % (k, v["numeric"]) for k, v in parameters.items()]
        logging.info(f"[parameters] Parsed {len(parameters)} parameter entries: {param_summary}")

        # ── STEP 2: Parse each worksheet ───────────────────────────────────────────────
        worksheet_defs = {}
        for worksheet in root.iter("worksheet"):
            w_name = worksheet.attrib.get("name")
            if not w_name:
                continue

            # Build worksheet-local token→caption map from datasource-dependencies
            field_map = {}
            for ds_deps in worksheet.findall(".//datasource-dependencies"):
                for col in ds_deps.findall("column"):
                    tok = col.attrib.get("name", "")
                    cap = (col.attrib.get("caption") or
                           tok.strip("[]").replace("&quot;", "").replace('"', ""))
                    if tok:
                        field_map[tok] = cap.strip()

            # Extract rows / cols shelf fields
            table = worksheet.find("table")
            rows_fields: list = []
            cols_fields: list = []
            if table is not None:
                self._extract_shelf_fields(table.find("rows"), rows_fields, field_map)
                self._extract_shelf_fields(table.find("cols"), cols_fields, field_map)

            # Extract worksheet-level filters
            categorical_filters: list = []
            date_filters: list = []
            top_filters: list = []
            relative_date_filters: list = []
            SKIP_COLS = {"MEASURE_NAMES", "MEASURE_VALUES", "NUMBER_OF_RECORDS"}
            for filt in worksheet.iter("filter"):
                filt_class = filt.attrib.get("class", "")
                field_attr = filt.attrib.get("column", "") or filt.attrib.get("field", "")
                if not field_attr:
                    continue
                agg, col = self._parse_field(field_attr)
                if not col or col in SKIP_COLS:
                    continue

                # ── Top N filter ──────────────────────────────────────────────────────
                if filt_class == "top":
                    top_el      = filt.find("top")
                    top_fld_el  = filt.find("top-field")
                    if top_el is not None:
                        direction = top_el.attrib.get("direction", "top")
                        try:
                            count = int(top_el.attrib.get("value", "10"))
                        except (ValueError, TypeError):
                            count = 10
                        by_col, by_agg_v = None, "sum"
                        if top_fld_el is not None:
                            tf_attr = top_fld_el.attrib.get("column", "")
                            by_agg_p, by_col_p = self._parse_field(tf_attr)
                            if by_col_p:
                                by_col   = by_col_p
                                by_agg_v = by_agg_p or "sum"
                        top_filters.append({
                            "col": col, "direction": direction,
                            "count": count, "by_col": by_col, "by_agg": by_agg_v,
                        })
                    continue

                # ── Relative-date filter ───────────────────────────────────────────────
                if filt_class == "relative-date":
                    date_el = filt.find("date-filter")
                    if date_el is not None:
                        start_date = date_el.attrib.get("start-date")
                        end_date   = date_el.attrib.get("end-date")
                        if start_date and end_date:
                            relative_date_filters.append({
                                "col": col, "type": "fixed",
                                "start_date": start_date, "end_date": end_date,
                            })
                        else:
                            granularity     = date_el.attrib.get("granularity", "day")
                            period_type     = date_el.attrib.get("period-type", "last")
                            periods_ago_raw = date_el.attrib.get("periods-ago", "")
                            param_name      = date_el.attrib.get("parameter-name", "").strip()
                            # Resolve periods_ago: literal integer or [Parameter N] reference
                            periods_ago = None
                            if not periods_ago_raw or periods_ago_raw.startswith("["):
                                # Parameter-driven — look up in parameters dict
                                if param_name:
                                    p_entry = parameters.get(param_name)
                                    if p_entry:
                                        periods_ago = p_entry.get("numeric")
                                    if periods_ago is None:
                                        # Also try stripping bracket from periods_ago_raw itself
                                        stripped = periods_ago_raw.strip("[]")
                                        p_entry2 = parameters.get(stripped)
                                        if p_entry2:
                                            periods_ago = p_entry2.get("numeric")
                            else:
                                try:
                                    periods_ago = int(periods_ago_raw)
                                except (ValueError, TypeError):
                                    nums = re.findall(r'\d+', periods_ago_raw)
                                    periods_ago = int(nums[0]) if nums else None
                            relative_date_filters.append({
                                "col": col, "type": "relative",
                                "granularity": granularity, "period_type": period_type,
                                "periods_ago": periods_ago, "param_name": param_name,
                            })
                    continue

                # ── Standard categorical / year-level date filter ─────────────────────
                values = [
                    gf.attrib.get("member", "").replace("&quot;", "").strip('"').strip()
                    for gf in filt.iter("groupfilter")
                    if gf.attrib.get("member")
                ]
                exclude = filt.attrib.get("exclude", "false").lower() == "true"
                if agg in ("yr", "mn", "qr") or "DATE" in col:
                    years = [int(v) for v in values if v.isdigit()]
                    if years:
                        date_filters.append({"col": col, "years": years})
                elif values:
                    categorical_filters.append({"col": col, "values": values, "exclude": exclude})

            # Extract marks-level measures (encodings inside panes)
            AGG_FUNCS = {"sum", "avg", "cnt", "cntd", "max", "min"}
            marks_measures: list = []
            for pane in worksheet.iter("pane"):
                for enc in pane.iter("encodings"):
                    for child in enc:
                        col_attr = child.attrib.get("column", "")
                        if not col_attr:
                            continue
                        if any(s in col_attr for s in ("Measure Values", "Multiple Values")):
                            continue
                        agg_m, col_m = self._parse_field(col_attr)
                        if agg_m and (agg_m in AGG_FUNCS or agg_m == "usr") and col_m and col_m not in SKIP_COLS:
                            marks_measures.append({"agg": agg_m, "col": col_m})

            logging.info(
                f"  [{w_name}] rows={[f['caption'] for f in rows_fields]} "
                f"cols={[f['caption'] for f in cols_fields]} "
                f"cat_filters={len(categorical_filters)} date_filters={len(date_filters)} "
                f"marks={len(marks_measures)}"
            )

            worksheet_defs[w_name] = {
                "rows_fields":            rows_fields,
                "cols_fields":            cols_fields,
                "categorical_filters":    categorical_filters,
                "date_filters":           date_filters,
                "top_filters":            top_filters,
                "relative_date_filters":  relative_date_filters,
                "marks_measures":         marks_measures,
                "field_catalog":          field_catalog,
                "field_map":              field_map,
            }

        return worksheet_defs

    # Internal Helpers (Logic from User's Snippet)
    def _parse_field(self, token):
        # Support tokens like [federated.123].[agg:col:type] or [:ColName] or [none:ColName:nk]
        # First, extract the content inside the LAST set of brackets
        inner_match = re.search(r'\[([^\[\]]+)\]$', token)
        if not inner_match:
            return None, None
        
        content = inner_match.group(1)
        # Split by colon
        parts = content.split(':')
        
        if len(parts) >= 3: # agg:col:type
            agg = parts[0].lower()
            col = parts[1]
        elif len(parts) == 2: # col:type or special:name
            if not parts[0]: # [:Measure Names]
                agg = None
                col = parts[1]
            else:
                agg = parts[0].lower()
                col = parts[1]
        else: # [FieldName]
            agg = "none"
            col = parts[0]
            
        # Clean col: remove leading dots/namespaces like [federated.xxx]
        # and convert to upper with underscores
        col = col.split('.')[-1]
        col = col.replace("&quot;", "").replace('"', "").strip()
        return agg, col.upper().replace(" ", "_")

    def _parse_fields_from_expr(self, expr):
        if not expr: return []
        fields = []
        seen = set()

        # First, try to match the federated.xxx.xxx pattern
        tokens = re.findall(r'\[federated[^\]]*\]\.\[[^\]]*\]', expr)
        if not tokens:
            # If no federated tokens, try simple single-bracket tokens like [agg:col:type]
            logging.info("    No federated tokens found, trying simple bracket tokens.")
            simple_tokens = re.findall(r'\[([^:\[\]]+):([^:\[\]]+):[^:\[\]]+\]', expr)
            for agg, col in simple_tokens:
                key = (agg.lower(), col.upper().replace(" ", "_"))
                if key not in seen:
                    seen.add(key)
                    fields.append({"agg": key[0], "col": key[1]})
            if fields:
                logging.info(f"    Found {len(fields)} simple fields in expr.")
            return fields

        for token in tokens:
            agg, col = self._parse_field(token)
            key = (agg, col)
            if agg and col and key not in seen:
                seen.add(key)
                fields.append({"agg": agg, "col": col})
        logging.info(f"    Found {len(fields)} fields in expr.")
        return fields

    def _rename_month_cols_wide(self, pivot: pd.DataFrame, measure: str) -> pd.DataFrame:
        """Rename numeric month columns into 'MMM_Measure' format for wide pivots."""
        import calendar
        pivot.columns = [
            f"{calendar.month_abbr[m]}_{measure}" if isinstance(m, int) and 1 <= m <= 12 
            else f"{m}_{measure}"
            for m in pivot.columns
        ]
        return pivot

    def _rename_month_cols_flat(self, df: pd.DataFrame, measure: str) -> pd.DataFrame:
        """Map numeric months to name strings for flattened (long) data."""
        import calendar
        col = f"MONTH_{measure}"
        if col in df.columns:
            df[col] = df[col].apply(lambda x: calendar.month_name[int(x)] if pd.notnull(x) and str(x).isdigit() and 1 <= int(x) <= 12 else x)
        return df

    def _find_col(self, col_name, df_cols, field_catalog=None):
        """
        Resolve a Tableau field name to an actual Hyper DataFrame column.

        Resolution order
        ────────────────
        1. Exact normalised match (UPPER + spaces→underscores)
        2. Caption lookup via field_catalog  (e.g. internal 'TOTAL_REV' → caption 'Total Revenue')
        3. Partial / contains match          (e.g. 'REVENUE' matches 'Total_Revenue')
        """
        if not col_name or not df_cols:
            return None

        target = col_name.upper().replace(" ", "_")

        # 1. Exact normalised match
        for c in df_cols:
            if c.upper().replace(" ", "_") == target:
                return c

        # 2. Caption fallback via field_catalog
        if field_catalog and target in field_catalog:
            caption = field_catalog[target].get("caption", "")
            if caption:
                cap_target = caption.upper().replace(" ", "_")
                for c in df_cols:
                    if c.upper().replace(" ", "_") == cap_target:
                        logging.debug(f"[_find_col] '{col_name}' → caption '{caption}' → '{c}'")
                        return c

        # 3. Partial / substring match (last resort — avoids false positives on short names)
        if len(target) >= 4:          # skip very short names to avoid wrong matches
            for c in df_cols:
                c_norm = c.upper().replace(" ", "_")
                if target in c_norm or c_norm in target:
                    logging.debug(f"[_find_col] '{col_name}' partial-matched → '{c}'")
                    return c

        return None

    def _get_group_cols(self, fields, df, field_catalog=None):
        group_cols = []
        df_cols = df.columns.tolist()
        for f in fields:
            agg, col_name = f["agg"], f["col"]
            col = self._find_col(col_name, df_cols, field_catalog=field_catalog)
            if not col:
                logging.warning(f"Column '{col_name}' not found in Hyper data (Normalized target: {col_name.upper().replace(' ', '_')})")
                continue
            
            if agg == "none":
                group_cols.append(col)
            elif agg == "yr":  # Year
                year_col = f"YEAR_{col}"
                def _get_y(v):
                    if hasattr(v, 'year'): return int(v.year)
                    try:
                        d = pd.to_datetime(v, errors='coerce')
                        return int(d.year) if pd.notnull(d) else 0
                    except: return 0
                df[year_col] = df[col].apply(_get_y)
                group_cols.append(year_col)
            elif agg == "mn":  # Month
                month_col = f"MONTH_{col}"
                def _get_m(v):
                    if hasattr(v, 'month'): return int(v.month)
                    try:
                        d = pd.to_datetime(v, errors='coerce')
                        return int(d.month) if pd.notnull(d) else 0
                    except: return 0
                df[month_col] = df[col].apply(_get_m)
                group_cols.append(month_col)
            elif agg == "qr":
                quarter_col = f"QUARTER_{col}"
                def _get_q(v):
                    if hasattr(v, 'quarter'): return int(v.quarter)
                    try:
                        d = pd.to_datetime(v, errors='coerce')
                        return int((d.month - 1) // 3 + 1) if pd.notnull(d) else 0
                    except: return 0
                df[quarter_col] = df[col].apply(_get_q)
                group_cols.append(quarter_col)
        return list(dict.fromkeys(group_cols)) # Dedupe

    def _get_measure_cols(self, fields, df, field_catalog=None):
        agg_map = {"sum": "sum", "avg": "mean", "cnt": "count", "cntd": "nunique", "max": "max", "min": "min"}
        measures = []
        seen = set()
        df_cols = df.columns.tolist()
        for f in fields:
            agg, col_name = f["agg"], f["col"]
            col = self._find_col(col_name, df_cols, field_catalog=field_catalog)
            if agg in agg_map and col and (agg, col) not in seen:
                seen.add((agg, col))
                measures.append((col, agg_map[agg]))
        return measures

    def _parse_formula_field(self, formula: str, df_cols: list, fc: dict):
        """
        Parse a simple Tableau calculation formula to extract a base Hyper column
        and its aggregation function.

        Handles:
          - "SUM([net_revenue_usd])"          → ("net_revenue_usd", "sum")
          - "AVG([net_revenue_usd])"          → ("net_revenue_usd", "mean")
          - "COUNT([event_id])"               → ("event_id", "count")
          - "COUNTD([customer_id])"           → ("customer_id", "nunique")
          - "{ FIXED [channel] : SUM([net_revenue_usd]) }" → base measure from FIXED LOD
          - Falls back to scanning all field references in the formula.

        Returns (actual_col, pandas_agg) or (None, None) if unresolvable.
        """
        if not formula:
            return None, None

        PANDAS_AGG = {"sum": "sum", "avg": "mean", "count": "count",
                      "countd": "nunique", "max": "max", "min": "min"}

        # ── Pattern 1: Simple aggregation  AGG([field]) ──────────────────────────────
        m = re.match(r'^\s*(SUM|AVG|COUNT|COUNTD|MAX|MIN)\(\[([^\]]+)\]\)\s*$',
                     formula.strip(), re.IGNORECASE)
        if m:
            tagg  = m.group(1).lower()
            tcol  = m.group(2)
            actual = self._find_col(tcol, df_cols, fc)
            if actual and tagg in PANDAS_AGG:
                return actual, PANDAS_AGG[tagg]

        # ── Pattern 2: LOD expr  { FIXED ... : AGG([field]) } ────────────────────────
        m2 = re.search(r'(SUM|AVG|COUNT|COUNTD|MAX|MIN)\(\[([^\]]+)\]\)',
                       formula, re.IGNORECASE)
        if m2:
            tagg  = m2.group(1).lower()
            tcol  = m2.group(2)
            actual = self._find_col(tcol, df_cols, fc)
            if actual and tagg in PANDAS_AGG:
                return actual, PANDAS_AGG[tagg]

        # ── Pattern 3: Any bracket field reference — only return measure fields ──
        refs = re.findall(r'\[([^\]]+)\]', formula)
        for ref in refs:
            actual = self._find_col(ref, df_cols, fc)
            if actual:
                entry = fc.get(ref.upper().replace(" ", "_"), {})
                if entry.get("role") == "measure":
                    return actual, "sum"
        # Do not return dimension fields — that would pollute groupby keys

        return None, None

    def _extract_shelf_fields(self, node, dest_list: list, field_map: dict):
        """
        Parse a Tableau <rows> or <cols> XML node into a list of field dicts.

        Each entry: {"agg": str, "col": str (normalized), "caption": str, "token": str}

        Handles both simple tokens like [none:FieldName:ok] and federated-prefixed
        tokens like [federated.12345].[sum:Revenue:qk].
        """
        if node is None or not node.text:
            return
        SKIP = {"[Measure Names]", "[Measure Values]", "[Number of Records]",
                "Measure Names", "Measure Values", "Number of Records"}
        # Match federated-prefixed compound tokens first, then simple bracket tokens
        tokens = re.findall(r'\[federated[^\]]*\]\.\[[^\]]+\]|\[[^\]]+\]', node.text)
        for tok in tokens:
            if any(s in tok for s in SKIP):
                continue
            agg, col = self._parse_field(tok)
            if not col:
                continue
            # Prefer worksheet-local caption; fall back to the parsed col name
            caption = field_map.get(tok) or col
            dest_list.append({"agg": agg or "none", "col": col,
                               "caption": caption, "token": tok})

    def _apply_categorical_filters(self, df, filters, field_catalog=None):
        df_cols = df.columns.tolist()
        for f in filters:
            col_name, values, exclude = f["col"], f["values"], f["exclude"]
            col = self._find_col(col_name, df_cols, field_catalog=field_catalog)
            if col:
                df = df[~df[col].isin(values)] if exclude else df[df[col].isin(values)]
            else:
                logging.warning(f"Categorical filter skipped: Column '{col_name}' not found.")
        return df

    def _apply_date_year_filters(self, df, filters, field_catalog=None):
        df_cols = df.columns.tolist()
        for f in filters:
            col_name, years = f["col"], f["years"]
            col = self._find_col(col_name, df_cols, field_catalog=field_catalog)
            if col:
                before = len(df)
                # Extra year safely (handles strings, datetime, and tableauhyperapi.Date)
                def get_year(val):
                    if hasattr(val, 'year'): return int(val.year)
                    try: return int(pd.to_datetime(val, errors='coerce').year)
                    except: return 0
                
                years_series = df[col].apply(get_year)
                df = df[years_series.isin(years)]
            else:
                logging.warning(f"Date filter skipped: Column '{col_name}' not found.")
        return df

    def _read_hyper(self, hyper_path: str) -> pd.DataFrame:
        """Read data from Hyper, prioritizing the largest table if multiple exist."""
        all_dfs = []
        with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
            with Connection(hyper.endpoint, hyper_path) as connection:
                for schema in connection.catalog.get_schema_names():
                    for table in connection.catalog.get_table_names(schema=schema):
                        with connection.execute_query(f"SELECT * FROM {table}") as result:
                            columns = [col.name.unescaped for col in result.schema.columns]
                            data = [list(row) for row in result]
                            if data:
                                all_dfs.append(pd.DataFrame(data, columns=columns))
        
        if not all_dfs:
            return pd.DataFrame()
        
        # Sort by row count descending and return the largest table (usually the main extract)
        all_dfs.sort(key=lambda x: len(x), reverse=True)
        return all_dfs[0]

    def _rename_month_cols(self, pivot, measure):
        pivot.columns = [f"{calendar.month_abbr[m]}_{measure}" if isinstance(m, int) and 1<=m<=12 else f"{m}_{measure}" for m in pivot.columns]
        return pivot
