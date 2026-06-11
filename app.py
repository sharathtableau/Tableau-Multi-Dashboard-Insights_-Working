import sys
sys.dont_write_bytecode = True  # Force Python to always use .py source files, ignore .pyc cache

import os
import logging
import requests
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, send_from_directory
from werkzeug.utils import secure_filename
from datetime import datetime
import json
import shutil
import uuid

from tableau_api import TableauAPI, TableauClient, TableauHyperExtractor
from image_processor import ImageProcessor
import config

# Configure logging
logging.basicConfig(level=logging.DEBUG)

from scheduler_service import report_scheduler
from job_executor import execute_preset_workflow
import threading as _threading
import time as _time

app = Flask(__name__)

# ── Selenium session caches ───────────────────────────────────────────────────
# Pre-warm cache: (token_prefix, wb_idx) → {'extractor', 'ts', 'ready'}
# Created in background threads when a dashboard is selected.
_warm_selenium_cache: dict = {}
_warm_selenium_lock  = _threading.Lock()
_WARM_SESSION_TTL    = 600   # seconds — discard pre-warmed drivers older than this

# Persistent driver cache: token_prefix → SeleniumExtractor (browser stays alive after download)
# Kept alive across multiple Show Data calls until /api/close-selenium-sessions is called
# (triggered by Save Preset).
_persistent_drivers: dict = {}
_persistent_drivers_lock  = _threading.Lock()

# Per-token download lock: prevents two concurrent zone_crosstab calls from
# using the same Selenium driver simultaneously (race condition / dialog interference).
_selenium_download_locks: dict = {}
_selenium_download_locks_lock = _threading.Lock()

def _get_selenium_download_lock(token_prefix: str) -> _threading.Lock:
    """Return a per-token Lock so only one zone_crosstab download runs at a time."""
    with _selenium_download_locks_lock:
        if token_prefix not in _selenium_download_locks:
            _selenium_download_locks[token_prefix] = _threading.Lock()
        return _selenium_download_locks[token_prefix]

# ── Workbook prefetch cache ────────────────────────────────────────────────────
# Key: (token_prefix, workbook_id) → {'wb_path': str, 'datasource_info': list, 'ts': float, 'ready': bool}
# TWB file + datasource info pre-downloaded when the user picks a workbook.
# Consumed by export_dashboard to skip re-downloading during Prepare Image.
_prefetch_wb_cache: dict = {}
_prefetch_wb_lock  = _threading.Lock()
_PREFETCH_WB_TTL   = 600   # seconds — cache valid for 10 min

# ── Dashboard image prefetch cache ────────────────────────────────────────────
# Key: (token_prefix, view_id) → {'pdf_path', 'png_path', 'ts', 'ready': bool}
# PDF export + PNG conversion run in the background when the user selects a dashboard.
# Consumed by export_dashboard to skip the 10–30 s Tableau PDF render on "Prepare Image".
_prefetch_img_cache: dict = {}
_prefetch_img_lock  = _threading.Lock()
_PREFETCH_IMG_TTL   = 300   # seconds — pre-fetched images expire after 5 min

app.secret_key = os.environ.get("SESSION_SECRET", "fallback-secret-key-for-dev")

# ── Server-side sessions ──────────────────────────────────────────────────────
# The app stores large data (CSV text, crop data, workbook paths) in the session.
# Flask's default client-side cookie session has a hard ~4KB browser limit, so the
# cookie silently overflowed and recent writes (e.g. a dashboard's crop_data) were
# dropped — which is why crops intermittently vanished from saved presets.
# Filesystem-backed sessions keep all data on the server; the cookie holds only an id.
from flask_session import Session
app.config['SESSION_TYPE']         = 'filesystem'
app.config['SESSION_FILE_DIR']     = os.path.join('data', 'flask_sessions')
app.config['SESSION_PERMANENT']    = False
app.config['SESSION_USE_SIGNER']   = True
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)
Session(app)

# Configure upload and output folders
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'output'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs('data', exist_ok=True)

PRESETS_FILE = 'data/presets.json'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def index():
    if 'tableau_token' not in session:
        return redirect(url_for('login'))
    
    if 'workbook_count' not in session:
        session['workbook_count'] = 2
        session['workbooks'] = []
        for i in range(2):
            session['workbooks'].append({
                'index': i,
                'project': '',
                'workbook': '',
                'dashboard': '',
                'cropped': False,
                'timestamp': None
            })

    projects = []
    if 'tableau_token' in session:
        try:
            tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
            tableau.token = session['tableau_token']
            tableau.site_id_response = session['tableau_site_id']
            tableau.user_id = session['tableau_user_id']
            projects = tableau.get_projects()
        except Exception as e:
            logging.error(f"Error getting projects for index: {e}")
            if "auth" in str(e).lower() or "401" in str(e):
                return redirect(url_for('logout'))

    return render_template('index.html', 
                          projects=projects,
                          workbooks_by_project={},
                          views_by_workbook={},
                          tableau_server=session.get('tableau_server', ''),
                          tableau_site=session.get('tableau_site', ''))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        site_id = request.form['site_id']
        server_url = request.form.get('server_url', 'https://us-east-1.online.tableau.com')
        
        # Check if using PAT or credentials
        token_name = request.form.get('token_name')
        token_key = request.form.get('token_key')
        username = request.form.get('username')
        password = request.form.get('password')
        
        try:
            tableau = TableauAPI(server_url, site_id)
            
            if token_name and token_key:
                token, site_id_response, user_id = tableau.authenticate_pat(token_name, token_key)
                session['username'] = token_name
                session['auth_method'] = 'pat'
            else:
                token, site_id_response, user_id = tableau.authenticate(username, password)
                session['username'] = username
                session['auth_method'] = 'credentials'
                # Store credentials for Selenium headless extraction
                session['tableau_username'] = username
                session['tableau_password'] = password
            
            # Store authentication info in session
            session['tableau_token'] = token
            session['tableau_site_id'] = site_id_response
            session['tableau_user_id'] = user_id
            session['tableau_server'] = server_url
            session['tableau_site'] = site_id
            
            flash('Successfully logged in to Tableau!', 'success')
            return redirect(url_for('index'))
            
        except Exception as e:
            flash(f'Login failed: {str(e)}', 'error')
            logging.error(f"Login error: {str(e)}")
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully', 'info')
    return redirect(url_for('login'))

@app.route('/extension')
def extension():
    """Serve the Tableau Extension HTML page."""
    return render_template('extension.html')

@app.route('/set_workbook_count', methods=['POST'])
def set_workbook_count():
    count = int(request.form.get('count', 2))
    session['workbook_count'] = count
    session['workbooks'] = []
    session['cropped_images'] = {}
    
    # Initialize workbook data structure
    for i in range(count):
        session['workbooks'].append({
            'index': i,
            'project': '',
            'workbook': '',
            'dashboard': '',
            'cropped': False,
            'timestamp': None
        })
    
    return redirect(url_for('index'))

@app.route('/get_projects')
def get_projects():
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        projects = tableau.get_projects()
        return jsonify({'projects': projects})
    except Exception as e:
        error_msg = str(e).lower()
        if "auth" in error_msg or "login" in error_msg or "credentials" in error_msg or "401" in error_msg:
            return jsonify({'error': 'Tableau session expired. Please log in again.'}), 401
        logging.error(f"Error getting projects: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/workbooks/<project_id>')
def get_workbooks_by_id(project_id):
    if 'tableau_token' not in session:
        logging.warning("Unauthorized access attempt to /api/workbooks")
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        logging.info(f"API CALL: Fetching workbooks for project ID: {project_id}")
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        workbooks = tableau.list_workbooks_in_project_by_id(project_id)
        logging.info(f"API SUCCESS: Found {len(workbooks)} workbooks for project {project_id}")
        return jsonify({'workbooks': workbooks})
    except Exception as e:
        error_msg = str(e).lower()
        if "auth" in error_msg or "login" in error_msg or "credentials" in error_msg or "401" in error_msg:
            logging.warning(f"Authentication error in get_workbooks_by_id: {str(e)}")
            return jsonify({'error': 'Tableau session expired. Please log in again.'}), 401
            
        logging.error(f"API ERROR in get_workbooks_by_id: {str(e)}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/api/datasources/<workbook_id>')
def get_datasources_by_workbook(workbook_id):
    """Get data sources for a workbook with their updated/refresh timestamps"""
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        datasources = tableau.get_workbook_datasources(workbook_id)
        return jsonify({'datasources': datasources})
    except Exception as e:
        logging.error(f"Error getting datasources: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/datasources/<workbook_id>/schema')
def get_datasource_schema(workbook_id):
    """Get upstream data source schema and custom SQL queries for a workbook"""
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        schema_info = tableau.get_workbook_upstream_tables(workbook_id)
        if schema_info:
            return jsonify({'schema_info': schema_info})
        else:
            return jsonify({'schema_info': 'No schema or custom SQL information could be retrieved (Metadata API might be disabled or workbook uses standard extracts).'})
    except Exception as e:
        logging.error(f"Error getting datasource schema: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/test_crosstab')
def show_crosstab_test_page():
    """Show the crosstab testing UI page."""
    if 'tableau_token' not in session:
        return redirect(url_for('login'))
    return render_template('test_crosstab.html')

@app.route('/api/test_crosstab/<workbook_id>')
def test_crosstab_export(workbook_id):
    """Debug endpoint to test crosstab CSV export for all sheets in a workbook."""
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        # Get views first
        views = tableau.get_views_in_workbook(workbook_id)
        views_info = []
        for v in views:
            views_info.append({
                'name': v.get('name', 'Unknown'),
                'id': v.get('id', 'Unknown'),
                'sheetType': v.get('sheetType', 'N/A'),
                'contentUrl': v.get('contentUrl', '')
            })
        
        # Try to export all sheets as CSV
        csv_data = tableau.export_all_sheets_as_csv(workbook_id, max_rows_per_sheet=500)
        
        result = {
            'success': bool(csv_data),
            'views_found': len(views_info),
            'views': views_info,
            'csv_retrieved': bool(csv_data),
            'csv_length': len(csv_data) if csv_data else 0,
            'sheet_count': csv_data.count('=== Sheet:') if csv_data else 0,
            'csv_preview': csv_data[:2000] if csv_data else None,
            'message': 'Crosstab export successful!' if csv_data else 'No CSV data retrieved. Check server logs for details.'
        }
        
        return jsonify(result)
        
    except Exception as e:
        logging.error(f"Error testing crosstab export: {str(e)}", exc_info=True)
        return jsonify({'error': str(e), 'success': False}), 500

@app.route('/api/views/<workbook_id>')
def get_views_by_id(workbook_id):
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        # Try GraphQL discovery first for more accurate counts (including hidden sheets)
        views = tableau.get_workbook_worksheets_graphql(workbook_id)
        if not views:
            views = tableau.get_views_in_workbook(workbook_id)
        
        # Resolve all contentUrls for the workbook views via REST API to have a lookup map
        rest_views = tableau.get_views_in_workbook(workbook_id)
        content_url_map = {v.get('id'): v.get('contentUrl') for v in rest_views if v.get('id')}
        
        # Count different types of views and ensure IDs are resolved
        view_details = []
        for v in views:
            v_id = v.get('id')
            v_name = v.get('name', 'Unknown')
            if not v_id:
                v_id = tableau.resolve_view_id_in_workbook(workbook_id, v_name)
            
            c_url = v.get('contentUrl') or content_url_map.get(v_id, '')
            
            detail = {
                'name': v_name,
                'type': v.get('sheetType', 'N/A'),
                'id': v_id or '',
                'contentUrl': c_url
            }
            
            # Add parent context if available (from GraphQL)
            if v.get('parent_dashboard_id'):
                p_id = v['parent_dashboard_id']
                detail['parent_dashboard_id'] = p_id
                detail['parent_dashboard_name'] = v.get('parent_dashboard_name')
                detail['parent_dashboard_content_url'] = content_url_map.get(p_id, '')
                
            view_details.append(detail)

        worksheets = [v for v in view_details if v.get('type', '').lower() == 'worksheet']
        dashboards_list = [v for v in view_details if v.get('type', '').lower() == 'dashboard']
        other_views = [v for v in view_details if v.get('type', '').lower() not in ['worksheet', 'dashboard']]

        # Expose only dashboards in the dropdown-facing 'views' list
        # (worksheets inside a dashboard are not selectable separately)
        dropdown_views = dashboards_list if dashboards_list else view_details

        # ── Store worksheet order in session so zone-crosstab can do positional
        #    matching when the Excel sheet titles are generic ("Sheet 1", etc.) ──
        worksheet_names = [v['name'].split(' > ')[-1].strip() for v in worksheets]
        if 'workbooks' in session:
            for idx, wb_entry in enumerate(session['workbooks']):
                if wb_entry.get('workbook_id') == workbook_id:
                    session['workbooks'][idx]['worksheet_order'] = worksheet_names
                    session.modified = True
                    break

        result = {
            'views': dropdown_views,
            'total_count': len(view_details),
            'worksheet_count': len(worksheets),
            'dashboard_count': len(dashboards_list),
            'other_count': len(other_views),
            'view_details': view_details,
            'worksheet_order': worksheet_names,
        }

        logging.info(f"Workbook {workbook_id} has {len(views)} total views: {len(worksheets)} worksheets, {len(dashboards_list)} dashboards")

        return jsonify(result)
    except Exception as e:
        logging.error(f"Error getting views: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/filters/<view_id>')
def get_filters(view_id):
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
        
    workbook_id = request.args.get('workbook_id')
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        # Get filters
        filters = tableau.get_view_filters(view_id)
        
        # Also get workbook parameters if available
        parameters = []
        if workbook_id:
            parameters = tableau.get_workbook_parameters(workbook_id)
            
        # Combine them (treating parameters as filters for UI purposes)
        # Use a dict to deduplicate by name
        unique_controls = {}
        for f in filters:
            name = f.get('name')
            if name:
                unique_controls[name] = f
        
        for p in parameters:
            name = p.get('name')
            if name and name not in unique_controls:
                unique_controls[name] = p
        
        combined = list(unique_controls.values())
        logging.info(f"Returning {len(combined)} unique controls (Filters: {len(filters)}, Params: {len(parameters)})")
        
        return jsonify({'filters': combined})
    except Exception as e:
        logging.error(f"Error getting filters: {str(e)}")
        return jsonify({'error': str(e)}), 500

# ── Value-aware filter discovery (fields from .twb, values from extract) ─────
_filter_options_cache: dict = {}
_FILTER_OPTIONS_TTL = 600  # seconds

@app.route('/api/filter-options/<workbook_id>')
def get_filter_options_api(workbook_id):
    """Discover a dashboard's real filters and their possible values so the UI
    can render dropdowns (exact casing) instead of free-text vf_ inputs."""
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401

    dashboard = (request.args.get('dashboard') or '').strip()
    cache_key = (workbook_id, dashboard.lower())
    cached = _filter_options_cache.get(cache_key)
    if cached and (_time.time() - cached['ts']) < _FILTER_OPTIONS_TTL:
        return jsonify({'filters': cached['data'], 'cached': True})

    try:
        from tableau_api import TableauHyperExtractor
        token = session['tableau_token']
        upload_dir = app.config.get('UPLOAD_FOLDER', 'uploads')

        # Reuse the prefetched workbook file when available (saves the download)
        wb_content = None
        pf_key = (token[:12], workbook_id)
        with _prefetch_wb_lock:
            pf = _prefetch_wb_cache.get(pf_key)
            wb_path = pf.get('wb_path') if (pf and pf.get('ready')) else None
        if wb_path and os.path.exists(wb_path):
            with open(wb_path, 'rb') as f:
                wb_content = f.read()

        extractor = TableauHyperExtractor(
            server_url=session['tableau_server'],
            site_id=session['tableau_site_id'],
            token=token,
            output_dir=upload_dir
        )
        if wb_content is None:
            wb_content = extractor.download_workbook(workbook_id)

        options = extractor.get_dashboard_filter_options(wb_content, dashboard)
        _filter_options_cache[cache_key] = {'ts': _time.time(), 'data': options}
        return jsonify({'filters': options})
    except Exception as e:
        logging.error(f"filter-options failed for workbook {workbook_id}: {e}")
        return jsonify({'filters': [], 'error': str(e)})

@app.route('/api/set_selection', methods=['POST'])
def set_selection():
    try:
        data = request.get_json()
        index = data.get('index')

        if 'workbooks' not in session:
            session['workbooks'] = []

        while len(session['workbooks']) <= index:
            session['workbooks'].append({})

        # Only store non-empty values, and keep display names and LUIDs in
        # SEPARATE keys. The name fields ('project'/'workbook'/'dashboard') feed
        # report section titles and preset name-matching — overwriting them with
        # IDs or empty strings silently corrupts presets and reports.
        updates = {}
        for key in ('project', 'workbook', 'dashboard',
                    'project_id', 'workbook_id', 'dashboard_id'):
            val = (data.get(key) or '').strip()
            if val:
                updates[key] = val
        session['workbooks'][index].update(updates)
        session.modified = True
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/prefetch-workbook', methods=['POST'])
def prefetch_workbook():
    """
    Pre-download the Tableau workbook TWB file and datasource info in the background
    as soon as the user selects a workbook — before they click "Prepare Image".

    This eliminates the TWB download step from export_dashboard, shaving several
    seconds off the Prepare Image processing time.

    Expected JSON body:
        workbook_id — Tableau workbook LUID (required)
    """
    if 'tableau_token' not in session:
        return jsonify({'status': 'skipped', 'reason': 'not authenticated'}), 200

    try:
        data        = request.get_json(force=True, silent=True) or {}
        workbook_id = (data.get('workbook_id') or '').strip()

        if not workbook_id:
            return jsonify({'status': 'skipped', 'reason': 'no workbook_id'}), 200

        token      = session.get('tableau_token', '')
        server_url = session.get('tableau_server', '').rstrip('/')
        site_id    = session.get('tableau_site_id', '')
        site_name  = session.get('tableau_site', '')
        token_prefix = token[:12]
        cache_key    = (token_prefix, workbook_id)

        with _prefetch_wb_lock:
            existing = _prefetch_wb_cache.get(cache_key)
            if existing and existing.get('ready') and \
               (_time.time() - existing.get('ts', 0)) < _PREFETCH_WB_TTL:
                logging.info(f"[prefetch-wb] Already cached for {workbook_id}")
                return jsonify({'status': 'cached'}), 200

            # Mark as in-progress so concurrent requests don't double-fetch
            _prefetch_wb_cache[cache_key] = {'ready': False, 'ts': _time.time(),
                                              'wb_path': None, 'datasource_info': []}

        upload_dir = app.config.get('UPLOAD_FOLDER', 'uploads')

        def _fetch_worker():
            try:
                from tableau_api import TableauAPI, TableauHyperExtractor

                # ── Download TWB file ────────────────────────────────────────
                extractor = TableauHyperExtractor(
                    server_url=server_url,
                    site_id=site_id,
                    token=token,
                    output_dir=upload_dir
                )
                wb_content = extractor.download_workbook(workbook_id)
                wb_path    = os.path.join(upload_dir, f"workbook_{workbook_id}.twb")
                with open(wb_path, 'wb') as f:
                    f.write(wb_content)
                logging.info(f"[prefetch-wb] TWB saved: {wb_path}")

                # ── Fetch datasource info ────────────────────────────────────
                tableau = TableauAPI(server_url, site_name)
                tableau.token            = token
                tableau.site_id_response = site_id
                try:
                    datasource_info = tableau.get_workbook_datasources(workbook_id)
                except Exception as ds_err:
                    logging.warning(f"[prefetch-wb] Datasource info failed: {ds_err}")
                    datasource_info = []

                with _prefetch_wb_lock:
                    entry = _prefetch_wb_cache.get(cache_key)
                    if entry is not None:
                        entry['wb_path']        = wb_path
                        entry['datasource_info'] = datasource_info
                        entry['ready']           = True
                        entry['ts']              = _time.time()
                        logging.info(f"[prefetch-wb] Ready for workbook_id={workbook_id}")

            except Exception as exc:
                logging.warning(f"[prefetch-wb] Failed for {workbook_id}: {exc}")
                with _prefetch_wb_lock:
                    _prefetch_wb_cache.pop(cache_key, None)

        t = _threading.Thread(target=_fetch_worker, daemon=True)
        t.start()

        return jsonify({'status': 'prefetching'}), 200

    except Exception as e:
        logging.error(f"prefetch_workbook error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/prefetch-dashboard-image', methods=['POST'])
def prefetch_dashboard_image():
    """
    Pre-render the Tableau dashboard as PDF + convert to PNG in the background
    the moment the user selects a dashboard — before they click "Prepare Image".

    This eliminates the 10–30 s Tableau PDF render wait from the critical path.

    Expected JSON body:
        view_id       — Tableau view LUID for the dashboard (required)
        workbook_index — int, index into session workbooks (for naming)
    """
    if 'tableau_token' not in session:
        return jsonify({'status': 'skipped', 'reason': 'not authenticated'}), 200

    try:
        data        = request.get_json(force=True, silent=True) or {}
        view_id     = (data.get('view_id') or '').strip()
        wb_idx      = int(data.get('workbook_index', 0))

        if not view_id or view_id == 'null':
            return jsonify({'status': 'skipped', 'reason': 'no view_id'}), 200

        token        = session.get('tableau_token', '')
        server_url   = session.get('tableau_server', '').rstrip('/')
        site_name    = session.get('tableau_site', '')
        site_id      = session.get('tableau_site_id', '')
        token_prefix = token[:12]
        cache_key    = (token_prefix, view_id)

        with _prefetch_img_lock:
            existing = _prefetch_img_cache.get(cache_key)
            if existing and existing.get('ready') and \
               (_time.time() - existing.get('ts', 0)) < _PREFETCH_IMG_TTL:
                logging.info(f"[prefetch-img] Already cached for view_id={view_id}")
                return jsonify({'status': 'cached'}), 200
            # Mark in-progress
            _prefetch_img_cache[cache_key] = {'ready': False, 'ts': _time.time(),
                                               'pdf_path': None, 'png_path': None}

        upload_dir = app.config.get('UPLOAD_FOLDER', 'uploads')

        def _img_fetch_worker():
            try:
                from tableau_api import TableauAPI
                from image_processor import ImageProcessor

                tableau = TableauAPI(server_url, site_name)
                tableau.token            = token
                tableau.site_id_response = site_id

                # ── Export PDF (no filters — captures default state) ──────────
                pdf_content = tableau.export_view_as_pdf(view_id, filters={})
                pdf_filename = f"prefetch_{wb_idx}_{view_id[:8]}_{int(_time.time())}.pdf"
                pdf_path     = os.path.join(upload_dir, pdf_filename)
                with open(pdf_path, 'wb') as f:
                    f.write(pdf_content)
                logging.info(f"[prefetch-img] PDF saved: {pdf_path}")

                # ── Convert PDF → PNG ─────────────────────────────────────────
                processor = ImageProcessor()
                png_path  = processor.pdf_to_png(pdf_path)
                logging.info(f"[prefetch-img] PNG ready: {png_path}")

                # ── Trim whitespace ───────────────────────────────────────────
                try:
                    trimmed = processor.trim_to_dashboard_size(png_path)
                    if trimmed != png_path:
                        png_path = trimmed
                except Exception:
                    pass

                with _prefetch_img_lock:
                    entry = _prefetch_img_cache.get(cache_key)
                    if entry is not None:
                        entry['pdf_path'] = pdf_path
                        entry['png_path'] = png_path
                        entry['ready']    = True
                        entry['ts']       = _time.time()
                        logging.info(f"[prefetch-img] Ready for view_id={view_id}")

            except Exception as exc:
                logging.warning(f"[prefetch-img] Failed for view_id={view_id}: {exc}")
                with _prefetch_img_lock:
                    _prefetch_img_cache.pop(cache_key, None)

        t = _threading.Thread(target=_img_fetch_worker, daemon=True)
        t.start()

        return jsonify({'status': 'prefetching'}), 200

    except Exception as e:
        logging.error(f"prefetch_dashboard_image error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/prewarm-selenium', methods=['POST'])
def prewarm_selenium():
    """
    Start a background Selenium session that logs into Tableau Online and
    navigates to the requested dashboard view so it is fully loaded before
    the user clicks "Show Data".

    Expected JSON body:
        workbook_index  — int, index into session['workbooks']
        content_url     — str, Tableau content_url for the selected view
                          (e.g. "WorkbookName/DashboardName")

    Returns immediately with {'status': 'warming'} — the driver is prepared in
    a background thread and cached under a key derived from the session token.
    """
    import re as _re

    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401

    try:
        data          = request.get_json(force=True, silent=True) or {}
        wb_idx        = int(data.get('workbook_index', 0))
        content_url   = (data.get('content_url') or '').strip()

        if not content_url:
            return jsonify({'status': 'skipped', 'reason': 'no content_url'}), 200

        server_url  = session.get('tableau_server', '').rstrip('/')
        site_name   = session.get('tableau_site', '')
        token       = session.get('tableau_token', '')
        username    = session.get('tableau_username', '')
        password    = session.get('tableau_password', '')

        # Fall back to config credentials (needed when logged in via PAT)
        if not username or not password:
            try:
                username = username or getattr(config, 'SELENIUM_USERNAME', '')
                password = password or getattr(config, 'SELENIUM_PASSWORD', '')
            except Exception:
                pass

        # Build view_url the same way zone-crosstab does
        raw_cu     = content_url.strip('/')
        browser_cu = _re.sub(r'/sheets(?=/|$)', '', raw_cu)
        cu_parts   = browser_cu.rsplit('/', 1)
        wb_part    = cu_parts[0] if len(cu_parts) == 2 else browser_cu
        view_part  = cu_parts[1] if len(cu_parts) == 2 else browser_cu
        site_prefix = f"/t/{site_name}" if site_name else ""
        view_url    = f"{server_url}{site_prefix}/views/{wb_part}/{view_part}"

        # Cache key: token prefix (first 12 chars) + workbook index
        token_prefix = token[:12]
        cache_key    = (token_prefix, wb_idx)

        with _warm_selenium_lock:
            # Evict stale / previously-warmed entry for the same slot
            old = _warm_selenium_cache.pop(cache_key, None)
            if old and old.get('extractor'):
                try:
                    old['extractor'].close_warmed_session()
                except Exception:
                    pass

            # Placeholder so zone-crosstab knows warming is in-progress
            _warm_selenium_cache[cache_key] = {'extractor': None, 'ts': _time.time(), 'ready': False}

        # Capture the persistent driver reference now (before the thread starts) so
        # the thread can safely call navigate_and_warm without race conditions.
        with _persistent_drivers_lock:
            persistent_ext = _persistent_drivers.get(token_prefix)

        def _warm_worker():
            try:
                from selenium_extractor import SeleniumExtractor
                dl_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'selenium_downloads')
                os.makedirs(dl_dir, exist_ok=True)

                ok = False
                used_extractor = None

                # ── Fast path: reuse persistent authenticated driver ──────────
                if persistent_ext is not None and persistent_ext._warmed_driver is not None:
                    logging.info(f"[prewarm] Reusing persistent driver for key={cache_key}")
                    ok = persistent_ext.navigate_and_warm(view_url)
                    if ok:
                        used_extractor = persistent_ext

                # ── Slow path: full login + navigate ─────────────────────────
                if not ok:
                    extractor = SeleniumExtractor(
                        username=username,
                        password=password,
                        download_dir=dl_dir,
                        token=token,
                        server_url=server_url,
                        site_name=site_name,
                    )
                    ok = extractor.warm_session(view_url)
                    if ok:
                        used_extractor = extractor

                with _warm_selenium_lock:
                    entry = _warm_selenium_cache.get(cache_key)
                    if entry is not None:
                        if ok and used_extractor:
                            entry['extractor'] = used_extractor
                            entry['ready']     = True
                            logging.info(f"[prewarm] Session ready for key={cache_key}")
                        else:
                            _warm_selenium_cache.pop(cache_key, None)
                            logging.warning(f"[prewarm] warm failed for key={cache_key}")
            except Exception as exc:
                logging.warning(f"[prewarm] Background warm failed: {exc}")
                with _warm_selenium_lock:
                    _warm_selenium_cache.pop(cache_key, None)

        t = _threading.Thread(target=_warm_worker, daemon=True)
        t.start()

        return jsonify({'status': 'warming', 'cache_key': f"{token[:6]}_{wb_idx}"}), 200

    except Exception as e:
        logging.error(f"prewarm_selenium error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/close-selenium-sessions', methods=['POST'])
def close_selenium_sessions():
    """
    Quit all Selenium browser sessions associated with the current user session.
    Called automatically after Save Preset so Chrome processes don't accumulate.
    """
    if 'tableau_token' not in session:
        return jsonify({'status': 'ok', 'closed': 0}), 200

    token_prefix = session.get('tableau_token', '')[:12]
    closed = 0

    # Close persistent driver
    with _persistent_drivers_lock:
        ext = _persistent_drivers.pop(token_prefix, None)
    if ext:
        try:
            ext.close_warmed_session()
            closed += 1
            logging.info(f"[close-selenium] Closed persistent driver for {token_prefix}")
        except Exception as e:
            logging.warning(f"[close-selenium] Error closing persistent driver: {e}")

    # Close any pending pre-warm sessions for this user
    with _warm_selenium_lock:
        stale_keys = [k for k in _warm_selenium_cache if k[0] == token_prefix]
        for k in stale_keys:
            entry = _warm_selenium_cache.pop(k, None)
            if entry and entry.get('extractor'):
                try:
                    entry['extractor'].close_warmed_session()
                    closed += 1
                except Exception:
                    pass

    logging.info(f"[close-selenium] Closed {closed} session(s) for token_prefix={token_prefix}")
    return jsonify({'status': 'ok', 'closed': closed}), 200


@app.route('/export_dashboard', methods=['POST'])
def export_dashboard():
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form
            
        view_id = data.get('view_id')
        view_type = data.get('view_type', '').lower()
        workbook_id = data.get('workbook_id')
        workbook_index = int(data.get('workbook_index', data.get('index', 0)))
        project_name = data.get('project_name', 'Unknown')
        workbook_name = data.get('workbook_name', 'Unknown')
        dashboard_name = data.get('dashboard_name', 'Unknown')
        filters = data.get('filters', {})
        
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']

        logging.info(f"DEBUG: export_dashboard - workbook_id: {workbook_id}, server: {session.get('tableau_server')}, token exists: {bool(tableau.token)}")

        # Determine the best view ID for the PDF image
        # If user selected a worksheet, we try to use its parent dashboard if possible for the image
        pdf_view_id = view_id
        if view_type != 'dashboard' and workbook_id:
            try:
                # Get all views to find a dashboard alternative
                workbook_views = tableau.get_workbook_worksheets_graphql(workbook_id)
                if not workbook_views:
                    workbook_views = tableau.get_views_in_workbook(workbook_id)
                
                dashboards = [v for v in workbook_views if v.get('sheetType', '').lower() == 'dashboard']
                if dashboards:
                    # Preference: a dashboard named exactly like the prefix in the worksheet name
                    # (e.g., if worksheet is "Dashboard > ProductView", try to find a dashboard named "Dashboard")
                    target_dashboard = None
                    if ' > ' in dashboard_name:
                        prefix = dashboard_name.split(' > ')[0]
                        target_dashboard = next((d for d in dashboards if d.get('name', '').lower() == prefix.lower()), None)
                    
                    if not target_dashboard:
                        target_dashboard = dashboards[0] # Fallback to first dashboard
                    
                    if target_dashboard:
                        resolved_id = target_dashboard.get('id')
                        if not resolved_id:
                            resolved_id = tableau.resolve_view_id_in_workbook(workbook_id, target_dashboard.get('name'))
                        
                        if resolved_id:
                            logging.info(f"Upgrading PDF export from worksheet '{dashboard_name}' (ID: {view_id}) to dashboard '{target_dashboard.get('name')}' (ID: {resolved_id})")
                            pdf_view_id = resolved_id
            except Exception as e:
                logging.warning(f"Failed to find dashboard alternative for PDF export: {e}")

        if isinstance(filters, str):
            try:
                filters = json.loads(filters)
            except:
                filters = {}

        logging.info(f"Exporting image from view {pdf_view_id} (Workbook Index: {workbook_index}) with filters: {filters}")

        processor   = ImageProcessor()
        token_pfx   = session.get('tableau_token', '')[:12]
        img_key     = (token_pfx, pdf_view_id)
        has_filters = bool(filters and any(v for v in filters.values() if v))

        pdf_path = None
        png_path = None

        # ── Check pre-fetched image cache (only when no custom filters applied) ─
        if not has_filters:
            with _prefetch_img_lock:
                pf_img = _prefetch_img_cache.get(img_key)
                if pf_img and pf_img.get('ready') and \
                   (_time.time() - pf_img.get('ts', 0)) < _PREFETCH_IMG_TTL and \
                   pf_img.get('png_path') and os.path.exists(pf_img['png_path']):
                    pdf_path = pf_img.get('pdf_path', '')
                    png_path = pf_img['png_path']
                    logging.info(f"[export_dashboard] ✓ Using pre-fetched PNG: {png_path}")
                elif pf_img and not pf_img.get('ready'):
                    # Still rendering — wait up to 25 s (covers most Tableau render times)
                    logging.info("[export_dashboard] Prefetch image in progress; waiting up to 25 s…")
                    wait_img = _time.time() + 25
                    while _time.time() < wait_img:
                        _time.sleep(0.5)
                        with _prefetch_img_lock:
                            pf2 = _prefetch_img_cache.get(img_key)
                        if pf2 and pf2.get('ready') and \
                           pf2.get('png_path') and os.path.exists(pf2['png_path']):
                            pdf_path = pf2.get('pdf_path', '')
                            png_path = pf2['png_path']
                            logging.info(f"[export_dashboard] ✓ Prefetch became ready: {png_path}")
                            break

        # ── Fallback: export PDF synchronously (filters present, or no prefetch hit) ─
        if not png_path:
            logging.info(f"[export_dashboard] No prefetch hit — exporting PDF now (has_filters={has_filters})")
            pdf_content  = tableau.export_view_as_pdf(pdf_view_id, filters=filters)
            pdf_filename = f"dashboard_{workbook_index}_{datetime.now().timestamp()}.pdf"
            pdf_path     = os.path.join(app.config['UPLOAD_FOLDER'], pdf_filename)
            with open(pdf_path, 'wb') as f:
                f.write(pdf_content)
            # Convert PDF to PNG
            png_path = processor.pdf_to_png(pdf_path)
            # Trim whitespace
            try:
                trimmed = processor.trim_to_dashboard_size(png_path)
                if trimmed != png_path:
                    png_path = trimmed
            except Exception as trim_err:
                logging.warning(f"PNG trim step failed (non-fatal): {trim_err}")
        
        # Update session data
        if 'workbooks' not in session:
            session['workbooks'] = []
        
        while len(session['workbooks']) <= workbook_index:
            session['workbooks'].append({})
        
        # Get data source info + workbook XML for this workbook
        workbook_id     = data.get('workbook_id')
        content_url     = data.get('content_url', '')
        datasource_info = []
        token           = session.get('tableau_token', '')
        token_prefix    = token[:12] if token else ''
        wb_cache_key    = (token_prefix, workbook_id) if workbook_id else None

        # ── Check prefetch cache (populated when workbook was selected) ────────
        prefetch_hit = False
        if wb_cache_key:
            with _prefetch_wb_lock:
                pf = _prefetch_wb_cache.get(wb_cache_key)
                if pf and pf.get('ready') and \
                   (_time.time() - pf.get('ts', 0)) < _PREFETCH_WB_TTL:
                    prefetch_hit    = True
                    datasource_info = pf.get('datasource_info') or []
                    prefetch_wb_path = pf.get('wb_path') or ''
                    logging.info(f"[export_dashboard] ✓ Using pre-fetched TWB: {prefetch_wb_path}")
                elif pf and not pf.get('ready'):
                    # Still downloading — wait up to 8 s
                    logging.info("[export_dashboard] Prefetch in progress; waiting up to 8 s…")
                    wait_dl = _time.time() + 8
                    while _time.time() < wait_dl:
                        _time.sleep(0.5)
                        pf2 = _prefetch_wb_cache.get(wb_cache_key)
                        if pf2 and pf2.get('ready'):
                            prefetch_hit    = True
                            datasource_info = pf2.get('datasource_info') or []
                            prefetch_wb_path = pf2.get('wb_path') or ''
                            logging.info(f"[export_dashboard] ✓ Prefetch became ready: {prefetch_wb_path}")
                            break

        # ── Fallback: fetch datasource info synchronously if not pre-fetched ──
        if not prefetch_hit and workbook_id:
            try:
                datasource_info = tableau.get_workbook_datasources(workbook_id)
            except Exception as ds_err:
                logging.warning(f"Could not fetch datasource info: {ds_err}")

        # ── Download / use the pre-fetched workbook XML (for zone mapping) ─────
        try:
            if workbook_id and token:
                if prefetch_hit and prefetch_wb_path and os.path.exists(prefetch_wb_path):
                    # Use the already-downloaded file — no network call needed
                    session['workbooks'][workbook_index]['wb_path'] = prefetch_wb_path
                    logging.info(f"[export_dashboard] Workbook XML reused from prefetch cache.")
                else:
                    # Download now (prefetch wasn't ready in time)
                    logging.info(f"[export_dashboard] Downloading TWB now (no prefetch hit).")
                    extractor = TableauHyperExtractor(
                        server_url=session.get('tableau_server', '').rstrip('/'),
                        site_id=session.get('tableau_site_id', ''),
                        token=token,
                        output_dir=app.config['UPLOAD_FOLDER']
                    )
                    wb_content = extractor.download_workbook(workbook_id)
                    wb_path    = os.path.join(app.config['UPLOAD_FOLDER'], f"workbook_{workbook_id}.twb")
                    with open(wb_path, 'wb') as f:
                        f.write(wb_content)
                    session['workbooks'][workbook_index]['wb_path'] = wb_path
                    logging.info(f"Workbook XML saved for zone mapping: {wb_path}")

        except Exception as wb_err:
            logging.warning(f"Workbook XML step failed (non-fatal): {wb_err}")

        csv_data = ''  # CSV is fetched on-demand, not during export

        # (no CSV file to save)
        
        session['workbooks'][workbook_index]['pdf_path']          = pdf_path
        session['workbooks'][workbook_index]['png_path']          = png_path
        session['workbooks'][workbook_index]['original_png_path'] = png_path  # never overwritten by crops
        session['workbooks'][workbook_index]['timestamp']       = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        session['workbooks'][workbook_index]['project']         = project_name
        session['workbooks'][workbook_index]['workbook']        = workbook_name
        session['workbooks'][workbook_index]['workbook_id']     = workbook_id
        session['workbooks'][workbook_index]['dashboard']       = dashboard_name
        session['workbooks'][workbook_index]['dashboard_name']  = dashboard_name
        session['workbooks'][workbook_index]['dashboard_view_id'] = pdf_view_id
        session['workbooks'][workbook_index]['content_url']     = content_url
        session['workbooks'][workbook_index]['applied_filters'] = filters
        session['workbooks'][workbook_index]['datasources']     = datasource_info
        session.modified = True

        return jsonify({
            'success':      True,
            'png_filename': os.path.basename(png_path),
            'csv_fetched':  False,   # CSV fetched on-demand after crop
            'csv_rows':     0,
            'csv_preview':  '',
        })
        
    except Exception as e:
        logging.error(f"Error exporting dashboard: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/data_insights', methods=['POST'])
def generate_data_insights():
    """Generate AI insights from CSV data using Gemini."""
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    csv_data = request.json.get('csv_data', '')
    png_filename = request.json.get('png_filename', '')
    # Prefer the cropped image so AI focuses on the specific selected area;
    # fall back to the full PNG if no crop has been made yet.
    cropped_filename = request.json.get('cropped_filename', '')
    workbook_index = request.json.get('workbook_index', None)

    if not csv_data and not png_filename and not cropped_filename:
        return jsonify({'error': 'No data (CSV or Image) provided for insights'}), 400

    try:
        processor = ImageProcessor()
        if not processor.gemini_client:
            return jsonify({'error': 'AI insights not available. Check Gemini API key configuration.'}), 500

        # Resolve the best image: cropped > full PNG
        image_path = None
        if cropped_filename:
            candidate = os.path.join(app.config['UPLOAD_FOLDER'], cropped_filename)
            if os.path.exists(candidate):
                image_path = candidate

        if not image_path and png_filename:
            candidate = os.path.join(app.config['UPLOAD_FOLDER'], png_filename)
            if os.path.exists(candidate):
                image_path = candidate

        # Pull the sheet-filtered csv_data from session if not supplied by caller
        if not csv_data and workbook_index is not None:
            try:
                wb = session.get('workbooks', [])[int(workbook_index)]
                csv_path = wb.get('csv_data_path', '')
                if csv_path and os.path.exists(csv_path):
                    with open(csv_path, 'r', encoding='utf-8') as _f:
                        csv_data = _f.read()
                else:
                    csv_data = wb.get('csv_data', '')
            except Exception:
                pass

        # If we have an image, use the more powerful _generate_ai_insights that combines both
        if image_path:
            logging.info(f"Generating combined Image+CSV insights (image={os.path.basename(image_path)}, csv={len(csv_data)} chars)")
            insights_list = processor._generate_ai_insights(image_path, "Dashboard", csv_data=csv_data)
            insights_text = "\n".join(insights_list)
            return jsonify({
                'success': True,
                'insights': insights_text
            })

        # Fallback to CSV only if no image
        from config import GEMINI_MODEL
        prompt = f"""You are a senior business intelligence analyst. Analyze the following data tables and provide actionable executive insights.

### DATA
{csv_data[:15000]}

### INSTRUCTIONS
1. Identifying trends and anomalies.
2. Cross-referencing tables.
3. Keeping it high-level for CEOs.

### RESPONSE FORMAT
- One bold Strategic Title.
- Three bullet points with data-backed insights.
"""
        response = processor.gemini_client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[prompt]
        )
        return jsonify({
            'success': True,
            'insights': response.text
        })
        
    except Exception as e:
        logging.error(f"Error generating data insights: {str(e)}")
        return jsonify({'error': f'Failed to generate insights: {str(e)}'}), 500

@app.route('/crop/<int:workbook_index>')
def crop_image(workbook_index):
    if 'tableau_token' not in session:
        return redirect(url_for('login'))
    
    if 'workbooks' not in session or workbook_index >= len(session['workbooks']):
        flash('Invalid workbook index', 'error')
        return redirect(url_for('index'))
    
    workbook = session['workbooks'][workbook_index]
    if 'png_path' not in workbook:
        flash('No image to crop for this workbook', 'error')
        return redirect(url_for('index'))
    
    png_filename = os.path.basename(workbook['png_path'])
    return render_template('crop.html', 
                         workbook_index=workbook_index, 
                         png_filename=png_filename,
                         workbook=workbook)

@app.route('/crop_dashboard', methods=['POST'])
def crop_dashboard():
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    try:
        data = request.get_json()
        workbook_index = data['workbook_index']
        crop_data = data['crop_data']
        
        workbook = session['workbooks'][workbook_index]

        # Resolve the original PNG path — always crop from the ORIGINAL dashboard image,
        # never from a previously cropped version.
        # Priority: original_png_path (set at export) → png_path → filename in request body
        original_path = (workbook.get('original_png_path') or workbook.get('png_path', ''))
        if not original_path or not os.path.exists(original_path):
            req_filename = data.get('filename', '')
            if req_filename:
                candidate = os.path.join(app.config['UPLOAD_FOLDER'], os.path.basename(req_filename))
                if os.path.exists(candidate):
                    original_path = candidate
                    session['workbooks'][workbook_index]['original_png_path'] = candidate
                    session['workbooks'][workbook_index]['png_path']          = candidate
                    session.modified = True
                    logging.info(f"crop_dashboard: original_png_path restored from request filename: {candidate}")

        if not original_path or not os.path.exists(original_path):
            return jsonify({'error': 'No image available for this workbook. Please re-export the dashboard first.'}), 400
        
        # Process the cropped image
        processor = ImageProcessor()
        cropped_path = processor.crop_image(original_path, crop_data)
        
        # Create thumbnail for preview
        thumbnail_path = processor.create_thumbnail(cropped_path)
        
        # Update session
        session['workbooks'][workbook_index]['cropped_path'] = cropped_path
        session['workbooks'][workbook_index]['thumbnail_path'] = thumbnail_path
        session['workbooks'][workbook_index]['cropped'] = True
        session['workbooks'][workbook_index]['crop_data'] = crop_data
        
        # --- NEW: Geometry Mapping to find the specific sheet ---
        new_csv_data = None
        csv_row_count = 0
        mapping_error = None
        matched_sheets = []
        try:
            from PIL import Image as PILImage
            with PILImage.open(original_path) as img:
                orig_w, orig_h = img.size
            
            wb_path = workbook.get('wb_path')
            dashboard_name = workbook.get('dashboard_name') or workbook.get('dashboard')

            if not wb_path:
                logging.warning(f"[zone-map] wb_path not in session for workbook {workbook_index} — workbook XML was never downloaded")
                mapping_error = "Workbook XML not downloaded — re-export this dashboard to enable zone mapping"
            elif not os.path.exists(wb_path):
                logging.warning(f"[zone-map] wb_path file missing: {wb_path}")
                mapping_error = f"Workbook XML file missing ({os.path.basename(wb_path)}) — re-export to refresh"
            elif not dashboard_name:
                logging.warning(f"[zone-map] dashboard_name not set for workbook {workbook_index}")
                mapping_error = "Dashboard name not set in session"

            if wb_path and os.path.exists(wb_path) and dashboard_name:
                with open(wb_path, "rb") as f:
                    wb_content = f.read()
                
                server_url = session.get('tableau_server', '').rstrip('/')
                extractor = TableauHyperExtractor(
                    server_url=server_url,
                    site_id=session.get('tableau_site_id', ''),
                    token=session.get('tableau_token', ''),
                    output_dir=app.config['UPLOAD_FOLDER']
                )
                
                # Identify specific sheet using geometry mapping
                # We pass the raw pixel crop_data and image_size; the extractor handles XML normalization
                # Include the PNG path so the extractor can pixel-scan for the exact y_offset
                crop_data_with_path = dict(crop_data)
                crop_data_with_path['_png_path'] = workbook.get('png_path', '')
                # Pull any filters that were active when this dashboard was exported
                saved_filters = workbook.get('applied_filters') or {}
                logging.info(f"Crop route: passing applied_filters to extract_and_parse: {saved_filters}")
                # Identify specific sheet(s) using geometry mapping
                _, matched_sheets, _ = extractor.extract_and_parse(
                    wb_content, dashboard_name, crop_data=crop_data_with_path, image_size=(orig_w, orig_h),
                    applied_filters=saved_filters
                )

                # Fallback: if zone mapping found nothing (e.g. single-view workbook or
                # crop didn't overlap any zone), re-run without crop_data to get all sheets.
                if not matched_sheets:
                    logging.info("Zone mapping returned no matches — falling back to all sheets in dashboard")
                    try:
                        _, matched_sheets, _ = extractor.extract_and_parse(
                            wb_content, dashboard_name, applied_filters=saved_filters
                        )
                    except Exception as _fb_err:
                        logging.warning(f"Fallback extract_and_parse failed: {_fb_err}")

                if matched_sheets:
                    logging.info(f"✓ Geometry Mapping matched crop to sheets: {matched_sheets}")
                    
                    full_csv_path = workbook.get('csv_data_path')
                    if full_csv_path and os.path.exists(full_csv_path):
                        with open(full_csv_path, 'r', encoding='utf-8') as f:
                            full_csv = f.read()
                        
                        lines = full_csv.splitlines()
                        filtered_lines = []
                        
                        # We collect data for ANY of the matched sheets
                        for sheet_name in matched_sheets:
                            in_sheet = False
                            target_norm = sheet_name.lower().replace(" ", "").replace("_", "")
                            
                            sheet_found = False
                            for line in lines:
                                if line.startswith("=== Sheet:"):
                                    line_norm = line.lower().replace(" ", "").replace("_", "")
                                    if target_norm in line_norm:
                                        in_sheet = True
                                        sheet_found = True
                                        filtered_lines.append(line)
                                    else:
                                        in_sheet = False
                                elif in_sheet:
                                    filtered_lines.append(line)
                            
                            if sheet_found:
                                logging.info(f"✓ Added data for sheet '{sheet_name}' from captured CSV.")
                        
                        if filtered_lines:
                            new_csv_data = "\n".join(filtered_lines)
                            session['workbooks'][workbook_index]['csv_data'] = new_csv_data
                            
                            # Count rows (excluding headers and metadata indicators)
                            csv_row_count = 0
                            for line in filtered_lines:
                                line = line.strip()
                                if not line or line.startswith('=== Sheet:') or line.startswith('...'):
                                    continue
                                csv_row_count += 1
                            # We subtract 1 for EACH sheet's CSV header
                            csv_row_count -= len(matched_sheets) 
                            if csv_row_count < 0: csv_row_count = 0
                
        except Exception as ge_err:
            import traceback
            err_msg = f"Geometry mapping failed: {ge_err}\n{traceback.format_exc()}"
            logging.error(err_msg)
            mapping_error = str(ge_err)
        
        # --- END Mapping Logic ---

        # Persist the matched sheet(s) in session so they survive page refresh
        if matched_sheets:
            session['workbooks'][workbook_index]['matched_sheet']  = matched_sheets[0]
            session['workbooks'][workbook_index]['matched_sheets'] = matched_sheets
            session.modified = True

        session.modified = True

        # Pull view/content info so the frontend can call /api/zone-crosstab
        _wb_sess = session['workbooks'][workbook_index]
        _dash_vid = _wb_sess.get('dashboard_view_id', '')
        _content_url = _wb_sess.get('content_url', '')
        _dashboard_name = _wb_sess.get('dashboard', '')

        _png_path = _wb_sess.get('png_path', '')
        return jsonify({
            'success': True,
            'cropped_filename': os.path.basename(cropped_path),
            'png_filename': os.path.basename(_png_path) if _png_path else '',
            'thumbnail_filename': os.path.basename(thumbnail_path),
            'csv_fetched': bool(new_csv_data),
            'csv_rows': csv_row_count,
            'csv_preview': new_csv_data[:50000] if new_csv_data else "",
            'matched_sheet': ", ".join(matched_sheets) if matched_sheets else None,
            'matched_sheets': matched_sheets,
            'mapping_error': mapping_error if not new_csv_data else None,
            'dashboard_view_id': _dash_vid,
            'content_url': _content_url,
            'dashboard': _dashboard_name,
        })
        
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        logging.error(f"Error saving crop: {str(e)}\n{error_detail}")
        return jsonify({
            'success': False,
            'error': str(e),
            'detail': error_detail
        }), 500

# Alias so the standalone crop.html page (cropper.js calls /save_crop) works too
@app.route('/save_crop', methods=['POST'])
def save_crop():
    return crop_dashboard()

@app.route('/combine', methods=['POST'])
def combine_images():
    print(f"\n>>> /combine CALLED — THIS IS THE REAL ENTRY POINT <<<", flush=True)
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    if 'workbooks' not in session:
        return jsonify({'error': 'No workbooks selected'}), 400

    # Export every dashboard that has a saved crop on disk. Stale placeholder
    # slots (left behind by changed selections) must not block the export.
    ready = [wb for wb in session['workbooks']
             if wb.get('cropped') and wb.get('cropped_path')
             and os.path.exists(wb.get('cropped_path', ''))]
    if not ready:
        return jsonify({'error': 'No cropped dashboards available — export and crop '
                                 'at least one dashboard first (crops are cleared after a reset).'}), 400
    
    try:
        # Get data from JSON request
        data = request.get_json()
        output_format = data.get('format', 'pptx')
        custom_filename = data.get('filename', 'dashboard_report')
        
        # Remove extension from filename if provided
        base_filename = custom_filename
        if custom_filename.lower().endswith(('.pdf', '.docx', '.pptx')):
            base_filename = os.path.splitext(custom_filename)[0]
        
        # Use default filename if empty
        if not base_filename.strip():
            base_filename = 'tableau_report'
        
        processor = ImageProcessor()

        # Create temporary output directory
        temp_dir = os.path.join(app.config['OUTPUT_FOLDER'], 'temp')
        os.makedirs(temp_dir, exist_ok=True)

        # Build image paths and metadata together from the SAME filtered list
        # so report sections always pair with the right dashboard.
        cropped_paths = [wb['cropped_path'] for wb in ready]
        summary_data = []
        for i, wb in enumerate(ready):
            summary_data.append({
                'section': i + 1,
                'project': wb.get('project') or 'Unknown',
                'workbook': wb.get('workbook') or 'Unknown',
                'dashboard': wb.get('dashboard') or f'Dashboard {i + 1}',
                'timestamp': wb.get('timestamp', 'Unknown'),
                'image_path': wb.get('cropped_path', ''),
                'applied_filters': wb.get('applied_filters', {}),
                'datasources': wb.get('datasources', []),
                'csv_data': wb.get('csv_data', '')
            })
        
        # Combine images
        if output_format == 'pdf':
            output_path = processor.combine_to_pdf_with_details(cropped_paths, temp_dir, base_filename, summary_data)
        elif output_format == 'pptx':
            output_path = processor.combine_to_pptx_with_details(cropped_paths, temp_dir, base_filename, summary_data)
        else:
            output_path = processor.combine_to_word_with_details(cropped_paths, temp_dir, base_filename, summary_data)
        
        # Return file for download. Only the generated report file is cleaned up —
        # the session's cropped/PNG images must survive so the user can export
        # again (e.g. a second format) or save a preset. /reset removes them.
        def cleanup_after_download():
            import threading
            import time
            def delayed_cleanup():
                time.sleep(120)
                try:
                    if os.path.exists(output_path):
                        os.remove(output_path)
                except:
                    pass
            threading.Thread(target=delayed_cleanup).start()

        cleanup_after_download()
        
        return send_file(output_path, as_attachment=True, download_name=custom_filename)
        
    except Exception as e:
        logging.error(f"Error combining images: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download')
def download_result():
    if 'last_output' not in session:
        flash('No file to download', 'error')
        return redirect(url_for('index'))
    
    output_path = session['last_output']
    if not os.path.exists(output_path):
        flash('Output file not found', 'error')
        return redirect(url_for('index'))
    
    return send_file(output_path, as_attachment=True)

@app.route('/reset')
def reset():
    # Clean up any uploaded files
    if 'workbooks' in session:
        for wb in session['workbooks']:
            for path_key in ['pdf_path', 'png_path', 'cropped_path']:
                if path_key in wb and os.path.exists(wb[path_key]):
                    try:
                        os.remove(wb[path_key])
                    except:
                        pass
    
    # Clear session data except authentication
    keys_to_keep = ['tableau_token', 'tableau_site_id', 'tableau_user_id', 
                   'tableau_server', 'tableau_site', 'username']
    session_copy = {k: v for k, v in session.items() if k in keys_to_keep}
    session.clear()
    session.update(session_copy)
    
    flash('Reset complete', 'info')
    return redirect(url_for('index'))

# --- Job & Preset Management ---

def load_presets():
    if not os.path.exists(PRESETS_FILE):
        return []
    try:
        with open(PRESETS_FILE, 'r') as f:
            return json.load(f)
    except:
        return []

def save_presets(presets):
    with open(PRESETS_FILE, 'w') as f:
        json.dump(presets, f, indent=4)

@app.route('/jobs')
def jobs():
    if 'tableau_token' not in session:
        return redirect(url_for('login'))
    return render_template('jobs.html', presets=load_presets())

@app.route('/api/presets', methods=['GET'])
def get_presets():
    return jsonify(load_presets())

@app.route('/api/save_preset', methods=['POST'])
def save_preset():
    if 'workbooks' not in session:
        return jsonify({'error': 'No active session data'}), 400
    
    data = request.json
    preset_name = data.get('name', f"Preset {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    # Capture relevant session data for the preset
    preset_data = {
        'id': str(uuid.uuid4()),
        'name': preset_name,
        'created_at': datetime.now().isoformat(),
        'server_url': session.get('tableau_server'),
        'site_id': session.get('tableau_site'),
        'output_format': data.get('format', 'pptx'),
        'workbooks': []
    }
    
    # Create presets directory if it doesn't exist
    presets_dir = os.path.join(app.static_folder, 'presets')
    os.makedirs(presets_dir, exist_ok=True)
    
    uncropped = []  # dashboards saved without a crop — will use the full image
    slide_no = 0
    for i, wb in enumerate(session['workbooks']):
        if wb.get('project') and wb.get('workbook') and wb.get('dashboard'):
            slide_no += 1
            workbook_entry = {
                'project': wb['project'],
                'workbook': wb['workbook'],
                'dashboard': wb['dashboard'],
                'workbook_id': wb.get('workbook_id', ''),
                'dashboard_id': wb.get('dashboard_view_id') or wb.get('dashboard_id', ''),
                'crop_data': wb.get('crop_data'),
                'filters': wb.get('applied_filters', {})
            }

            # Flag any dashboard that has no crop — these render the full dashboard
            if not wb.get('crop_data'):
                uncropped.append(f"#{slide_no} ({wb.get('dashboard') or 'Dashboard'})")

            # Save preview image if available
            if wb.get('cropped_path') and os.path.exists(wb['cropped_path']):
                preview_filename = f"{preset_data['id']}_{i}.png"
                preview_path = os.path.join(presets_dir, preview_filename)
                try:
                    shutil.copy2(wb['cropped_path'], preview_path)
                    workbook_entry['preview_image'] = f"presets/{preview_filename}"
                except Exception as e:
                    logging.error(f"Failed to save preset preview: {e}")

            preset_data['workbooks'].append(workbook_entry)

    presets = load_presets()
    presets.append(preset_data)
    save_presets(presets)

    resp = {'success': True, 'preset': preset_data}
    if uncropped:
        resp['warning'] = (
            "Saved, but these dashboards have NO crop and will use the FULL dashboard image: "
            + ", ".join(uncropped)
            + ". Crop each dashboard (you should see a cropped thumbnail) and save again if you want the cropped area."
        )
        logging.warning(f"Preset '{preset_name}' saved with uncropped dashboards: {uncropped}")
    return jsonify(resp)

@app.route('/api/delete_preset/<preset_id>', methods=['DELETE'])
def delete_preset(preset_id):
    presets = load_presets()
    presets = [p for p in presets if p['id'] != preset_id]
    save_presets(presets)
    return jsonify({'success': True})

@app.route('/api/run_job/<preset_id>', methods=['POST'])
def run_job(preset_id):
    import sys
    print(f"\n>>> RUN_JOB CALLED for preset_id={preset_id} <<<", flush=True)
    sys.stdout.flush()
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    presets = load_presets()
    preset = next((p for p in presets if p['id'] == preset_id), None)
    
    if not preset:
        return jsonify({'error': 'Preset not found'}), 404
    
    try:
        tableau = TableauAPI(session['tableau_server'], session['tableau_site'])
        tableau.token = session['tableau_token']
        tableau.site_id_response = session['tableau_site_id']
        tableau.user_id = session['tableau_user_id']
        
        for i, wb_config in enumerate(preset['workbooks']):
            # Use the shared executor for consistent logic
            report_path = execute_preset_workflow(
                preset, 
                server_url=session.get('tableau_server'),
                site_id=session.get('tableau_site'),
                token_name=session.get('username') if session.get('auth_method') == 'pat' else None,
                token_key=None # We don't have the key in session most likely, 
                              # execute_preset_workflow will fallback to config or fail if needed.
                              # Actually, in a session context, we might rely on the existing token.
            )
            
            filename = os.path.basename(report_path)
            return jsonify({
                'success': True, 
                'download_url': url_for('download_result_file', filename=filename)
            })
        
    except Exception as e:
        logging.error(f"Job execution failed: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/download_job/<filename>')
def download_result_file(filename):
    path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "File not found", 404

@app.route('/api/update_job_settings/<preset_id>', methods=['POST'])
def update_job_settings(preset_id):
    data = request.json
    presets = load_presets()
    
    for preset in presets:
        if preset['id'] == preset_id:
            if 'schedule' in data:
                preset['schedule'] = data['schedule']
            if 'recipients' in data:
                # Expecting comma separated string from UI
                preset['recipients'] = [r.strip() for r in data['recipients'].split(',') if r.strip()]
            if 'is_enabled' in data:
                preset['is_enabled'] = data['is_enabled']
            if 'message' in data:
                preset['message'] = data['message']
            if 'repeat_type' in data:
                preset['repeat_type'] = data['repeat_type']
            if 'interval' in data:
                preset['interval'] = data['interval']
            if 'days' in data:
                preset['days'] = data['days']
            break
            
    save_presets(presets)
    # Reload schedules
    report_scheduler.load_and_schedule_jobs()
    
    return jsonify({'success': True})

# --- End Job Management ---

@app.route('/image/<filename>')
def serve_image(filename):
    """Serve uploaded images"""
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))

@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# ─────────────────────────────────────────────────────────────────────────────
# Debug Tool routes
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/debug')
def debug_tool():
    """Serve the standalone Zone Debug HTML tool."""
    debug_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'debug_zone_tool_v1.html')
    return send_file(debug_path, mimetype='text/html')


@app.route('/api/debug/sheet-data', methods=['POST'])
def debug_sheet_data():
    """
    Debug endpoint: extract and reconstruct crosstab data for a specific
    worksheet from an uploaded Tableau .twb / .twbx workbook.

    Form fields:
        workbook      — the uploaded workbook file (required)
        sheet         — worksheet name to reconstruct (required)
        dashboard     — dashboard name hint for context (optional)
        max_rows      — preview row cap, default 500 (optional)
    """
    import pandas as pd
    import io as _io

    try:
        if 'workbook' not in request.files:
            return jsonify({'error': 'No workbook file in request (field: workbook)'}), 400

        wb_file      = request.files['workbook']
        sheet_name   = (request.form.get('sheet', '') or '').strip()
        dash_name    = (request.form.get('dashboard', '') or '').strip()
        max_rows     = int(request.form.get('max_rows', 500) or 500)

        if not sheet_name:
            return jsonify({'error': 'sheet parameter is required'}), 400

        wb_content = wb_file.read()
        if not wb_content:
            return jsonify({'error': 'Uploaded workbook is empty'}), 400

        # TableauHyperExtractor works entirely from local bytes — no server auth needed
        import tempfile, os as _os
        tmp_dir = tempfile.mkdtemp(prefix='dbg_hyper_')
        try:
            extractor = TableauHyperExtractor('', '', '', output_dir=tmp_dir)

            df_master, dashboard_sheets, worksheet_defs = extractor.extract_and_parse(
                wb_content,
                dash_name or sheet_name   # pass sheet_name as fallback so logs are readable
            )
        finally:
            # tmp_dir is cleaned up by extractor internally; remove dir if still present
            try:
                import shutil as _shutil
                _shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass

        if sheet_name not in worksheet_defs:
            available = sorted(worksheet_defs.keys())
            return jsonify({
                'error': f"Sheet '{sheet_name}' not found in workbook.",
                'available_sheets': available
            }), 404

        # ── Build filter summary for the response ──────────────────────────────────
        defn = worksheet_defs[sheet_name]
        filters_summary = {
            'categorical': [
                {'col': f['col'], 'values': f['values'][:10],
                 'exclude': f.get('exclude', False)}
                for f in defn.get('categorical_filters', [])
            ],
            'date':     defn.get('date_filters', []),
            'top':      defn.get('top_filters', []),
            'relative': defn.get('relative_date_filters', []),
        }

        # ── Detect layout-only worksheets (table-calc patterns) ────────────────────
        LAYOUT_TOKENS = {'MIN(0)', 'MAX(0)', 'INDEX', 'INDEX()', 'RANK', 'RANK()',
                         'AGG(MIN(0))', 'AGG(MAX(0))'}
        layout_warning = None
        all_shelf_captions = (
            [f.get('caption', '').upper() for f in defn.get('rows_fields', [])] +
            [f.get('caption', '').upper() for f in defn.get('cols_fields', [])]
        )
        if any(t in all_shelf_captions for t in LAYOUT_TOKENS) or (
            not defn.get('rows_fields') and not defn.get('cols_fields')
        ):
            layout_warning = (
                "This worksheet uses Tableau table calculations "
                "(e.g. INDEX, RANK, MIN(0)) on its shelves. "
                "The reconstructed data shows raw aggregated values "
                "and will NOT match Tableau's calculated layout exactly."
            )

        if df_master is None or df_master.empty:
            return jsonify({
                'error': 'No Hyper extract found in this workbook. '
                         'Live-connection workbooks cannot be queried offline.'
            }), 400

        # Reconstruct for just this one sheet
        csv_str = extractor.reconstruct_csv(df_master, [sheet_name], worksheet_defs)
        if not csv_str or not csv_str.strip():
            return jsonify({'error': 'Reconstruction returned no data for this sheet.'}), 400

        # Parse the CSV section for the sheet (skip the "=== Sheet ===" header lines)
        lines = csv_str.splitlines()
        data_lines = []
        in_sheet = False
        for line in lines:
            if f'=== {sheet_name} ===' in line or f"Sheet: {sheet_name}" in line:
                in_sheet = True
                continue
            if in_sheet:
                if line.startswith('===') and sheet_name not in line:
                    break  # next sheet section starts
                data_lines.append(line)
        csv_block = '\n'.join(data_lines).strip() if data_lines else csv_str.strip()

        try:
            df_sheet = pd.read_csv(_io.StringIO(csv_block))
        except Exception:
            # Fallback: try the whole csv_str
            df_sheet = pd.read_csv(_io.StringIO(csv_str))

        # JSON-safe: replace NaN/NaT with None
        df_sheet = df_sheet.where(pd.notnull(df_sheet), None)

        columns   = list(df_sheet.columns)
        total     = len(df_sheet)
        preview   = df_sheet.head(max_rows)
        rows_data = preview.values.tolist()

        return jsonify({
            'success':         True,
            'sheet_name':      sheet_name,
            'dashboard':       dash_name,
            'columns':         columns,
            'rows':            rows_data,
            'total_rows':      total,
            'preview_rows':    len(rows_data),
            'filters_applied': filters_summary,
            'layout_warning':  layout_warning,
        })

    except Exception as exc:
        logging.error(f"debug_sheet_data error: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/debug-session')
def debug_session():
    """Returns session state for debugging — login status, workbook data, and candidate URLs."""
    import urllib.parse as _up
    logged_in   = 'tableau_token' in session
    server_url  = session.get('tableau_server', '')
    site_name   = session.get('tableau_site', '')
    workbooks   = session.get('workbooks', [])
    out = {
        'logged_in':   logged_in,
        'server_url':  server_url,
        'site_name':   site_name,
        'workbook_count': len(workbooks),
        'workbooks': [],
    }
    import re as _re2
    site_prefix = f"/t/{site_name}" if site_name else ""
    for i, wb in enumerate(workbooks):
        content_url   = wb.get('content_url', '')
        dashboard     = wb.get('dashboard', '')
        dashboard_vid = wb.get('dashboard_view_id', '')
        matched_sheet = wb.get('matched_sheet', '')
        raw_cu        = content_url.strip('/')
        browser_cu    = _re2.sub(r'/sheets(?=/|$)', '', raw_cu)
        cu_parts      = browser_cu.rsplit('/', 1)
        workbook_part = cu_parts[0] if len(cu_parts) == 2 else browser_cu
        view_part     = cu_parts[1] if len(cu_parts) == 2 else browser_cu
        enc_sheet     = _up.quote(matched_sheet) if matched_sheet else '<sheet>'
        enc_dash      = _up.quote(dashboard) if dashboard else _up.quote(view_part)
        candidates = []
        if workbook_part and enc_dash:
            candidates.append(f"{server_url}{site_prefix}/views/{workbook_part}/{enc_dash}/crosstabs/sheets/{enc_sheet}")
        if browser_cu and browser_cu != f"{workbook_part}/{view_part}":
            candidates.append(f"{server_url}{site_prefix}/views/{browser_cu}/crosstabs/sheets/{enc_sheet}")
        out['workbooks'].append({
            'index': i,
            'dashboard': dashboard,
            'dashboard_view_id': dashboard_vid,
            'content_url': content_url,
            'matched_sheet': matched_sheet,
            'has_png': bool(wb.get('png_path')),
            'has_original_png': bool(wb.get('original_png_path')),
            'candidate_urls': candidates,
        })
    return jsonify(out)


@app.route('/api/zone-crosstab', methods=['POST'])
def zone_crosstab():
    """
    Download the crosstab data for a specific embedded worksheet using zone mapping.

    Strategy (in order):
      1. Selenium (headless Chrome) — injects REST token as workgroup_session_id cookie,
         clicks Download → Crosstab → selects the matched sheet → downloads Excel.
         Falls back to username/password login if cookie auth is rejected.
      2. REST API  — GET /views/{view_id}/crosstab/excel
      3. Internal URL — direct HTTP request with workgroup_session_id cookie

    JSON body:
        workbook_index  — index into session['workbooks'] (required)
        sheet_name      — matched worksheet name from zone mapping (required)
        max_rows        — preview row cap, default 500 (optional)
    """
    import re as _re
    import io as _io
    import csv as _csv
    import pandas as pd
    import openpyxl
    import urllib.parse as _urlparse

    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated', 'auth_required': True}), 401

    try:
        body       = request.get_json(force=True, silent=True) or {}
        wb_idx     = int(body.get('workbook_index', 0))
        sheet_name = (body.get('sheet_name') or '').strip()
        max_rows   = int(body.get('max_rows', 500) or 500)

        if not sheet_name:
            return jsonify({'error': 'sheet_name is required'}), 400

        if 'workbooks' not in session or wb_idx >= len(session['workbooks']):
            return jsonify({'error': f'Workbook index {wb_idx} not found in session'}), 404

        wb            = session['workbooks'][wb_idx]
        dashboard_vid = wb.get('dashboard_view_id', '')
        content_url   = wb.get('content_url', '')
        dashboard     = wb.get('dashboard', '')
        server_url    = session.get('tableau_server', '').rstrip('/')
        site_name     = session.get('tableau_site', '')
        site_id_res   = session.get('tableau_site_id', '')
        token         = session.get('tableau_token', '')
        username      = session.get('tableau_username', '')
        password      = session.get('tableau_password', '')

        # ── Fall back to config credentials for Selenium (PAT login stores no password) ──
        if not username or not password:
            try:
                username = username or getattr(config, 'SELENIUM_USERNAME', '')
                password = password or getattr(config, 'SELENIUM_PASSWORD', '')
                if username and password:
                    logging.info("[zone-crosstab] Using config.SELENIUM_USERNAME/PASSWORD for Selenium.")
            except Exception:
                pass

        # Build browser-friendly content_url:
        # Tableau Online REST API returns "WorkbookName/sheets/ViewName";
        # the browser URL uses "WorkbookName/ViewName" — strip /sheets/.
        raw_cu     = content_url.strip('/')
        browser_cu = _re.sub(r'/sheets(?=/|$)', '', raw_cu)
        cu_parts   = browser_cu.rsplit('/', 1)
        wb_part    = cu_parts[0] if len(cu_parts) == 2 else browser_cu
        view_part  = cu_parts[1] if len(cu_parts) == 2 else browser_cu

        site_prefix = f"/t/{site_name}" if site_name else ""
        view_url    = f"{server_url}{site_prefix}/views/{wb_part}/{view_part}"

        tableau = TableauAPI(server_url, site_name)
        tableau.token            = token
        tableau.site_id_response = site_id_res

        def _norm(s): return ''.join(s.split()).lower()

        df_result   = None
        method_used = None
        tried_urls  = []

        # ── Helper: find the right Excel sheet for sheet_name ─────────────────
        def _find_excel_sheet(wb_xl, target_sheet_name):
            """
            Tableau REST API names Excel worksheets 'Sheet 1', 'Sheet 2', … not
            by the actual Tableau sheet names.  Use three strategies to find the
            correct worksheet:

            1. Exact / partial TITLE match (works when Tableau uses real names).
            2. CONTENT match — search the first 8 rows of every worksheet for
               cell values that contain the key words of target_sheet_name.
               e.g. for "Region" we look for a sheet whose cells include "region".
            3. ALPHABETICAL POSITIONAL match using all worksheet names stored in
               session (if available) — sorts them and uses the index.
            4. First worksheet fallback.
            """
            worksheets = wb_xl.worksheets
            if not worksheets:
                return None, 'empty workbook'

            def _sn(s): return ''.join(str(s).lower().split())
            tgt_norm  = _sn(target_sheet_name)
            # words > 2 chars for content search
            tgt_words = [w.lower() for w in target_sheet_name.split() if len(w) > 2]

            # ── 1. Title match ──────────────────────────────────────────────
            for ws in worksheets:
                t = _sn(ws.title)
                if t == tgt_norm or tgt_norm in t:
                    logging.info(f"[zone-crosstab] Sheet match by title: '{ws.title}'")
                    return ws, f'title match ({ws.title})'

            # ── 2. Content match ────────────────────────────────────────────
            # For each worksheet scan first 8 rows; score by how many
            # target words appear in cell values.
            best_ws    = None
            best_score = 0
            sheet_scores = []
            for ws in worksheets:
                score = 0
                try:
                    rows = list(ws.iter_rows(min_row=1, max_row=8, values_only=True))
                    all_text = ' '.join(
                        str(v).lower() for row in rows for v in row if v is not None
                    )
                    for word in tgt_words:
                        if word in all_text:
                            score += 1
                except Exception:
                    pass
                sheet_scores.append((ws, score))
                if score > best_score:
                    best_score = score
                    best_ws    = ws

            logging.info(f"[zone-crosstab] Content scores for '{target_sheet_name}': "
                         + ', '.join(f"'{ws.title}'={sc}" for ws, sc in sheet_scores))

            if best_score > 0:
                logging.info(f"[zone-crosstab] Sheet match by content: '{best_ws.title}'")
                return best_ws, f'content match ({best_ws.title})'

            # ── 3. Alphabetical positional match ───────────────────────────
            known_order = wb.get('worksheet_order') or []
            if known_order:
                sorted_names = sorted(known_order, key=str.lower)
                try:
                    idx = next(i for i, s in enumerate(sorted_names) if _sn(s) == tgt_norm)
                    if idx < len(worksheets):
                        ws = worksheets[idx]
                        logging.info(f"[zone-crosstab] Sheet match by position {idx}: '{ws.title}'")
                        return ws, f'positional match ({ws.title})'
                except StopIteration:
                    pass

            # ── 4. Fallback ─────────────────────────────────────────────────
            logging.warning(f"[zone-crosstab] No match for '{target_sheet_name}'; "
                            f"using first sheet '{worksheets[0].title}'")
            return worksheets[0], f'fallback ({worksheets[0].title})'

        # ── REST API strategy is intentionally disabled ───────────────────────
        # Tableau REST API always returns the dashboard's DEFAULT sheet regardless
        # of which sheet was requested — only Selenium can select the correct sheet
        # from the Download Crosstab dialog.
        logging.info(f"[zone-crosstab] Skipping REST API; using Selenium for sheet='{sheet_name}'")

        # ── Strategy: Selenium headless browser (only strategy — REST API disabled) ──
        token_prefix = token[:12]
        # Acquire per-token download lock: prevents two concurrent requests from
        # fighting over the same headless browser session (dialog/navigation race).
        _dl_lock = _get_selenium_download_lock(token_prefix)
        _dl_lock.acquire()
        try:
            from selenium_extractor import SeleniumExtractor

            dl_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'selenium_downloads')
            os.makedirs(dl_dir, exist_ok=True)

            # ── 1. Check pre-warm cache for this workbook slot ────────────────
            cache_key      = (token_prefix, wb_idx)
            warm_extractor = None

            with _warm_selenium_lock:
                entry = _warm_selenium_cache.get(cache_key)
                if entry:
                    age = _time.time() - entry.get('ts', 0)
                    if entry.get('ready') and entry.get('extractor') and age < _WARM_SESSION_TTL:
                        warm_extractor = entry['extractor']
                        _warm_selenium_cache.pop(cache_key)   # consume
                        logging.info(f"[zone-crosstab] Reusing pre-warmed session (age {age:.0f}s)")
                    elif not entry.get('ready') and age < _WARM_SESSION_TTL:
                        logging.info("[zone-crosstab] Pre-warm in progress; polling up to 40 s…")

            # ── 2. If not ready yet, wait for the warming thread ──────────────
            if warm_extractor is None:
                wait_deadline = _time.time() + 40
                while _time.time() < wait_deadline:
                    with _warm_selenium_lock:
                        entry = _warm_selenium_cache.get(cache_key)
                        if entry and entry.get('ready') and entry.get('extractor'):
                            warm_extractor = entry['extractor']
                            _warm_selenium_cache.pop(cache_key)
                            logging.info("[zone-crosstab] Pre-warm became ready; reusing.")
                            break
                        elif not entry:
                            break   # warming failed or never started
                    _time.sleep(1)

            df_sel = None
            used_extractor = None  # track which extractor performed the download

            if warm_extractor is not None:
                # ── Fast path: browser already on the dashboard ───────────────
                logging.info(f"[zone-crosstab] Selenium (warmed keep-alive): sheet={sheet_name!r}")
                df_sel = warm_extractor.download_keep_alive(sheet_name)
                used_extractor = warm_extractor
            else:
                # ── Slow path: full login + navigate + download ───────────────
                # Check if there is a persistent (previously-used) driver for reuse
                with _persistent_drivers_lock:
                    persist_ext = _persistent_drivers.get(token_prefix)

                if persist_ext is not None and persist_ext._warmed_driver is not None:
                    logging.info(f"[zone-crosstab] Selenium (persistent navigate): view_url={view_url!r}")
                    ok = persist_ext.navigate_and_warm(view_url)
                    if ok:
                        df_sel = persist_ext.download_keep_alive(sheet_name)
                        used_extractor = persist_ext
                    else:
                        logging.warning("[zone-crosstab] Persistent driver navigate failed; creating fresh driver.")
                        with _persistent_drivers_lock:
                            _persistent_drivers.pop(token_prefix, None)

                if df_sel is None:
                    fresh_ext = SeleniumExtractor(
                        username=username,
                        password=password,
                        download_dir=dl_dir,
                        token=token,
                        server_url=server_url,
                        site_name=site_name,
                    )
                    logging.info(f"[zone-crosstab] Selenium (fresh): view_url={view_url!r}, sheet={sheet_name!r}")
                    # warm first so we can use download_keep_alive (keeps driver alive)
                    ok = fresh_ext.warm_session(view_url)
                    if ok:
                        df_sel = fresh_ext.download_keep_alive(sheet_name)
                        used_extractor = fresh_ext
                    else:
                        # Last resort: single-shot (driver quits after)
                        df_sel = fresh_ext.download_single_sheet(view_url, sheet_name)

            # ── Store the extractor in persistent cache (driver stays alive) ──
            if used_extractor is not None and used_extractor._warmed_driver is not None:
                with _persistent_drivers_lock:
                    old_ext = _persistent_drivers.get(token_prefix)
                    if old_ext and old_ext is not used_extractor:
                        try:
                            old_ext.close_warmed_session()
                        except Exception:
                            pass
                    _persistent_drivers[token_prefix] = used_extractor
                logging.info(f"[zone-crosstab] Browser kept alive in persistent cache for token {token_prefix}")

            if df_sel is not None and len(df_sel) > 0:
                df_result   = df_sel
                method_used = f"Selenium headless browser (sheet: {sheet_name})"
                logging.info(f"[zone-crosstab] ✓ Selenium: {len(df_result)} rows")
            else:
                logging.warning("[zone-crosstab] Selenium returned empty result; trying internal URL.")
        except ImportError:
            logging.warning("[zone-crosstab] selenium_extractor not importable; skipping.")
        except Exception as sel_err:
            logging.warning(f"[zone-crosstab] Selenium failed: {sel_err}")
        finally:
            _dl_lock.release()

        # ── Strategy 3: Internal browser URL (direct HTTP + cookie) ─────────
        if df_result is None:
            try:
                enc_sheet  = _urlparse.quote(sheet_name)
                enc_dash   = _urlparse.quote(dashboard) if dashboard else _urlparse.quote(view_part)
                domain     = server_url.replace('https://', '').replace('http://', '').split('/')[0]
                base_url   = f"{server_url}{site_prefix}/views/{wb_part}/{enc_dash}/crosstabs/sheets/{enc_sheet}"
                logging.info(f"[zone-crosstab] Internal URL: {base_url}")

                for fmt in ['xlsx', 'csv']:
                    trial_url = f"{base_url}?:format={fmt}&:embed=yes&:showVizHome=no"
                    sess_r = requests.Session()
                    sess_r.cookies.set('workgroup_session_id', token, domain=domain, path='/')
                    sess_r.cookies.set('XSRF-TOKEN', token, domain=domain, path='/')
                    r2 = sess_r.get(trial_url, timeout=60,
                                    headers={'X-Tableau-Auth': token, 'X-XSRF-TOKEN': token,
                                             'Referer': server_url,
                                             'User-Agent': 'Mozilla/5.0 (compatible; TableauCropper)'},
                                    allow_redirects=True)
                    tried_urls.append({'url': trial_url, 'status': r2.status_code,
                                       'ct': r2.headers.get('content-type', '')})
                    logging.info(f"[zone-crosstab] {fmt} → {r2.status_code}, size={len(r2.content)}")

                    if r2.status_code == 200 and len(r2.content) > 50:
                        target = _norm(sheet_name)
                        if fmt == 'csv':
                            try:
                                df_result   = pd.read_csv(_io.StringIO(r2.text))
                                method_used = "Internal URL (CSV)"
                                break
                            except Exception:
                                pass
                        else:
                            try:
                                wb_xl2 = openpyxl.load_workbook(_io.BytesIO(r2.content), read_only=True)
                                ws2 = next(
                                    (ws for ws in wb_xl2.worksheets
                                     if _norm(ws.title) == target or target in _norm(ws.title)),
                                    wb_xl2.worksheets[0] if wb_xl2.worksheets else None
                                )
                                if ws2:
                                    out2 = _io.StringIO()
                                    _csv.writer(out2).writerows(
                                        r for r in ws2.iter_rows(values_only=True)
                                        if any(c is not None for c in r)
                                    )
                                    wb_xl2.close()
                                    csv2 = out2.getvalue().strip()
                                    if csv2:
                                        df_result   = pd.read_csv(_io.StringIO(csv2))
                                        method_used = f"Internal URL (XLSX: {ws2.title})"
                                        break
                            except Exception as xlsx_err:
                                logging.warning(f"[zone-crosstab] XLSX parse error: {xlsx_err}")
            except Exception as e:
                logging.warning(f"[zone-crosstab] Internal URL failed: {e}")

        # ── All strategies exhausted ───────────────────────────────────────────
        if df_result is None:
            return jsonify({
                'error': (
                    f"Could not download crosstab for '{sheet_name}'. "
                    "Selenium needs username+password credentials — "
                    "ensure you logged in with email/password (not PAT only)."
                ),
                'dashboard_view_id': dashboard_vid,
                'content_url': content_url,
                'view_url': view_url,
                'dashboard': dashboard,
                'tried_urls': tried_urls,
            }), 502

        df_result = df_result.where(pd.notnull(df_result), None)
        total     = len(df_result)
        preview   = df_result.head(max_rows)
        cols      = list(df_result.columns)

        # Use pandas to_json then re-parse to guarantee Python-native types
        # (no numpy.int64/float64). Flask's jsonify can't serialize numpy types
        # and an unhandled TypeError produces a garbled/empty response that
        # Safari throws as "The string did not match the expected pattern".
        import json as _json
        rows_data = _json.loads(preview.to_json(orient='records', default_handler=str))

        payload = {
            'success':      True,
            'source':       'tableau_server',
            'method':       method_used,
            'sheet_name':   sheet_name,
            'dashboard':    dashboard,
            'columns':      cols,
            'rows':         rows_data,
            'total_rows':   int(total),
            'preview_rows': len(rows_data),
        }
        resp = app.response_class(
            response=_json.dumps(payload, ensure_ascii=False, allow_nan=False, default=str),
            status=200,
            mimetype='application/json',
        )
        return resp

    except Exception as exc:
        logging.error(f"zone_crosstab error: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/save-zone-csv', methods=['POST'])
def save_zone_csv():
    """
    Persist the zone crosstab result (columns + rows from the Show Data panel) into
    the session so that PPT/Word/PDF export and AI insights can use it.

    Body JSON:
      workbook_index  — int, which workbook slot to update
      columns         — list of column names (from zone-crosstab response)
      rows            — list of row dicts  (from zone-crosstab response)
      sheet_name      — the matched sheet name (for the CSV header)
    """
    if 'tableau_token' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    try:
        body         = request.get_json(force=True, silent=True) or {}
        wb_idx       = int(body.get('workbook_index', 0))
        columns      = body.get('columns', [])
        rows         = body.get('rows', [])
        sheet_name   = body.get('sheet_name', 'Sheet')

        if 'workbooks' not in session or wb_idx >= len(session['workbooks']):
            return jsonify({'error': 'Workbook index out of range'}), 404

        if not columns:
            return jsonify({'ok': False, 'reason': 'no columns'}), 200

        # Build a simple CSV string: header + data rows
        import csv as _csv, io as _io
        buf = _io.StringIO()
        writer = _csv.DictWriter(buf, fieldnames=columns, extrasaction='ignore')
        buf.write(f'=== Sheet: {sheet_name} ===\n')
        writer.writeheader()
        writer.writerows(rows)
        csv_text = buf.getvalue()

        # Store in session — this replaces/overrides the old Embedding-API CSV so
        # AI insights and export always see the fresh live Selenium-downloaded data.
        session['workbooks'][wb_idx]['csv_data']         = csv_text
        session['workbooks'][wb_idx]['zone_csv_sheet']   = sheet_name
        session.modified = True

        logging.info(f"[save-zone-csv] Saved {len(rows)} rows for workbook {wb_idx} / sheet '{sheet_name}'")
        return jsonify({'ok': True, 'rows': len(rows), 'sheet': sheet_name})

    except Exception as exc:
        logging.error(f"save_zone_csv error: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/debug/view-csv', methods=['POST'])
def debug_view_csv():
    """
    Download the underlying CSV data for a Tableau worksheet directly from Tableau Server.

    Uses the authenticated Flask session (tableau_server, tableau_token, tableau_site_id).

    Two-step lookup:
      1. Search for the workbook by name  →  GET /workbooks?filter=name:eq:{name}
      2. List views in that workbook      →  GET /workbooks/{id}/views
      3. Fuzzy-match the sheet_name to a published view
      4. If no direct match, return available_views so the user can pick

    Note: Worksheets embedded inside a Dashboard are NOT separate published views on
    Tableau Server.  Only the Dashboard itself (and any standalone sheets) appear as views.
    If the matched sheet is an embedded worksheet, the caller should pick the Dashboard view.

    JSON body:
        workbook_name  — workbook name as shown on Tableau Server (required)
        sheet_name     — worksheet / view name from zone mapping (required)
        view_id        — optional: skip lookup and use this view ID directly
        max_rows       — row cap for preview, default 500 (optional)
    """
    import pandas as pd
    import io as _io
    import urllib.parse as _urlparse

    try:
        # ── Auth check ────────────────────────────────────────────────────────
        if 'tableau_token' not in session:
            return jsonify({
                'error': 'Not authenticated. Please log in to Tableau first at /',
                'auth_required': True,
            }), 401

        server_url  = session.get('tableau_server', '').rstrip('/')
        site_id_res = session.get('tableau_site_id', '')
        token       = session.get('tableau_token', '')

        if not server_url or not token:
            return jsonify({'error': 'Tableau session missing server or token.'}), 401

        data          = request.get_json(force=True, silent=True) or {}
        workbook_name = (data.get('workbook_name') or '').strip()
        sheet_name    = (data.get('sheet_name')    or '').strip()
        direct_vid    = (data.get('view_id')       or '').strip()
        max_rows      = int(data.get('max_rows', 500) or 500)

        if not workbook_name or not sheet_name:
            return jsonify({'error': 'workbook_name and sheet_name are required.'}), 400

        # ── Build TableauAPI using existing session ───────────────────────────
        tableau = TableauAPI(server_url, session.get('tableau_site', ''))
        tableau.token            = token
        tableau.site_id_response = site_id_res
        tableau.user_id          = session.get('tableau_user_id', '')
        hdrs = {"X-Tableau-Auth": token, "Accept": "application/json"}
        api  = f"{server_url}/api/{tableau.api_version}/sites/{site_id_res}"

        def _norm(s): return ''.join(s.split()).lower()

        # ── If caller already knows the view ID, skip lookup ─────────────────
        all_views   = []
        workbook_id = None

        if not direct_vid:
            # Step 1: Find the workbook by name
            enc_wb  = _urlparse.quote(workbook_name)
            wb_url  = f"{api}/workbooks?filter=name:eq:{enc_wb}&pageSize=100"
            wb_res  = requests.get(wb_url, headers=hdrs, timeout=30)

            if wb_res.status_code == 401:
                return jsonify({
                    'error': 'Tableau token expired. Please log out and log in again.',
                    'auth_required': True,
                }), 401

            if wb_res.status_code != 200:
                return jsonify({
                    'error': (f"Tableau Server returned {wb_res.status_code} while searching "
                              f"for workbook '{workbook_name}'. "
                              "Check that the name matches exactly as shown on Tableau Server."),
                    'detail': wb_res.text[:400],
                }), 502

            workbooks = wb_res.json().get('workbooks', {}).get('workbook', [])
            if isinstance(workbooks, dict):
                workbooks = [workbooks]

            if not workbooks:
                return jsonify({
                    'error': (f"Workbook '{workbook_name}' not found on Tableau Server. "
                              "The name must match exactly (case-sensitive) as shown in the server UI."),
                    'hint': "Try browsing to the workbook in Tableau Server and copying its exact name.",
                }), 404

            workbook_id = workbooks[0].get('id', '')
            logging.info(f"[debug/view-csv] Workbook '{workbook_name}' → ID {workbook_id}")

            # Step 2: List all views in that workbook
            views_url = f"{api}/workbooks/{workbook_id}/views?includeUsageStatistics=false"
            vw_res    = requests.get(views_url, headers=hdrs, timeout=30)

            if vw_res.status_code == 200:
                all_views = vw_res.json().get('views', {}).get('view', [])
                if isinstance(all_views, dict):
                    all_views = [all_views]
            else:
                logging.warning(f"[debug/view-csv] Could not list views: {vw_res.status_code}")

            logging.info(
                f"[debug/view-csv] {len(all_views)} published views in '{workbook_name}': "
                f"{[v.get('name') for v in all_views]}")

        # ── Step 3: Match the sheet name to a published view ─────────────────
        target = _norm(sheet_name)
        matched_view = None

        if not direct_vid:
            # Exact match
            matched_view = next(
                (v for v in all_views if v.get('name', '').lower() == sheet_name.lower()), None)
            # Normalised (strip spaces)
            if not matched_view:
                matched_view = next(
                    (v for v in all_views if _norm(v.get('name', '')) == target), None)
            # Substring
            if not matched_view:
                matched_view = next(
                    (v for v in all_views
                     if _norm(v.get('name', '')) in target or target in _norm(v.get('name', ''))),
                    None)

        if not direct_vid and not matched_view:
            # Sheet is likely a worksheet embedded inside a dashboard — not a standalone view.
            available = [{'name': v.get('name',''), 'id': v.get('id','')} for v in all_views]
            return jsonify({
                'error': (
                    f"'{sheet_name}' is not a published standalone view in '{workbook_name}'. "
                    "It is probably a worksheet inside a Dashboard. "
                    "Pick a view from available_views below to download its data instead."
                ),
                'embedded_sheet': sheet_name,
                'available_views': available,
                'workbook_id': workbook_id,
            }), 404

        view_id   = direct_vid or matched_view.get('id', '')
        view_name = (matched_view.get('name', sheet_name) if matched_view else sheet_name)
        logging.info(f"[debug/view-csv] Downloading data for view '{view_name}' (ID: {view_id})")

        # ── Step 4: Download view data as CSV ─────────────────────────────────
        csv_text = tableau._try_csv_endpoint(view_id, {})
        if not csv_text:
            csv_text = tableau._try_crosstab_endpoint(view_id, {})

        if not csv_text:
            return jsonify({
                'error': (f"Tableau returned no data for view '{view_name}'. "
                          "The view may require browser-level rendering or contain no data."),
                'view_id': view_id,
            }), 502

        # ── Step 5: Parse CSV → JSON ──────────────────────────────────────────
        clean_csv = '\n'.join(
            l for l in csv_text.splitlines() if not l.startswith('... (truncated'))
        try:
            df = pd.read_csv(_io.StringIO(clean_csv))
        except Exception as parse_err:
            return jsonify({
                'error': f"Could not parse Tableau CSV: {parse_err}",
                'raw_preview': csv_text[:500],
            }), 500

        df        = df.where(pd.notnull(df), None)
        total     = len(df)
        rows_data = df.head(max_rows).values.tolist()

        return jsonify({
            'success':       True,
            'source':        'tableau_server',
            'sheet_name':    view_name,
            'workbook_name': workbook_name,
            'view_id':       view_id,
            'columns':       list(df.columns),
            'rows':          rows_data,
            'total_rows':    total,
            'preview_rows':  len(rows_data),
        })

    except Exception as exc:
        logging.error(f"debug_view_csv error: {exc}", exc_info=True)
        return jsonify({'error': str(exc)}), 500


if __name__ == '__main__':
    # Initial load of schedules
    report_scheduler.load_and_schedule_jobs()
    app.run(host='0.0.0.0', port=5002, debug=True)
